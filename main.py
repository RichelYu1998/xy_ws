# -*- coding: utf-8 -*-
# 标准库
import argparse
import asyncio
import base64
import ctypes
import glob
import gzip
import importlib.metadata as im
import io
import ipaddress
import json
import logging
import os
import platform
import random
import secrets
import re
import select
import shlex
import shutil
import smtplib
import socket
import ssl
import subprocess
import sys
import threading
import time
import struct
import hashlib
import urllib.error
import urllib.parse
import urllib.request
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from functools import wraps
from html import escape
from pathlib import Path
try:
    import psutil
except ImportError:
    psutil = None

try:
    from prometheus_client import Counter, Histogram, Gauge
except ImportError:
    Counter = Histogram = Gauge = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import pymysql
except ImportError:
    pymysql = None

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    from playwright.async_api import async_playwright
except ImportError:
    async_playwright = None
try:
    from fastapi import FastAPI, HTTPException, Query, Header, Depends, File, UploadFile, Form, BackgroundTasks
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import JSONResponse, FileResponse, StreamingResponse, HTMLResponse
    import uvicorn
except ImportError:
    FastAPI = HTTPException = Query = Header = Depends = File = UploadFile = Form = BackgroundTasks = None
    CORSMiddleware = None
    JSONResponse = FileResponse = StreamingResponse = None
try:
    from starlette.middleware.gzip import GZipMiddleware
except ImportError:
    GZipMiddleware = None
    uvicorn = None

try:
    from starlette.requests import Request
    from starlette.responses import Response
except ImportError:
    Request = Response = None
from typing import List, Dict, Optional, Any, Callable, TypeVar, Union, Tuple
try:
    from pydantic import BaseModel, Field, ValidationError, field_validator
except ImportError:
    try:
        from pydantic import BaseModel, Field, ValidationError
        try:
            from pydantic import field_validator
        except ImportError:
            from pydantic import validator as field_validator
    except ImportError:
        BaseModel = Field = ValidationError = field_validator = None

try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
except ImportError:
    Fernet = hashes = PBKDF2HMAC = None

try:
    from packaging import version as packaging_version
except ImportError:
    packaging_version = None

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
_module_logger = logging.getLogger('main')
logger = logging.getLogger('FileCleaner')

def get_version_info():
    """获取详细的 Python 版本信息"""
    return {
        "version": sys.version.split()[0],
        "version_info": sys.version_info[:3],
        "major": sys.version_info.major,
        "minor": sys.version_info.minor,
        "micro": sys.version_info.micro,
        "releaselevel": sys.version_info.releaselevel,
        "serial": sys.version_info.serial,
    }

def check_python_version(min_version=(3, 0)):
    """
    检查 Python 版本是否满足最低要求
    
    Args:
        min_version: 最低支持的 Python 版本元组，如 (3, 9)
    
    Returns:
        bool: 是否满足要求
    """
    current = sys.version_info[:2]
    
    print("=" * 60)
    print("🐍 Python 版本兼容性检查")
    print("=" * 60)
    print(f"✅ 当前 Python 版本: {sys.version}")
    print(f"✅ 版本号: {sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}")
    print(f"✅ 要求最低版本: {'.'.join(map(str, min_version))}+")
    print("-" * 60)
    
    if current < min_version:
        print(f"❌ 错误: Python 版本过低!")
        print(f"   需要: >={'.'.join(map(str, min_version))}")
        print(f"   当前: {'.'.join(map(str, current))}")
        print()
        print("💡 请升级 Python 版本或使用正确的 Python 解释器运行")
        print("=" * 60)
        return False
    
    print(f"✅ 版本检查通过! (>={'.'.join(map(str, min_version))})")
    
    check_features(current)
    
    print("=" * 60)
    return True

def check_features(version):
    """检查特定版本的特性支持"""
    print("\n📋 特性支持检查:")
    print("-" * 40)
    
    features = {
        (3, 6): "f-string, 变量注解",
        (3, 7): "dataclass, asyncio.run()",
        (3, 8): "海象运算符(:=), positional-only参数",
        (3, 9): "字典合并运算符(|), 类型泛型",
        (3, 10): "模式匹配(match/case), 更好的错误消息",
        (3, 11): "异常组(ExceptGroup*), Tomllib",
        (3, 12): "类型参数语法, 改进的错误消息",
        (3, 13): "实验性的自由线程CPython",
        (3, 14): "即将发布的特性",
    }
    
    for min_ver, feature in sorted(features.items()):
        status = "✅" if version >= min_ver else "⚠️"
        print(f"{status} Python {'.'.join(map(str, min_ver))}: {feature}")

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, Exception):
        pass

TIMEOUT_CONFIG = {
    'socket_connect': int(os.environ.get('TIMEOUT_SOCKET_CONNECT', '5')),
    'socket_read': int(os.environ.get('TIMEOUT_SOCKET_READ', '10')),
    'http_request': int(os.environ.get('TIMEOUT_HTTP_REQUEST', '30')),
    'subprocess_kill': int(os.environ.get('TIMEOUT_SUBPROCESS_KILL', '3')),
    'subprocess_wait': int(os.environ.get('TIMEOUT_SUBPROCESS_WAIT', '10')),
    'browser_page_load': int(os.environ.get('TIMEOUT_BROWSER_PAGE_LOAD', '60')),
    'browser_element': int(os.environ.get('TIMEOUT_BROWSER_ELEMENT', '30')),
    'tunnel_startup': int(os.environ.get('TIMEOUT_TUNNEL_STARTUP', '15')),
    'tunnel_heartbeat': int(os.environ.get('TIMEOUT_TUNNEL_HEARTBEAT', '300')),
    'email_send': int(os.environ.get('TIMEOUT_EMAIL_SEND', '30')),
    'file_operation': int(os.environ.get('TIMEOUT_FILE_OPERATION', '10')),
}

if not hasattr(subprocess, 'CREATE_NO_WINDOW'):
    subprocess.CREATE_NO_WINDOW = 0x08000000 if platform.system() == 'Windows' else 0


class AppException(Exception):
    """统一异常类 - 所有业务异常都使用此类"""
    
    CATEGORY_FILE = 'FILE'
    CATEGORY_NETWORK = 'NETWORK'
    CATEGORY_AUTH = 'AUTH'
    CATEGORY_BROWSER = 'BROWSER'
    CATEGORY_PARSE = 'PARSE'
    CATEGORY_CONFIG = 'CONFIG'
    CATEGORY_EXCEL = 'EXCEL'
    CATEGORY_EMAIL = 'EMAIL'
    CATEGORY_PERMISSION = 'PERMISSION'
    CATEGORY_RESOURCE = 'RESOURCE'
    CATEGORY_VALIDATION = 'VALIDATION'
    CATEGORY_DATABASE = 'DATABASE'
    
    _CATEGORY_CODES = {
        CATEGORY_FILE: 'FILE_ERROR',
        CATEGORY_NETWORK: 'NETWORK_ERROR',
        CATEGORY_AUTH: 'AUTH_ERROR',
        CATEGORY_BROWSER: 'BROWSER_ERROR',
        CATEGORY_PARSE: 'PARSE_ERROR',
        CATEGORY_CONFIG: 'CONFIG_ERROR',
        CATEGORY_EXCEL: 'EXCEL_ERROR',
        CATEGORY_EMAIL: 'EMAIL_ERROR',
        CATEGORY_PERMISSION: 'PERMISSION_ERROR',
        CATEGORY_RESOURCE: 'RESOURCE_ERROR',
        CATEGORY_VALIDATION: 'VALIDATION_ERROR',
        CATEGORY_DATABASE: 'DATABASE_ERROR',
    }
    
    def __init__(self, message: str, category: str = None, code: str = None, details: Any = None):
        self.message = message
        self.category = category or 'APP'
        self.code = code or self._CATEGORY_CODES.get(self.category, 'APP_ERROR')
        self.details = details or {}
        super().__init__(self.message)
    
    @classmethod
    def file_error(cls, message: str, file_path: str = None, operation: str = None, **kwargs):
        """文件操作异常"""
        details = {'file_path': file_path, 'operation': operation}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_FILE, details=details)
    
    @classmethod
    def network_error(cls, message: str, url: str = None, status_code: int = None, **kwargs):
        """网络请求异常"""
        details = {'url': url, 'status_code': status_code}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_NETWORK, details=details)
    
    @classmethod
    def auth_error(cls, message: str, cookie_file: str = None, **kwargs):
        """认证异常"""
        details = {'cookie_file': cookie_file}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_AUTH, details=details)
    
    @classmethod
    def browser_error(cls, message: str, browser_error: str = None, **kwargs):
        """浏览器操作异常"""
        details = {'browser_error': browser_error}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_BROWSER, details=details)
    
    @classmethod
    def parse_error(cls, message: str, data_type: str = None, raw_data: Any = None, **kwargs):
        """数据解析异常"""
        details = {'data_type': data_type, 'raw_data': str(raw_data)[:200] if raw_data else None}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_PARSE, details=details)
    
    @classmethod
    def config_error(cls, message: str, config_key: str = None, **kwargs):
        """配置异常"""
        details = {'config_key': config_key}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_CONFIG, details=details)
    
    @classmethod
    def excel_error(cls, message: str, excel_file: str = None, sheet_name: str = None, **kwargs):
        """Excel操作异常"""
        details = {'excel_file': excel_file, 'sheet_name': sheet_name}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_EXCEL, details=details)
    
    @classmethod
    def email_error(cls, message: str, smtp_host: str = None, recipient: str = None, **kwargs):
        """邮件发送异常"""
        details = {'smtp_host': smtp_host, 'recipient': recipient}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_EMAIL, details=details)
    
    @classmethod
    def permission_error(cls, message: str, path: str = None, operation: str = None, **kwargs):
        """权限异常"""
        details = {'path': path, 'operation': operation}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_PERMISSION, details=details)
    
    @classmethod
    def resource_error(cls, message: str, resource_type: str = None, resource_id: str = None, **kwargs):
        """资源异常"""
        details = {'resource_type': resource_type, 'resource_id': resource_id}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_RESOURCE, details=details)
    
    @classmethod
    def validation_error(cls, message: str, field: str = None, **kwargs):
        """数据验证异常"""
        details = {'field': field}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_VALIDATION, details=details)
    
    @classmethod
    def database_error(cls, message: str, table: str = None, query: str = None, **kwargs):
        """数据库异常"""
        details = {'table': table, 'query': query}
        details.update(kwargs)
        return cls(message, category=cls.CATEGORY_DATABASE, details=details)
    
    def to_dict(self) -> dict:
        return {
            'error': self.__class__.__name__,
            'category': self.category,
            'code': self.code,
            'message': self.message,
            'details': self.details
        }


class ExceptionHandler:
    """统一异常处理器"""
    
    _instance = None
    _logger = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self._setup_logger()
        self._error_counts = {}
        self._error_history = []
        self._max_history = 100
        self._suppressed_errors = {}
        self._suppression_window = 60
    
    def _setup_logger(self):
        """设置日志记录器"""
        if self._logger is None:
            self._logger = logging.getLogger('ExceptionHandler')
            self._logger.setLevel(logging.ERROR)
            if not self._logger.handlers:
                handler = logging.StreamHandler()
                handler.setFormatter(logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S'
                ))
                self._logger.addHandler(handler)
    
    def handle(self, error: Exception, context: str = '', show_traceback: bool = True) -> str:
        """处理异常并返回错误信息"""
        error_type = type(error).__name__
        error_msg = str(error)
        
        self._error_counts[error_type] = self._error_counts.get(error_type, 0) + 1
        
        error_record = {
            'timestamp': datetime.now().isoformat(),
            'type': error_type,
            'message': error_msg,
            'context': context
        }
        self._error_history.append(error_record)
        if len(self._error_history) > self._max_history:
            self._error_history.pop(0)
        
        if isinstance(error, AppException):
            full_msg = f"[{error.code}] {error.message}"
        else:
            full_msg = f"[{error_type}] {error_msg}"
        
        if context:
            full_msg = f"{context}: {full_msg}"
        
        print(f'错误: {full_msg}')
        if show_traceback:
            traceback.print_exc()
        
        self._logger.error(full_msg, exc_info=show_traceback)
        
        return full_msg
    
    def try_execute(self, func: Callable, default: Any = None, context: str = '') -> Any:
        """统一异常处理包装器 - 用于需要捕获异常并返回默认值的场景"""
        try:
            return func()
        except Exception as e:
            self.handle(e, context)
            return default
    
    def try_execute_with_error(self, func: Callable, context: str = '') -> Tuple[Any, str]:
        """统一异常处理包装器 - 用于需要获取错误信息的场景"""
        try:
            return func(), None
        except Exception as e:
            error_msg = self.handle(e, context)
            return None, error_msg
    
    def get_error_counts(self) -> dict:
        """获取错误统计"""
        return self._error_counts.copy()
    
    def get_error_history(self, limit: int = 10) -> List[dict]:
        """获取错误历史"""
        return self._error_history[-limit:]
    
    def retry_on_exception(self, func: Callable, max_retries: int = 3, delay: float = 1.0, context: str = '') -> Any:
        """带重试的异常处理"""
        last_error = None
        for attempt in range(max_retries):
            try:
                return func()
            except Exception as e:
                last_error = e
                if attempt < max_retries - 1:
                    time.sleep(delay * (attempt + 1))
        if last_error:
            self.handle(last_error, context)
        raise last_error
    
    def suppress_duplicate_errors(self, error_key: str, window_seconds: int = None) -> bool:
        """抑制重复错误日志（避免日志爆炸）"""
        if window_seconds is None:
            window_seconds = self._suppression_window
        
        current_time = time.time()
        if error_key in self._suppressed_errors:
            last_time = self._suppressed_errors[error_key]
            if current_time - last_time < window_seconds:
                return True
        
        self._suppressed_errors[error_key] = current_time
        return False
    
    def get_category_stats(self) -> dict:
        """获取按类别分组的错误统计"""
        category_counts = {}
        for record in self._error_history:
            error_type = record.get('type', 'Unknown')
            category_counts[error_type] = category_counts.get(error_type, 0) + 1
        return category_counts
    
    def clear_old_suppressions(self, max_age: int = 300):
        """清理过期的错误抑制记录"""
        current_time = time.time()
        keys_to_remove = [
            key for key, timestamp in self._suppressed_errors.items()
            if current_time - timestamp > max_age
        ]
        for key in keys_to_remove:
            del self._suppressed_errors[key]


class ExceptionContext:
    """异常处理上下文管理器"""
    
    def __init__(self, context: str = '', default: Any = None, show_traceback: bool = True):
        self.context = context
        self.default = default
        self.show_traceback = show_traceback
        self.handler = ExceptionHandler()
        self.result = None
        self.error = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is not None:
            self.error = self.handler.handle(exc_val, self.context, self.show_traceback)
            self.result = self.default
            return True
        return False
    
    def get_result(self) -> Tuple[Any, str]:
        """获取结果和错误信息"""
        return self.result, self.error


T = TypeVar('T')


def safe_call(func: Callable[..., T], *args, default: T = None, context: str = '', **kwargs) -> T:
    """安全调用函数，返回默认值"""
    handler = ExceptionHandler()
    return handler.try_execute(lambda: func(*args, **kwargs), default, context)


def safe_call_with_error(func: Callable[..., T], *args, context: str = '', **kwargs) -> Tuple[T, str]:
    """安全调用函数，返回(结果, 错误信息)"""
    handler = ExceptionHandler()
    return handler.try_execute_with_error(lambda: func(*args, **kwargs), context)


def handle_error(error: Exception, context: str = '') -> str:
    """处理已捕获的异常"""
    handler = ExceptionHandler()
    return handler.handle(error, context)


def safe_execute_func(func: Callable, default: Any = None, context: str = '') -> Any:
    """统一异常处理包装器 - 用于需要捕获异常并返回默认值的场景"""
    handler = ExceptionHandler()
    return handler.try_execute(func, default, context)


def auto_clean_temp_dir():
    """检查temp目录大小，超过3MB立即清理所有文件"""
    temp_dir = os.path.join(PROJECT_DIR, 'temp')
    if not os.path.isdir(temp_dir):
        return
    temp_size = 0
    for f in os.listdir(temp_dir):
        fp = os.path.join(temp_dir, f)
        if os.path.isfile(fp):
            temp_size += os.path.getsize(fp)
    if temp_size > 3 * 1024 * 1024:
        cleaned = 0
        for f in os.listdir(temp_dir):
            fp = os.path.join(temp_dir, f)
            if os.path.isfile(fp):
                safe_execute_func(lambda: os.remove(fp), context='auto_clean_temp_dir清理')
                cleaned += 1
        print(f"[Clean] temp目录超过3MB({temp_size / (1024 * 1024):.1f}MB)，已清理{cleaned}个文件")


def safe_execute_with_error(func: Callable, context: str = '') -> Tuple[Any, str]:
    """统一异常处理包装器 - 用于需要获取错误信息的场景"""
    handler = ExceptionHandler()
    return handler.try_execute_with_error(func, context)


def handle_exception(e: Exception, context: str = '') -> str:
    """统一异常处理函数 - 用于已捕获异常的场景"""
    handler = ExceptionHandler()
    return handler.handle(e, context)


def exception_handler(context: str = '', default: Any = None, reraise: bool = False, custom_exc: type = None):
    """统一异常处理装饰器
    
    Args:
        context: 错误上下文描述
        default: 异常时返回的默认值
        reraise: 是否在异常时重新抛出统一的自定义异常
        custom_exc: 自定义异常类型（如果reraise=True）
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except AppException:
                raise
            except Exception as e:
                handler = ExceptionHandler()
                handler.handle(e, context)
                if reraise and custom_exc:
                    raise custom_exc(str(e)) from e
                return default
        return wrapper
    return decorator


def file_operation_handler(operation: str):
    """文件操作异常处理装饰器"""
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except PermissionError as e:
                raise AppException.permission_error(
                    f"文件操作{operation}失败（权限不足）",
                    path=str(args[0]) if args else kwargs.get('path'),
                    operation=operation
                ) from e
            except FileNotFoundError as e:
                raise AppException.file_error(
                    f"文件操作{operation}失败（文件不存在）",
                    file_path=str(args[0]) if args else kwargs.get('path'),
                    operation=operation
                ) from e
            except OSError as e:
                raise AppException.file_error(
                    f"文件操作{operation}失败（系统错误）: {e}",
                    file_path=str(args[0]) if args else kwargs.get('path'),
                    operation=operation
                ) from e
        return wrapper
    return decorator


def network_handler(url: str = None):
    """网络请求异常处理装饰器"""
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except urllib.error.HTTPError as e:
                raise AppException.network_error(
                    f"网络请求失败（HTTP {e.code}）",
                    url=url or str(args[0]) if args else None,
                    status_code=e.code
                ) from e
            except urllib.error.URLError as e:
                raise AppException.network_error(
                    f"网络请求失败（连接错误）: {e.reason}",
                    url=url or str(args[0]) if args else None
                ) from e
            except Exception as e:
                raise AppException.network_error(
                    f"网络请求失败: {e}",
                    url=url or str(args[0]) if args else None
                ) from e
        return wrapper
    return decorator


def json_handler(context: str = ''):
    """JSON解析异常处理装饰器"""
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except json.JSONDecodeError as e:
                raise AppException.parse_error(
                    f"JSON解析失败: {e}",
                    data_type='json',
                    raw_data=str(e)
                ) from e
        return wrapper
    return decorator


def excel_handler(operation: str = '操作'):
    """Excel操作异常处理装饰器"""
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except PermissionError as e:
                if "sharing violation" in str(e).lower() or "另一个程序" in str(e) or "正在使用" in str(e):
                    raise AppException.excel_error(
                        f"Excel文件{operation}失败（共享违规）",
                        excel_file=str(args[0]) if args else kwargs.get('excel_file')
                    ) from e
                raise AppException.excel_error(
                    f"Excel文件{operation}失败（权限不足）",
                    excel_file=str(args[0]) if args else kwargs.get('excel_file')
                ) from e
            except Exception as e:
                raise AppException.excel_error(
                    f"Excel文件{operation}失败: {e}",
                    excel_file=str(args[0]) if args else kwargs.get('excel_file')
                ) from e
        return wrapper
    return decorator

PSUTIL_AVAILABLE = psutil is not None
PROMETHEUS_AVAILABLE = Counter is not None
FLASK_RESTX_AVAILABLE = False  # 已迁移到 FastAPI，不再使用 flask-restx
PYDANTIC_AVAILABLE = BaseModel is not None

if pd is None:
    print("警告: pandas未安装，Excel对比功能将不可用")

if async_playwright is None:
    print("警告: playwright未安装，浏览器自动化功能将不可用")
    print("请运行: pip install playwright && playwright install chromium")

if openpyxl is None:
    print("警告: openpyxl未安装，Excel功能将不可用")

file_write_lock = threading.Lock()

# Web日志文件路径
web_log_file = None

class TeeOutput:
    """同时输出到控制台和文件"""
    def __init__(self, original, log_file_path=None):
        self.original = original
        self.log_file_path = log_file_path
        self.file = None
        if log_file_path:
            self._init_log_file(log_file_path)
    
    def _init_log_file(self, log_file_path, retry_count=0):
        max_retries = 3
        try:
            log_dir = os.path.dirname(log_file_path)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)
            
            if os.path.exists(log_file_path):
                try:
                    test_fd = os.open(log_file_path, os.O_WRONLY | os.O_APPEND)
                    os.close(test_fd)
                except OSError as e:
                    if retry_count < max_retries:
                        backup_path = f"{log_file_path}.locked_{time.strftime('%H%M%S')}"
                        try:
                            os.rename(log_file_path, backup_path)
                            print(f"[TeeOutput] ⚠️ 日志文件被锁定，已备份为: {backup_path}")
                        except Exception as e:
                            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                            pass
                        time.sleep(0.5 * (retry_count + 1))
                        return self._init_log_file(log_file_path, retry_count + 1)
                    else:
                        raise
            
            self.file = open(log_file_path, 'a', encoding='utf-8')
            
        except PermissionError as e:
            if retry_count < max_retries:
                alt_path = f"{log_file_path}.{time.strftime('%Y%m%d_%H%M%S')}"
                print(f"[TeeOutput] ⚠️ 权限不足，尝试使用备用文件: {alt_path}")
                time.sleep(0.3 * (retry_count + 1))
                return self._init_log_file(alt_path, retry_count + 1)
            else:
                print(f"[TeeOutput] ❌ 无法打开日志文件（已重试{max_retries}次），将仅输出到控制台")
                print(f"[TeeOutput]    文件路径: {log_file_path}")
                print(f"[TeeOutput]    错误: {e}")
                self.file = None
                
        except Exception as e:
            if retry_count < max_retries:
                time.sleep(0.2 * (retry_count + 1))
                return self._init_log_file(log_file_path, retry_count + 1)
            else:
                print(f"[TeeOutput] ❌ 初始化失败（已重试{max_retries}次）: {e}")
                self.file = None
    
    def write(self, text):
        _output_text = text
        
        if text.strip():
            _full_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            
            _has_timestamp = (
                text.strip().startswith(f'[{_full_timestamp[:10]}') or 
                text.strip().startswith(f'[{_full_timestamp[:4]}')
            )
            
            _is_flask_access_log = (
                ' - - [' in text and 
                ('"GET ' in text or '"POST ' in text or '"HEAD ' in text or 
                 '"PUT ' in text or '"DELETE ' in text or '"PATCH ' in text or
                 '"OPTIONS ' in text)
            )
            
            if not _has_timestamp:
                if _is_flask_access_log:
                    _access_match = re.search(r'^(\S+)\s+-\s+-\s+\[([^\]]+)\]\s+"([^"]+)"\s+(\d+)\s*(.*)', text.strip())
                    if _access_match:
                        _client_ip, _flask_time, _request_line, _status_code, _extra = _access_match.groups()
                        _output_text = f"[{_full_timestamp}] {_client_ip} {_request_line} {_status_code}\n"
                    else:
                        _lines = text.split('\n')
                        _timestamped_lines = []
                        for _line in _lines:
                            if _line.strip():
                                _timestamped_lines.append(f"[{_full_timestamp}] {_line}")
                            else:
                                _timestamped_lines.append(_line)
                        _output_text = '\n'.join(_timestamped_lines)
                else:
                    _lines = text.split('\n')
                    _timestamped_lines = []
                    for _line in _lines:
                        if _line.strip():
                            _timestamped_lines.append(f"[{_full_timestamp}] {_line}")
                        else:
                            _timestamped_lines.append(_line)
                    _output_text = '\n'.join(_timestamped_lines)
        
        self.original.write(_output_text)
        
        if self.file:
            safe_execute_func(
                lambda: (self.file.write(_output_text), self.file.flush()),
                context='TeeOutput写入'
            )
    
    def flush(self):
        self.original.flush()
        if self.file:
            safe_execute_func(
                lambda: self.file.flush(),
                context='TeeOutput刷新'
            )
    
    def close(self):
        if self.file:
            safe_execute_func(
                lambda: self.file.close(),
                context='TeeOutput关闭'
            )
    
    def isatty(self):
        return False
    
    def __del__(self):
        """析构函数，确保文件资源被正确释放"""
        try:
            self.close()
        except Exception as e:
            logger.debug(f"Silent exception: {e}")

def setup_web_logging():
    """设置Web模式下的日志输出（追加模式，保留shell脚本已写入的完整启动日志）"""
    global web_log_file
    web_log_file = PathManager.get_web_output_file()
    need_header = True
    if os.path.exists(web_log_file):
        try:
            with open(web_log_file, 'r', encoding='utf-8') as f:
                content = f.read().strip()
            if content:
                need_header = False
        except Exception as e:
            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
            pass
    if need_header:
        def _write_header():
            with open(web_log_file, 'a', encoding='utf-8') as f:
                f.write("=" * 50 + "\nSzwego商品爬虫 - Web服务\n" + "=" * 50 + "\n")
        safe_execute_func(_write_header, context='setup_web_logging')
    sys.stdout = TeeOutput(sys.stdout, web_log_file)
    sys.stderr = TeeOutput(sys.stderr, web_log_file)

def log_print(*args, **kwargs):
    """同时输出到控制台和 web_output.log（自动添加时间戳）"""
    global web_log_file
    msg = ' '.join(str(a) for a in args)
    _log_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _msg_with_timestamp = f"[{_log_timestamp}] {msg}"
    print(_msg_with_timestamp, **kwargs)
    if web_log_file:
        def _write_log():
            with open(web_log_file, 'a', encoding='utf-8') as f:
                f.write(_msg_with_timestamp + '\n')
        safe_execute_func(_write_log, context='log_print')

def format_size(size_bytes: int) -> str:
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} PB"


def setup_logger(log_file: Optional[str] = None, log_level: int = logging.INFO, stream=None) -> logging.Logger:
    logger = logging.getLogger('FileCleaner')
    logger.setLevel(log_level)
    logger.handlers.clear()
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_handler = logging.StreamHandler(stream)
    console_handler.setLevel(log_level)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    if log_file:
        safe_execute_func(
            lambda: logger.addHandler(logging.FileHandler(log_file, encoding='utf-8')),
            context='setup_logger'
        )
        for h in logger.handlers:
            if isinstance(h, logging.FileHandler):
                h.setLevel(log_level)
                h.setFormatter(formatter)
    return logger


IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.svg'}
VIDEO_EXTENSIONS = {'.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm', '.m4v'}
MEDIA_EXTENSIONS = IMAGE_EXTENSIONS | VIDEO_EXTENSIONS
EXCLUDE_EXTENSIONS = {'.log', '.sh', '.py', '.bat', '.json', '.md', '.txt', '.html', '.htm', '.sql', '.xml', '.yml', '.yaml', '.ini', '.cfg', '.conf'}
EXCLUDE_FOLDERS = {'file', 'config', '__pycache__', 'clean', '.venv', 'templates', '.git', '.idea', 'node_modules', '.vscode', 'static'}
EXCLUDE_FILE_NAMES = {'.DS_Store', 'Thumbs.db', '.gitkeep', '.gitignore'}  # 特殊文件名保护


def clean_old_files(
        directory: str,
        dry_run: bool = False,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("开始清理旧文件")
    logger.info(f"清理目录: {directory}")
    logger.info(f"保留规则: 最新的一组文件（'_'前完全一致的为一组）")
    logger.info(f"排序方式: 按文件下载到本地的毫秒时间")
    logger.info(f"测试模式: {'是' if dry_run else '否'}")
    logger.info("=" * 60)

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    if not directory.is_dir():
        logger.error(f"路径不是目录: {directory}")
        return

    matched_files = []
    logger.info("扫描文件中...")
    for file in directory.iterdir():
        if file.is_file():
            ext = file.suffix.lower()
            if ext in MEDIA_EXTENSIONS:
                with ExceptionContext(f"clean_old_files扫描 {file.name}", default=None) as ctx:
                    file_stat = file.stat()
                    mtime = file_stat.st_mtime
                    download_time = datetime.fromtimestamp(mtime)

                    name_without_ext = file.stem
                    group_key = name_without_ext.split('_')[0] if '_' in name_without_ext else name_without_ext

                    matched_files.append({
                        'file': file,
                        'group_key': group_key,
                        'ext': ext,
                        'is_image': ext in IMAGE_EXTENSIONS,
                        'is_video': ext in VIDEO_EXTENSIONS,
                        'mtime': mtime,
                        'size': file_stat.st_size,
                        'download_time': download_time
                    })

    if not matched_files:
        logger.warning("没有找到图片或视频文件")
        return

    matched_files.sort(key=lambda x: x['mtime'], reverse=True)
    logger.info(f"按文件下载到本地的毫秒时间排序（从最新到最旧）")

    total_files = len(matched_files)
    image_count = sum(1 for f in matched_files if f['is_image'])
    video_count = sum(1 for f in matched_files if f['is_video'])
    total_size = sum(f['size'] for f in matched_files)

    logger.info(f"找到 {total_files} 个符合条件的文件")
    logger.info(f"  - 图片: {image_count} 个")
    logger.info(f"  - 视频: {video_count} 个")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    groups = {}
    for file_info in matched_files:
        group_key = file_info['group_key']
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(file_info)

    def get_group_latest_mtime(group_key):
        group_files = groups[group_key]
        return max(f['mtime'] for f in group_files)

    sorted_group_keys = sorted(groups.keys(), key=get_group_latest_mtime, reverse=True)

    logger.info(f"共 {len(groups)} 组文件（按'_'前部分分组）")
    for i, group_key in enumerate(sorted_group_keys[:5], 1):
        group_files = groups[group_key]
        latest_file = max(group_files, key=lambda x: x['mtime'])
        group_time = latest_file['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        group_images = sum(1 for f in group_files if f['is_image'])
        group_videos = sum(1 for f in group_files if f['is_video'])
        logger.info(
            f"  组{i}: {group_key} - {len(group_files)}个文件 (图片:{group_images}, 视频:{group_videos}, 最新下载时间:{group_time})")

    if len(sorted_group_keys) > 5:
        logger.info(f"  ... 还有 {len(sorted_group_keys) - 5} 组")

    latest_group_key = sorted_group_keys[0]
    latest_group = groups[latest_group_key]
    latest_file = max(latest_group, key=lambda x: x['mtime'])
    latest_time_str = latest_file['download_time'].strftime('%Y-%m-%d %H:%M:%S')

    logger.info(f"\n将保留最新的一组文件（组名: {latest_group_key}, 最新下载时间: {latest_time_str}）")
    logger.info(f"  该组共 {len(latest_group)} 个文件")
    logger.info(f"  将删除其他 {len(sorted_group_keys) - 1} 组旧文件，共 {total_files - len(latest_group)} 个文件")

    logger.info("\n文件列表（从最新到最旧）：")
    for i, file_info in enumerate(matched_files[:10], 1):
        is_latest = file_info['group_key'] == latest_group_key
        status = "保留" if is_latest else "删除"
        file_type = "图片" if file_info['is_image'] else "视频"
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        logger.info(
            f"{i:3d}. [{status}] {file_info['file'].name} ({file_type}, {format_size(file_info['size'])}, 下载时间: {download_time_str})")

    if len(matched_files) > 10:
        logger.info(f"... 还有 {len(matched_files) - 10} 个文件")

    files_to_delete = [f for f in matched_files if f['group_key'] != latest_group_key]

    if not files_to_delete:
        logger.info("没有需要删除的文件")
        logger.info("清理完成")
        return

    delete_size = sum(f['size'] for f in files_to_delete)
    logger.info(f"\n准备删除 {len(files_to_delete)} 个旧文件，释放空间: {format_size(delete_size)}")

    for file_info in files_to_delete:
        file_type = "图片" if file_info['is_image'] else "视频"
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        logger.info(
            f"  - {file_info['file'].name} ({file_type}, {format_size(file_info['size'])}, 下载时间: {download_time_str})")

    if dry_run:
        logger.info("\n[测试模式] 未实际删除文件")
        logger.info("清理完成（测试模式）")
        return

    logger.info("\n开始删除文件...")
    deleted_count = 0
    failed_count = 0
    deleted_size = 0

    for file_info in files_to_delete:
        with ExceptionContext(f"删除文件 {file_info['file'].name}", default=False) as ctx:
            file_info['file'].unlink()
            deleted_count += 1
            deleted_size += file_info['size']
            logger.info(f"已删除: {file_info['file'].name} ({format_size(file_info['size'])})")
            if ctx.error:
                failed_count += 1

    logger.info("\n" + "=" * 60)
    logger.info("清理完成")
    logger.info(f"成功删除: {deleted_count} 个文件，释放空间: {format_size(deleted_size)}")
    if failed_count > 0:
        logger.warning(f"删除失败: {failed_count} 个文件")
    logger.info("=" * 60)


def clean_old_files_by_time(
        directory: str,
        minutes: int = 5,
        dry_run: bool = False,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("开始按时间清理旧文件")
    logger.info(f"清理目录: {directory}")
    logger.info(f"删除规则: 删除 {minutes} 分钟前下载到本地的所有文件")
    logger.info(f"测试模式: {'是' if dry_run else '否'}")
    logger.info("=" * 60)

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    if not directory.is_dir():
        logger.error(f"路径不是目录: {directory}")
        return

    matched_files = []
    logger.info("扫描文件中...")

    current_time = datetime.now()
    cutoff_time = current_time.timestamp() - (minutes * 60)

    for file in directory.iterdir():
        if file.is_file():
            ext = file.suffix.lower()
            if ext in MEDIA_EXTENSIONS:
                with ExceptionContext(f"clean_old_files_by_time扫描 {file.name}", default=None) as ctx:
                    file_stat = file.stat()
                    mtime = file_stat.st_mtime
                    download_time = datetime.fromtimestamp(mtime)

                    matched_files.append({
                        'file': file,
                        'ext': ext,
                        'is_image': ext in IMAGE_EXTENSIONS,
                        'is_video': ext in VIDEO_EXTENSIONS,
                        'mtime': mtime,
                        'size': file_stat.st_size,
                        'download_time': download_time
                    })

    if not matched_files:
        logger.warning("没有找到图片或视频文件")
        return

    matched_files.sort(key=lambda x: x['mtime'], reverse=True)

    total_files = len(matched_files)
    image_count = sum(1 for f in matched_files if f['is_image'])
    video_count = sum(1 for f in matched_files if f['is_video'])
    total_size = sum(f['size'] for f in matched_files)

    logger.info(f"找到 {total_files} 个符合条件的文件")
    logger.info(f"  - 图片: {image_count} 个")
    logger.info(f"  - 视频: {video_count} 个")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    cutoff_time_str = datetime.fromtimestamp(cutoff_time).strftime('%Y-%m-%d %H:%M:%S')
    current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')

    logger.info(f"\n当前时间: {current_time_str}")
    logger.info(f"删除时间阈值: {cutoff_time_str} ( {minutes} 分钟前)")

    files_to_delete = [f for f in matched_files if f['mtime'] < cutoff_time]
    files_to_keep = [f for f in matched_files if f['mtime'] >= cutoff_time]

    logger.info(f"\n将保留 {len(files_to_keep)} 个文件（{minutes} 分钟内下载的）")
    logger.info(f"将删除 {len(files_to_delete)} 个文件（{minutes} 分钟前下载的）")

    if not files_to_delete:
        logger.info("没有需要删除的文件")
        logger.info("清理完成")
        return

    delete_size = sum(f['size'] for f in files_to_delete)
    logger.info(f"\n准备删除 {len(files_to_delete)} 个旧文件，释放空间: {format_size(delete_size)}")

    logger.info("\n将要删除的文件列表：")
    for i, file_info in enumerate(files_to_delete, 1):
        file_type = "图片" if file_info['is_image'] else "视频"
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        time_diff = (current_time - file_info['download_time']).total_seconds() / 60
        logger.info(
            f"{i:3d}. {file_info['file'].name} ({file_type}, {format_size(file_info['size'])}, 下载时间: {download_time_str}, {time_diff:.1f}分钟前)")

    if dry_run:
        logger.info("\n[测试模式] 未实际删除文件")
        logger.info("清理完成（测试模式）")
        return

    logger.info("\n开始删除文件...")
    deleted_count = 0
    failed_count = 0
    deleted_size = 0

    for file_info in files_to_delete:
        with ExceptionContext(f"删除文件 {file_info['file'].name}", default=False) as ctx:
            file_info['file'].unlink()
            deleted_count += 1
            deleted_size += file_info['size']
            logger.info(f"已删除: {file_info['file'].name} ({format_size(file_info['size'])})")
            if ctx.error:
                failed_count += 1

    logger.info("\n" + "=" * 60)
    logger.info("清理完成")
    logger.info(f"成功删除: {deleted_count} 个文件，释放空间: {format_size(deleted_size)}")
    if failed_count > 0:
        logger.warning(f"删除失败: {failed_count} 个文件")
    logger.info("=" * 60)


def list_files(
        directory: str,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("扫描文件列表")
    logger.info(f"扫描目录: {directory}")
    logger.info(f"排序方式: 按文件下载到本地的毫秒时间")
    logger.info("扫描模式: 递归扫描所有子目录")
    logger.info("=" * 60)

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    matched_files = []
    scanned_dirs = 0
    logger.info("扫描文件中...")
    
    for root, dirs, files in os.walk(directory):
        root_path = Path(root)
        scanned_dirs += 1
        
        for file in files:
            file_path = root_path / file
            ext = Path(file).suffix.lower()
            if ext in MEDIA_EXTENSIONS:
                with ExceptionContext(f"list_files扫描 {file_path}", default=None) as ctx:
                    file_stat = file_path.stat()
                    mtime = file_stat.st_mtime
                    download_time = datetime.fromtimestamp(mtime)

                    matched_files.append({
                        'file': file_path,
                        'relative_path': str(file_path.relative_to(directory)),
                        'ext': ext,
                        'is_image': ext in IMAGE_EXTENSIONS,
                        'is_video': ext in VIDEO_EXTENSIONS,
                        'mtime': mtime,
                        'size': file_stat.st_size,
                        'download_time': download_time
                    })

    if not matched_files:
        logger.warning("没有找到图片或视频文件")
        return

    matched_files.sort(key=lambda x: x['mtime'], reverse=True)

    total_files = len(matched_files)
    image_count = sum(1 for f in matched_files if f['is_image'])
    video_count = sum(1 for f in matched_files if f['is_video'])
    total_size = sum(f['size'] for f in matched_files)

    logger.info(f"扫描了 {scanned_dirs} 个目录")
    logger.info(f"找到 {total_files} 个符合条件的文件")
    logger.info(f"  - 图片: {image_count} 个")
    logger.info(f"  - 视频: {video_count} 个")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    logger.info("\n文件列表：")
    for i, file_info in enumerate(matched_files, 1):
        file_type = "图片" if file_info['is_image'] else "视频"
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        logger.info(
            f"{i:3d}. {file_info['relative_path']} ({file_type}, {format_size(file_info['size'])}, 下载时间: {download_time_str})")

    logger.info("\n扫描完成")


def clean_all_files(
        directory: str,
        dry_run: bool = False,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("开始删除所有文件和文件夹")
    logger.info(f"清理目录: {directory}")
    logger.info(f"删除规则: 删除所有文件和文件夹，除了 log, sh, py, bat 格式的文件")
    logger.info(f"测试模式: {'是' if dry_run else '否'}")
    logger.info("=" * 60)

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    if not directory.is_dir():
        logger.error(f"路径不是目录: {directory}")
        return

    files_to_delete = []
    folders_to_delete = []
    logger.info("扫描文件和文件夹中...")

    for item in directory.iterdir():
        if item.is_file():
            ext = item.suffix.lower()
            if item.name in EXCLUDE_FILE_NAMES:
                logger.info(f"保留文件: {item.name}")
            elif ext not in EXCLUDE_EXTENSIONS:
                files_to_delete.append(item)
            else:
                logger.info(f"保留文件: {item.name}")
        elif item.is_dir():
            if item.name not in EXCLUDE_FOLDERS:
                folders_to_delete.append(item)
            else:
                logger.info(f"保留文件夹: {item.name}")

    total_files = len(files_to_delete)
    total_folders = len(folders_to_delete)
    total_size = sum(
        safe_call(lambda f=item: f.stat().st_size, default=0, context='clean_all_files获取文件大小')
        for item in files_to_delete
    )

    logger.info(f"找到 {total_files} 个文件和 {total_folders} 个文件夹")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    if not files_to_delete and not folders_to_delete:
        logger.warning("没有需要删除的文件或文件夹")
        logger.info("清理完成")
        return

    logger.info(f"\n将删除 {total_files} 个文件和 {total_folders} 个文件夹")

    if files_to_delete:
        logger.info("\n将要删除的文件列表：")
        for i, item in enumerate(files_to_delete, 1):
            file_size = safe_call(lambda f=item: f.stat().st_size, default=0, context='clean_all_files显示文件')
            logger.info(f"{i:3d}. {item.name} ({format_size(file_size)})")

    if folders_to_delete:
        logger.info("\n将要删除的文件夹列表：")
        for i, folder in enumerate(folders_to_delete, 1):
            logger.info(f"{i:3d}. {folder.name}")

    if dry_run:
        logger.info("\n[测试模式] 未实际删除文件和文件夹")
        logger.info("清理完成（测试模式）")
        return

    logger.info("\n开始删除文件...")
    deleted_files_count = 0
    deleted_folders_count = 0
    deleted_size = 0
    failed_count = 0

    for file in files_to_delete:
        with ExceptionContext(f"删除文件 {file.name}", default=False) as ctx:
            file_size = file.stat().st_size
            file.unlink()
            deleted_files_count += 1
            deleted_size += file_size
            logger.info(f"已删除文件: {file.name} ({format_size(file_size)})")
            if ctx.error:
                failed_count += 1

    for folder in folders_to_delete:
        with ExceptionContext(f"删除文件夹 {folder.name}", default=False) as ctx:
            shutil.rmtree(folder)
            deleted_folders_count += 1
            logger.info(f"已删除文件夹: {folder.name}")
            if ctx.error:
                failed_count += 1

    logger.info("\n" + "=" * 60)
    logger.info("清理完成")
    logger.info(f"成功删除: {deleted_files_count} 个文件，{deleted_folders_count} 个文件夹，释放空间: {format_size(deleted_size)}")
    if failed_count > 0:
        logger.warning(f"删除失败: {failed_count} 个项目")
    logger.info("=" * 60)


def clean_png_files(
        directory: str,
        dry_run: bool = False,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("开始删除PNG文件")
    logger.info(f"清理目录: {directory}")
    logger.info(f"删除规则: 删除所有以.png结尾的文件，不保留任何文件")
    logger.info(f"测试模式: {'是' if dry_run else '否'}")
    logger.info("=" * 60)

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    if not directory.is_dir():
        logger.error(f"路径不是目录: {directory}")
        return

    matched_files = []
    logger.info("扫描文件中...")

    for file in directory.iterdir():
        if file.is_file() and file.name.lower().endswith('.png'):
            file_stat = file.stat()
            mtime = file_stat.st_mtime
            download_time = datetime.fromtimestamp(mtime)

            matched_files.append({
                'file': file,
                'main_num': file.name,
                'sub_num': 0,
                'ext': '.png',
                'is_image': True,
                'is_video': False,
                'mtime': mtime,
                'size': file_stat.st_size,
                'download_time': download_time
            })

    if not matched_files:
        logger.warning("没有找到PNG文件")
        return

    matched_files.sort(key=lambda x: x['mtime'], reverse=True)

    total_files = len(matched_files)
    total_size = sum(f['size'] for f in matched_files)

    logger.info(f"找到 {total_files} 个PNG文件")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    logger.info(f"\n将删除所有 {total_files} 个PNG文件，释放空间: {format_size(total_size)}")

    logger.info("\n文件列表：")
    for i, file_info in enumerate(matched_files, 1):
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        logger.info(
            f"{i:3d}. {file_info['file'].name} ({format_size(file_info['size'])}, 下载时间: {download_time_str})")

    if dry_run:
        logger.info("\n[测试模式] 未实际删除文件")
        logger.info("清理完成（测试模式）")
        return

    logger.info("\n开始删除文件...")
    deleted_count = 0
    failed_count = 0
    deleted_size = 0

    for file_info in matched_files:
        with ExceptionContext(f"删除PNG文件 {file_info['file'].name}", default=False) as ctx:
            file_info['file'].unlink()
            deleted_count += 1
            deleted_size += file_info['size']
            logger.info(f"已删除: {file_info['file'].name} ({format_size(file_info['size'])})")
            if ctx.error:
                failed_count += 1

    logger.info("\n" + "=" * 60)
    logger.info("清理完成")
    logger.info(f"成功删除: {deleted_count} 个文件，释放空间: {format_size(deleted_size)}")
    if failed_count > 0:
        logger.warning(f"删除失败: {failed_count} 个文件")
    logger.info("=" * 60)


def clean_media_files(
        directory: str,
        dry_run: bool = False,
        log_file: Optional[str] = None,
        log_level: int = logging.INFO,
        stream=None
) -> None:
    logger = setup_logger(log_file, log_level, stream)
    logger.info("=" * 60)
    logger.info("开始删除媒体文件")
    logger.info(f"清理目录: {directory}")
    logger.info(f"删除规则: 删除所有以.png、.jpg、.gif和.mp4结尾的文件，不保留任何文件")
    logger.info(f"测试模式: {'是' if dry_run else '否'}")
    logger.info("=" * 60)

    valid_extensions = {'.png', '.jpg', '.gif', '.mp4'}

    directory = Path(directory)

    if not directory.exists():
        logger.error(f"目录不存在: {directory}")
        return

    if not directory.is_dir():
        logger.error(f"路径不是目录: {directory}")
        return

    matched_files = []
    logger.info("扫描文件中...")

    for file in directory.iterdir():
        if file.is_file():
            ext = file.name.lower().split('.')[-1] if '.' in file.name else ''
            ext_with_dot = f'.{ext}' if ext else ''

            if ext_with_dot in valid_extensions:
                file_stat = file.stat()
                mtime = file_stat.st_mtime
                download_time = datetime.fromtimestamp(mtime)

                matched_files.append({
                    'file': file,
                    'main_num': file.name,
                    'sub_num': 0,
                    'ext': ext_with_dot,
                    'is_image': ext_with_dot in {'.png', '.jpg', '.gif'},
                    'is_video': ext_with_dot == '.mp4',
                    'mtime': mtime,
                    'size': file_stat.st_size,
                    'download_time': download_time
                })

    if not matched_files:
        logger.warning("没有找到PNG、JPG、GIF或MP4文件")
        return

    matched_files.sort(key=lambda x: x['mtime'], reverse=True)

    total_files = len(matched_files)
    png_count = sum(1 for f in matched_files if f['ext'] == '.png')
    jpg_count = sum(1 for f in matched_files if f['ext'] == '.jpg')
    gif_count = sum(1 for f in matched_files if f['ext'] == '.gif')
    video_count = sum(1 for f in matched_files if f['is_video'])
    total_size = sum(f['size'] for f in matched_files)

    logger.info(f"找到 {total_files} 个媒体文件")
    logger.info(f"  - PNG图片: {png_count} 个")
    logger.info(f"  - JPG图片: {jpg_count} 个")
    logger.info(f"  - GIF图片: {gif_count} 个")
    logger.info(f"  - MP4视频: {video_count} 个")
    logger.info(f"  - 总大小: {format_size(total_size)}")

    logger.info(f"\n将删除所有 {total_files} 个媒体文件，释放空间: {format_size(total_size)}")

    logger.info("\n文件列表：")
    for i, file_info in enumerate(matched_files, 1):
        if file_info['ext'] == '.png':
            file_type = "PNG图片"
        elif file_info['ext'] == '.jpg':
            file_type = "JPG图片"
        elif file_info['ext'] == '.gif':
            file_type = "GIF图片"
        else:
            file_type = "MP4视频"
        download_time_str = file_info['download_time'].strftime('%Y-%m-%d %H:%M:%S')
        logger.info(
            f"{i:3d}. {file_info['file'].name} ({file_type}, {format_size(file_info['size'])}, 下载时间: {download_time_str})")

    if dry_run:
        logger.info("\n[测试模式] 未实际删除文件")
        logger.info("清理完成（测试模式）")
        return

    logger.info("\n开始删除文件...")
    deleted_count = 0
    failed_count = 0
    deleted_size = 0

    for file_info in matched_files:
        with ExceptionContext(f"删除媒体文件 {file_info['file'].name}", default=False) as ctx:
            file_info['file'].unlink()
            deleted_count += 1
            deleted_size += file_info['size']
            if file_info['ext'] == '.png':
                file_type = "PNG图片"
            elif file_info['ext'] == '.jpg':
                file_type = "JPG图片"
            elif file_info['ext'] == '.gif':
                file_type = "GIF图片"
            else:
                file_type = "MP4视频"
            logger.info(f"已删除: {file_info['file'].name} ({file_type}, {format_size(file_info['size'])})")
            if ctx.error:
                failed_count += 1

    logger.info("\n" + "=" * 60)
    logger.info("清理完成")
    logger.info(f"成功删除: {deleted_count} 个文件，释放空间: {format_size(deleted_size)}")
    if failed_count > 0:
        logger.warning(f"删除失败: {failed_count} 个文件")
    logger.info("=" * 60)


def run_cleaner():
    """文件清理工具主函数"""
    print_separator()
    print('文件清理工具')
    print_separator()
    
    default_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"默认清理目录: {default_dir}")
    
    while True:
        print('\n请选择清理模式：')
        print('1. 按组清理（保留最新的一组文件）')
        print('2. 按时间清理（删除指定分钟前的文件）')
        print('3. 列出文件（不删除）')
        print('4. 删除所有文件和文件夹')
        print('5. 删除PNG文件')
        print('6. 删除媒体文件（PNG/JPG/GIF/MP4）')
        print('0. 返回主菜单')
        print_separator()
        
        try:
            choice = input('请输入选项 (0-6): ').strip()
        except (EOFError, KeyboardInterrupt):
            print('\n已退出清理工具')
            return
        
        if choice == '0':
            print('返回主菜单')
            return
        
        # 获取目录
        dir_input = input(f'请输入清理目录（回车使用默认目录 {default_dir}）: ').strip()
        directory = dir_input if dir_input else default_dir
        
        # 询问是否测试模式
        dry_run_input = input('是否启用测试模式（不实际删除文件）？(y/n): ').strip().lower()
        dry_run = dry_run_input in ['y', 'yes', '是']
        
        log_file = os.path.join(directory, 'clean_files.log')
        
        if choice == '1':
            clean_old_files(directory=directory, dry_run=dry_run, log_file=log_file)
        elif choice == '2':
            try:
                minutes = int(input('请输入分钟数（默认5分钟）: ').strip() or '5')
                clean_old_files_by_time(directory=directory, minutes=minutes, dry_run=dry_run, log_file=log_file)
            except ValueError:
                print('无效的分钟数，使用默认值5分钟')
                clean_old_files_by_time(directory=directory, minutes=5, dry_run=dry_run, log_file=log_file)
        elif choice == '3':
            list_files(directory=directory, log_file=log_file)
        elif choice == '4':
            clean_all_files(directory=directory, dry_run=dry_run, log_file=log_file)
        elif choice == '5':
            clean_png_files(directory=directory, dry_run=dry_run, log_file=log_file)
        elif choice == '6':
            clean_media_files(directory=directory, dry_run=dry_run, log_file=log_file)
        else:
            print('无效的选项')
        
        input('\n按回车键继续...')

def print_separator(char='=', length=60):
    """打印分隔线"""
    print(char * length)

def get_version_from_readme():
    """从 README.md 自动解析最新版本号"""
    readme_path = os.path.join(PROJECT_DIR, 'README.md')
    try:
        with open(readme_path, 'r', encoding='utf-8') as f:
            content = f.read()
        match = re.search(r'###\s+v(\d+\.\d+\.\d+)', content)
        if not match:
            match = re.search(r'##\s+v(\d+\.\d+\.\d+)', content)
        if match:
            return match.group(1)
    except Exception as e:
        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
        pass
    return "0.0.0"

def jsonify(*args, **kwargs):
    """FastAPI兼容层：模拟Flask的jsonify函数"""
    if args and isinstance(args[0], dict):
        data = args[0]
        if 'status_code' in kwargs:
            pass
        return data
    return kwargs if kwargs else (args[0] if args else {})

VERSION = get_version_from_readme()

# 统一环境检测类
class Environment:
    """统一环境检测和管理"""
    
    # 系统类型
    SYSTEM = platform.system()  # 'Windows', 'Darwin'(Mac), 'Linux'
    
    # 是否为Windows系统
    IS_WINDOWS = SYSTEM == 'Windows'
    
    # 是否为Mac系统
    IS_MAC = SYSTEM == 'Darwin'
    
    # 是否为Linux系统
    IS_LINUX = SYSTEM == 'Linux'
    
    EXE_SUFFIX = '.exe' if IS_WINDOWS else ''
    NODE_PROCESS_NAME = 'node' + EXE_SUFFIX
    HOSTC_PROCESS_NAME = 'node' + EXE_SUFFIX if IS_WINDOWS else 'hostc'
    
    @staticmethod
    def get_venv_python():
        """获取虚拟环境Python路径"""
        venv_dir = os.path.join(PROJECT_DIR, '.venv')
        scripts_dir = os.path.join(venv_dir, 'Scripts' if Environment.IS_WINDOWS else 'bin')
        python_name = os.path.basename(sys.executable) if sys.executable else ('python' + Environment.EXE_SUFFIX)
        return os.path.join(scripts_dir, python_name)
    
    @staticmethod
    def _get_playwright_browsers_dir():
        """获取Playwright浏览器缓存目录，优先读取PLAYWRIGHT_BROWSERS_PATH环境变量"""
        env_path = os.environ.get('PLAYWRIGHT_BROWSERS_PATH')
        if env_path and os.path.isdir(env_path):
            return env_path
        if Environment.IS_WINDOWS:
            base = os.environ.get('LOCALAPPDATA', '')
        elif Environment.IS_MAC:
            base = os.path.join(os.environ.get('HOME', ''), 'Library', 'Caches')
        elif Environment.IS_LINUX:
            base = os.path.join(os.environ.get('HOME', ''), '.cache')
        else:
            return None
        pw_dir = os.path.join(base, 'ms-playwright') if base else None
        return pw_dir if pw_dir and os.path.isdir(pw_dir) else None

    @staticmethod
    def _find_playwright_chromium():
        """在Playwright缓存目录中动态搜索Chromium可执行文件，不硬编码子目录名"""
        pw_dir = Environment._get_playwright_browsers_dir()
        if not pw_dir:
            return None
        base_names = ('chrome', 'chromium')
        target_names = tuple(n + Environment.EXE_SUFFIX for n in base_names) if Environment.IS_WINDOWS else base_names
        chromium_dirs = sorted(
            [d for d in os.listdir(pw_dir) if d.startswith('chromium-')],
            reverse=True
        )
        for cdir in chromium_dirs:
            root = os.path.join(pw_dir, cdir)
            for dirpath, dirnames, filenames in os.walk(root):
                for fname in filenames:
                    if fname.lower() in target_names:
                        fpath = os.path.join(dirpath, fname)
                        if os.access(fpath, os.X_OK) or Environment.IS_WINDOWS:
                            return fpath
        return None

    @staticmethod
    def _find_system_chrome():
        """搜索系统已安装的Chrome，优先读取CHROME_PATH环境变量"""
        env_path = os.environ.get('CHROME_PATH')
        if env_path and os.path.isfile(env_path):
            return env_path
        if Environment.IS_WINDOWS:
            candidates = []
            for env_var in ('PROGRAMFILES', 'PROGRAMFILES(X86)', 'LOCALAPPDATA'):
                base = os.environ.get(env_var, '')
                if base:
                    candidates.append(os.path.join(base, 'Google', 'Chrome', 'Application', 'chrome' + Environment.EXE_SUFFIX))
        elif Environment.IS_MAC:
            candidates = [
                os.path.join('/Applications', 'Google Chrome.app', 'Contents', 'MacOS', 'Google Chrome'),
            ]
        elif Environment.IS_LINUX:
            candidates = []
            linux_dir = os.environ.get('CHROME_LINUX_DIR', '/chrome-linux64')
            if linux_dir:
                candidates.append(os.path.join(linux_dir, 'chrome'))
            candidates.append(os.path.join('/usr', 'bin', 'google-chrome'))
            candidates.append(os.path.join('/usr', 'bin', 'chromium-browser'))
        else:
            return None
        for path in candidates:
            if path and os.path.isfile(path):
                return path
        return None

    @staticmethod
    def get_chrome_path():
        """获取Chrome浏览器路径，支持Windows、Mac和Linux系统
        
        优先级：
        1. Playwright内置Chromium（存在则返回None，让Playwright自己找）
        2. 系统Chrome（作为fallback）
        3. 都找不到返回None（由launch的try-except兜底自动安装）
        """
        if Environment._find_playwright_chromium():
            return None
        return Environment._find_system_chrome()
    
    @staticmethod
    def get_browser_args():
        """获取浏览器启动参数，根据系统类型返回不同的参数"""
        browser_args = ['--no-sandbox', '--disable-setuid-sandbox', '--disable-blink-features=AutomationControlled']
        
        if Environment.IS_WINDOWS:
            browser_args.append('--disable-gpu')
        elif Environment.IS_LINUX:
            browser_args.extend(['--disable-gpu', '--disable-dev-shm-usage'])
        
        return browser_args
    
    @staticmethod
    def get_user_agent():
        """获取用户代理字符串，根据系统类型返回不同的UA（动态版本号）"""
        chrome_versions = ['120.0.0.0', '121.0.0.0', '122.0.0.0', '123.0.0.0', '124.0.0.0',
                          '125.0.0.0', '126.0.0.0', '127.0.0.0', '128.0.0.0', '129.0.0.0']
        chrome_version = secrets.choice(chrome_versions)

        if Environment.IS_WINDOWS:
            return f'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version} Safari/537.36'
        elif Environment.IS_MAC:
            return f'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version} Safari/537.36'
        elif Environment.IS_LINUX:
            return f'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version} Safari/537.36'
        else:
            return f'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version} Safari/537.36'
    
    @staticmethod
    def get_system_info():
        """获取系统信息"""
        return {
            'system': Environment.SYSTEM,
            'is_windows': Environment.IS_WINDOWS,
            'is_mac': Environment.IS_MAC,
            'is_linux': Environment.IS_LINUX,
            'venv_python': Environment.get_venv_python(),
            'project_dir': PROJECT_DIR
        }
    
    @staticmethod
    def test_pip_mirror(mirror_url, timeout=None):
        """测试pip镜像源速度"""
        if timeout is None:
            timeout = TIMEOUT_CONFIG['http_request']
        try:
            start_time = time.time()
            urllib.request.urlopen(mirror_url, timeout=timeout)
            elapsed_time = time.time() - start_time
            return elapsed_time
        except (urllib.error.URLError, urllib.error.HTTPError, socket.timeout, OSError) as e:
            return None
    
    @staticmethod
    def get_fastest_pip_mirror():
        """获取最快的pip镜像源"""
        mirrors = [
            ('https://mirrors.aliyun.com/pypi/simple/', 'mirrors.aliyun.com'),
            ('https://pypi.tuna.tsinghua.edu.cn/simple/', 'pypi.tuna.tsinghua.edu.cn'),
            ('https://mirrors.cloud.tencent.com/pypi/simple/', 'mirrors.cloud.tencent.com'),
            ('https://mirrors.ustc.edu.cn/pypi/simple/', 'mirrors.ustc.edu.cn'),
            ('https://pypi.douban.com/simple/', 'pypi.douban.com')
        ]
        
        fastest_mirror = mirrors[0]
        min_time = float('inf')
        
        for mirror_url, host in mirrors:
            elapsed = Environment.test_pip_mirror(mirror_url)
            if elapsed is not None and elapsed < min_time:
                min_time = elapsed
                fastest_mirror = (mirror_url, host)
        
        return fastest_mirror
    
    @staticmethod
    def kill_process_by_name(process_name):
        """跨系统终止进程"""
        try:
            if not re.match(r'^[a-zA-Z0-9._-]+$', process_name):
                print(f"⚠️ 无效的进程名: {process_name}")
                return
            if Environment.IS_WINDOWS:
                subprocess.run(['taskkill', '/F', '/IM', process_name], capture_output=True, timeout=TIMEOUT_CONFIG['subprocess_wait'])
            else:
                subprocess.run(['pkill', '-f', process_name], capture_output=True, timeout=TIMEOUT_CONFIG['subprocess_wait'])
        except (subprocess.SubprocessError, OSError, FileNotFoundError) as e:
            print(f"⚠️ 终止进程失败: {e}")
    
    @staticmethod
    def get_default_viewport():
        """动态获取默认浏览器视口大小（根据系统屏幕分辨率）"""
        try:
            if Environment.IS_WINDOWS:
                user32 = ctypes.windll.user32
                width = user32.GetSystemMetrics(0)
                height = user32.GetSystemMetrics(1)
                return {'width': min(width, 1920), 'height': min(height - 100, 1080)}
            elif Environment.IS_MAC or Environment.IS_LINUX:
                try:
                    result = subprocess.run(['xdpyinfo'], capture_output=True, text=True, timeout=2)
                    match = re.search(r'dimensions:\s*(\d+)\s*x\s*(\d+)', result.stdout)
                    if match:
                        return {'width': min(int(match.group(1)), 1920), 'height': min(int(match.group(2)) - 100, 1080)}
                except (subprocess.SubprocessError, OSError, FileNotFoundError) as e:
                    print(f"⚠️ 获取屏幕分辨率失败: {e}")
                return {'width': 1920, 'height': 1080}
        except Exception as e:
            print(f"⚠️ 获取默认视口失败: {e}")
            return {'width': 1920, 'height': 1080}

    @staticmethod
    def check_process_running(process_name):
        """跨系统检查进程是否运行"""
        try:
            if Environment.IS_WINDOWS:
                result = subprocess.run(['tasklist', '/FI', f'IMAGENAME eq {process_name}'], capture_output=True, text=True, timeout=TIMEOUT_CONFIG['subprocess_kill'])
                return process_name in result.stdout
            else:
                result = subprocess.run(['pgrep', '-f', process_name], capture_output=True, text=True, timeout=TIMEOUT_CONFIG['subprocess_kill'])
                return result.returncode == 0
        except subprocess.TimeoutExpired as e:
            print(f"⚠️ 检查进程运行状态超时（{TIMEOUT_CONFIG['subprocess_kill']}秒）: {e}")
            return False
        except (subprocess.SubprocessError, OSError, FileNotFoundError) as e:
            print(f"⚠️ 检查进程运行状态失败: {e}")
            return False

# Windows上的emoji安全打印
def safe_print(*args, **kwargs):
    """安全打印，处理Windows上的emoji编码问题"""
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        # Windows上无法打印emoji，替换为ASCII字符
        safe_args = []
        for arg in args:
            if isinstance(arg, str):
                # 替换emoji为ASCII字符
                safe_arg = arg
                emoji_map = {
                    '[搜索]': '[检查]',
                    '❌': '[错误]',
                    '[OK]': '[OK]',
                    '⚠️': '[警告]',
                    '[FAIL]': '[失败]',
                    '📝': '[说明]',
                    '💡': '[提示]',
                    '🚀': '[启动]',
                    '🎯': '[目标]',
                    '📊': '[数据]',
                    '🔧': '[设置]',
                    '🎉': '[完成]'
                }
                for emoji, replacement in emoji_map.items():
                    safe_arg = safe_arg.replace(emoji, replacement)
                safe_args.append(safe_arg)
            else:
                safe_args.append(arg)
        print(*safe_args, **kwargs)

# 使用Environment类的VENV_PYTHON，自动创建虚拟环境
def get_python_executable():
    """获取Python可执行文件路径，优先使用虚拟环境，不存在则创建"""
    venv_python = Environment.get_venv_python()
    if os.path.exists(venv_python):
        return venv_python
    # 虚拟环境不存在，创建它
    print(f"虚拟环境不存在，正在创建...")
    try:
        subprocess.run(
            [sys.executable, '-m', 'venv', '.venv'],
            check=True,
            capture_output=True
        )
        print(f"虚拟环境创建成功: .venv")
        return venv_python
    except Exception as e:
        print(f"创建虚拟环境失败: {e}")
        # 创建失败，fallback到系统Python
        return sys.executable or ('python' + Environment.EXE_SUFFIX if Environment.IS_WINDOWS else 'python3')

VENV_PYTHON = get_python_executable()



def sec_sp(base_dir, user_path):
    """Secure path join - prevent path traversal attacks."""
    safe_base = os.path.realpath(base_dir)
    target = os.path.realpath(os.path.join(safe_base, user_path))
    if not target.startswith(safe_base + os.sep) and target != safe_base:
        return None, f"Path traversal blocked: {user_path}"
    return target, None

app = FastAPI(
    title="Szwego商品爬虫",
    description="Szwego商品爬虫Web服务",
    version="3.8.90.01",
    docs_url=None,
    redoc_url=None,
    openapi_url=None)

app.add_middleware(GZipMiddleware, minimum_size=1000)

if CORSMiddleware:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=[
            "http://localhost",
            "http://localhost:8888",
            "http://127.0.0.1",
            "http://127.0.0.1:8888",
            "http://localhost:5000",
            "http://127.0.0.1:5000",
            "http://localhost:8080",
            "http://127.0.0.1:8080",
        ],
        allow_credentials=True,
        allow_methods=["GET", "POST", "OPTIONS"],
        allow_headers=["Content-Type", "Authorization", "X-Requested-With", "X-API-Key"],
    )

# ============================================================
# CSRF 防护常量 (v3.8.89.29) - API Key 在 ConfigManager 定义后初始化
# ============================================================
LOCAL_TRUSTED_ORIGINS = frozenset([
    'http://localhost', 'http://127.0.0.1',
    'http://localhost:8888', 'http://127.0.0.1:8888',
    'http://localhost:5000', 'http://127.0.0.1:5000',
    'http://localhost:8080', 'http://127.0.0.1:8080',
])

def _is_private_ip(ip_str):
    try:
        return ipaddress.ip_address(ip_str) in (
            ipaddress.ip_network('10.0.0.0/8'),
            ipaddress.ip_network('172.16.0.0/12'),
            ipaddress.ip_network('192.168.0.0/16'),
        )
    except (ValueError, TypeError):
        return False

WRITE_METHODS = frozenset(['POST', 'PUT', 'PATCH', 'DELETE'])
CSRF_EXEMPT_PATHS = frozenset(['/api/bootstrap'])
WEB_API_KEY = None

if PROMETHEUS_AVAILABLE:
    REQUEST_COUNT = Counter('http_requests_total', '总请求数', ['method', 'endpoint', 'status'])
    REQUEST_LATENCY = Histogram('http_request_duration_seconds', '请求延迟', ['method', 'endpoint'], buckets=[0.005, 0.01, 0.025, 0.05, 0.1, 0.25, 0.5, 1.0, 2.5, 5.0])
    ACTIVE_TASKS_GAUGE = Gauge('active_tasks', '活跃任务数')
else:
    REQUEST_COUNT = None
    REQUEST_LATENCY = None
    ACTIVE_TASKS_GAUGE = None

if FLASK_RESTX_AVAILABLE:
    _api = None

def get_daily_profit_report_from_excel(excel_file):
    """从Excel的'每日利润'sheet的A列中查找以'截止'开头的报表文本
    
    Args:
        excel_file: Excel文件路径
        
    Returns:
        str: 报表文本，如果未找到则返回None
    """
    if openpyxl is None:
        return None
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_name = '每日利润'
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]
        report_text = None
        # 在A列中搜索以"截止"开头的单元格
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip().startswith('截止'):
                    report_text = cell.value.strip()
                    break
            if report_text:
                break
        wb.close()
        return report_text
    except Exception as e:
        print(f"读取每日利润报表失败: {e}")
        return None

def get_excel_files_with_report():
    """获取Excel文件列表和每日利润报表
    
    Returns:
        tuple: (excel_files_list, daily_profit_report)
    """
    excel_files_list = []
    config_file = os.path.join(PROJECT_DIR, 'config', 'config.json')
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        excel_files = config.get('excel_files', [])
        for path in excel_files:
            expanded_path = os.path.expanduser(path)
            if os.path.exists(expanded_path):
                excel_files_list.append(expanded_path)
    
    if not excel_files_list:
        excel_files_list = [os.path.join(PROJECT_DIR, 'config', '本地商品表格.xlsx')]
    
    excel_files_list = list(dict.fromkeys(os.path.abspath(f) for f in excel_files_list))
    
    daily_profit_report = None
    for excel_file in excel_files_list:
        if os.path.exists(excel_file):
            if daily_profit_report is None:
                daily_profit_report = get_daily_profit_report_from_excel(excel_file)
            break
    
    return excel_files_list, daily_profit_report

@app.exception_handler(Exception)
async def handle_api_exception(request: Request, exc: Exception):
    """FastAPI全局异常处理器 - 统一处理所有未捕获的异常"""
    error_msg = handle_exception(exc, 'FastAPI API')
    return JSONResponse(
        status_code=500,
        content={'error': error_msg, 'success': False, 'code': getattr(exc, 'code', 'UNKNOWN')}
    )


# ============================================================
# API速率限制器 (v3.8.70)
# ============================================================
class RateLimiter:
    """IP级别速率限制器"""

    def __init__(self, max_requests=100, window_seconds=60):
        self.max_requests = max_requests
        self.window_seconds = window_seconds
        self.requests = {}
        self._lock = threading.Lock()

    def is_allowed(self, client_ip):
        """检查是否允许请求"""
        current_time = time.time()
        with self._lock:
            if client_ip not in self.requests:
                self.requests[client_ip] = []
            self.requests[client_ip] = [
                t for t in self.requests[client_ip]
                if current_time - t < self.window_seconds
            ]
            if len(self.requests[client_ip]) >= self.max_requests:
                return False
            self.requests[client_ip].append(current_time)
            return True

    def get_retry_after(self, client_ip):
        """获取重试等待时间"""
        with self._lock:
            if client_ip not in self.requests or not self.requests[client_ip]:
                return 0
            oldest = min(self.requests[client_ip])
            return max(0, int(self.window_seconds - (time.time() - oldest)) + 1)

api_rate_limiter = RateLimiter(max_requests=200, window_seconds=60)
upload_rate_limiter = RateLimiter(max_requests=10, window_seconds=60)


# ============================================================
# Pydantic输入验证 (v3.8.71)
# ============================================================
class RunCommandRequest(BaseModel):
    command: str = Field(..., min_length=1, max_length=10000)

    @field_validator('command')
    def validate_command_safe(cls, v):
        dangerous = [
            'rm -rf /', 'rm -rf /*', 'rm -rf ~', 'rm -rf *',
            'mkfs', 'shutdown', 'reboot', 'halt', 'poweroff',
            'format', 'del /f /q C:\\', 'del /f /s /q',
            'dd if=', 'dd of=/dev/sd', 'dd of=/dev/hd',
            '> /dev/sd', '> /dev/hd',
            'chmod -R 777 /', 'chown -R',
            ':(){ :|:& };:', 'fork bomb',
            'wget http', 'curl http', 'chmod +x',
            'nc -l', 'nc -e', 'bash -i', 'sh -i',
            'python -c', 'perl -e', 'ruby -e',
            'eval ', 'exec ', 'source /dev/tcp',
            'crontab -r', 'systemctl stop', 'systemctl disable',
            'iptables -F', 'ufw disable',
            'userdel', 'usermod -L', 'passwd -d',
            'netsh advfirewall set allprofiles state off',
            'reg delete', 'reg add',
        ]
        v_lower = v.lower()
        for p in dangerous:
            if p.lower() in v_lower:
                raise ValueError(f'检测到危险命令: {p}')
        return v.strip()


class TaskInputRequest(BaseModel):
    task_id: str = Field(..., min_length=1, max_length=50)
    user_input: str = Field('', max_length=10000)


class KillTaskRequest(BaseModel):
    task_id: str = Field(..., min_length=1, max_length=50)


class SKUCompareRequest(BaseModel):
    skus: str = Field(None, max_length=50000)


def validate_request(model_class, data):
    if not PYDANTIC_AVAILABLE:
        if not data:
            return None, '请求体不能为空'
        return data, None
    try:
        validated = model_class(**data)
        return validated.dict() if hasattr(validated, 'dict') else validated.model_dump(), None
    except ValidationError as e:
        errors = e.errors() if hasattr(e, 'errors') else []
        msg = '; '.join([f"{err.get('loc', ('?',))[0]}: {err.get('msg', '')}" for err in errors])
        return None, f'输入验证失败: {msg}'


def rate_limit(limiter, endpoint_name='API'):
    """速率限制装饰器"""
    def decorator(f):
        @wraps(f)
        async def decorated(request: Request, *args, **kwargs):
            client_ip = request.client.host if request.client else 'unknown'
            if not limiter.is_allowed(client_ip):
                retry_after = limiter.get_retry_after(client_ip)
                return JSONResponse(
                    status_code=429,
                    content={
                        'error': '请求过于频繁，请稍后再试',
                        'code': 'RATE_LIMIT_EXCEEDED',
                        'retry_after': retry_after,
                        'endpoint': endpoint_name,
                    },
                    headers={'Retry-After': str(retry_after)}
                )
            
            if asyncio.iscoroutinefunction(f):
                return await f(*args, **kwargs)
            else:
                return f(*args, **kwargs)
        return decorated
    return decorator


def safe_read_json(file_path, default=None):
    """安全读取JSON文件，解析失败返回默认值"""
    if default is None:
        default = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default
    except Exception:
        return default


# ============================================================
# JSON文件缓存管理器 (v3.8.70)
# ============================================================
class FileCacheManager:
    """JSON文件缓存管理器"""

    def __init__(self, ttl_seconds=30):
        self._cache = {}
        self._ttl = ttl_seconds
        self._lock = threading.Lock()

    def read_json(self, file_path, default=None):
        if default is None:
            default = {}
        current_time = time.time()
        with self._lock:
            if file_path in self._cache:
                cached_data, cache_time = self._cache[file_path]
                if current_time - cache_time < self._ttl:
                    if os.path.exists(file_path):
                        if os.path.getmtime(file_path) <= cache_time:
                            return cached_data
                    del self._cache[file_path]
        data = safe_read_json(file_path, default)
        with self._lock:
            self._cache[file_path] = (data, current_time)
        return data

    def invalidate(self, file_path=None):
        with self._lock:
            if file_path:
                self._cache.pop(file_path, None)
            else:
                self._cache.clear()

    def get_stats(self):
        with self._lock:
            return {'cached_files': len(self._cache), 'files': list(self._cache.keys())}

json_cache = FileCacheManager(ttl_seconds=30)


processes = {}
tasks = {}
_processes_lock = threading.Lock()
_tasks_lock = threading.Lock()

def run_command_background(task_id, command):
    try:
        with _tasks_lock:
            tasks[task_id]['status'] = 'running'
        
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        
        if isinstance(command, str):
            cmd_list = shlex.split(command)
        else:
            cmd_list = list(command)
        
        process = subprocess.Popen(
            cmd_list,
            shell=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            stdin=subprocess.DEVNULL,
            cwd=PROJECT_DIR,
            text=True,
            encoding='utf-8',
            errors='replace',
            bufsize=1,
            env=env
        )
        
        with _processes_lock:
            processes[task_id] = process
        
        stdout_lines = []
        while True:
            if process.poll() is not None:
                remaining = process.stdout.read()
                if remaining:
                    stdout_lines.append(remaining)
                break
            
            try:
                if Environment.IS_WINDOWS:
                    time.sleep(0.1)
                    line = process.stdout.readline()
                    if line:
                        stdout_lines.append(line)
                else:
                    readable, _, _ = select.select([process.stdout], [], [], 0.1)
                    if readable:
                        line = process.stdout.readline()
                        if line:
                            stdout_lines.append(line)
            except Exception as e:
                handle_exception(e, 'run_command_background读取输出')
            
            with _tasks_lock:
                tasks[task_id]['output'] = ''.join(stdout_lines)
        
        process.wait()
        with _tasks_lock:
            tasks[task_id]['returncode'] = process.returncode
            tasks[task_id]['output'] = ''.join(stdout_lines)
            tasks[task_id]['status'] = 'completed'
    except Exception as e:
        handle_exception(e, 'run_command_background')
        with _tasks_lock:
            tasks[task_id]['error'] = str(e)
            tasks[task_id]['status'] = 'error'


class PathManager:
    """路径管理类，统一处理跨系统路径问题"""
    
    @staticmethod
    def get_config_dir():
        """获取配置文件目录"""
        return os.path.join(PROJECT_DIR, 'config')
    
    @staticmethod
    def get_file_dir():
        """获取输出文件目录"""
        return os.path.join(PROJECT_DIR, 'file')
    
    @staticmethod
    def get_config_file():
        """获取配置文件路径"""
        return os.path.join(PathManager.get_config_dir(), 'config.json')
    
    @staticmethod
    def get_cookie_file():
        """获取Cookie文件路径"""
        return os.path.join(PathManager.get_config_dir(), 'cookies.json')
    
    @staticmethod
    def get_output_file():
        """获取输出文件路径"""
        return os.path.join(PathManager.get_file_dir(), 'output.json')
    
    @staticmethod
    def get_input_file():
        """获取输入文件路径"""
        return os.path.join(PathManager.get_config_dir(), 'input_stock_numbers.txt')
    
    @staticmethod
    def get_json_filename(date_str):
        """获取JSON文件名"""
        return f"{date_str}微购相册(小旭数码).json"
    
    @staticmethod
    def get_cache_filename(date_str):
        """获取缓存文件名"""
        return f"{date_str}微购相册(小旭数码)_cache.json"
    
    @staticmethod
    def get_json_file_path(date_str):
        """获取JSON文件完整路径"""
        return os.path.join(PathManager.get_file_dir(), PathManager.get_json_filename(date_str))
    
    @staticmethod
    def get_cache_file_path(date_str):
        """获取缓存文件完整路径"""
        return os.path.join(PathManager.get_file_dir(), PathManager.get_cache_filename(date_str))
    
    @staticmethod
    def get_diff_log_file(date_str):
        """获取差异日志文件路径"""
        return os.path.join(PathManager.get_file_dir(), f'diff_log_{date_str}.json')
    
    @staticmethod
    def get_duplicate_log_file():
        """获取重复日志文件路径"""
        return os.path.join(PathManager.get_file_dir(), 'duplicate_log.json')
    
    @staticmethod
    def get_tunnel_url_file():
        """获取隧道URL文件路径"""
        return os.path.join(PathManager.get_file_dir(), 'tunnel_url.txt')
    
    @staticmethod
    def get_web_output_file():
        """获取Web输出日志文件路径"""
        return os.path.join(PathManager.get_file_dir(), 'web_output.log')
    _url_source_config = {
        'primary_source': 'tunnel_url.txt',
        'enable_logging': True,
        'enable_health_check': True,
        'auto_sync_interval': 300,
        'validate_url': True,
        'url_validation_timeout': 5,
        'skip_validation': False
    }
    
    _last_url_source_log = {}
    _url_health_check_time = 0
    
    @staticmethod
    def get_public_url_from_web_log(skip_validation=False, quiet=False):
        """获取公网地址（统一入口） - 以 tunnel_url.txt 为权威源
        
        数据流向：
        hostc → tunnel_url.txt (权威源) → web_output.log (镜像) → 前端显示
        
        策略：
        1. 优先从 tunnel_url.txt 读取（权威源）
        2. 如果 tunnel_url.txt 的 URL 不可用，尝试 web_output.log
        3. 两个都失败则返回 None
        
        Args:
            skip_validation: 跳过URL可用性验证（调用方会自行验证时使用，避免双重验证）
            quiet: 静默模式，减少日志输出（心跳循环等高频调用时使用）
        """
        
        config = PathManager._url_source_config
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result_url = None
        url_source = None
        should_log = config['enable_logging'] and not quiet
        
        if should_log:
            print(f"[{current_time}] [URL-Source] [搜索] 开始获取公网地址...")
        
        # ========== 策略1：从 tunnel_url.txt 读取（权威源）==========
        try:
            tunnel_file = PathManager.get_tunnel_url_file()
            
            if should_log:
                print(f"[{current_time}] [URL-Source] 📂 尝试读取: {tunnel_file}")
            
            if os.path.exists(tunnel_file):
                with open(tunnel_file, 'r', encoding='utf-8') as f:
                    tunnel_content = f.read()
                
                if should_log:
                    print(f"[{current_time}] [URL-Source] 📄 文件大小: {len(tunnel_content)} 字符")
                
                # 新格式：hostc: https://xxx.hostc.dev
                hostc_match = re.search(r'^hostc:\s*(https://[^\s]+)', tunnel_content, re.MULTILINE)
                cf_match = re.search(r'^cloudflare:\s*(https://[^\s]+)', tunnel_content, re.MULTILINE)
                
                # 旧格式兼容：Public URL: https://...
                tunnel_match = re.search(r'Public URL:\s*(https://[^\s]+)', tunnel_content)
                if not tunnel_match:
                    # 回退：匹配任意 hostc.dev URL
                    tunnel_match = re.search(r'(https://[a-zA-Z0-9_-]+\.hostc\.dev)', tunnel_content)
                
                # 优先使用新格式的 hostc URL，其次使用旧格式
                if hostc_match:
                    candidate_url = hostc_match.group(1).rstrip('/')
                elif tunnel_match:
                    candidate_url = tunnel_match.group(1).rstrip('/')
                else:
                    candidate_url = None
                
                if candidate_url:
                    
                    if should_log:
                        print(f"[{current_time}] [URL-Source] ✅ 从 tunnel_url.txt 提取到候选URL: {candidate_url}")
                    
                    should_validate = config['validate_url'] and candidate_url and not skip_validation
                    
                    if should_validate:
                        is_valid = PathManager._validate_url_accessibility(candidate_url, config['url_validation_timeout'])
                        if is_valid:
                            result_url = candidate_url
                            url_source = 'tunnel_url.txt (validated)'
                            
                            if should_log:
                                print(f"[{current_time}] [URL-Source] ✅✅✅ URL验证通过！来源: tunnel_url.txt")
                                print(f"[{current_time}] [URL-Source] 🎯 最终URL: {result_url}")
                        else:
                            if should_log:
                                print(f"[{current_time}] [URL-Source] ⚠️ tunnel_url.txt 中的URL不可用，尝试备用源...")
                    else:
                        result_url = candidate_url
                        url_source = 'tunnel_url.txt' + (' (skip_validation)' if skip_validation else ' (no validation)')
                        
                        if should_log:
                            print(f"[{current_time}] [URL-Source] ✅ 跳过验证，直接使用URL")
                            print(f"[{current_time}] [URL-Source] 🎯 最终URL: {result_url}")
                else:
                    if should_log:
                        print(f"[{current_time}] [URL-Source] ❌ tunnel_url.txt 中未找到有效URL格式")
            else:
                if should_log:
                    print(f"[{current_time}] [URL-Source] ⚠️ tunnel_url.txt 文件不存在")
                    
        except Exception as e:
            if should_log:
                print(f"[{current_time}] [URL-Source] ❌ 读取 tunnel_url.txt 失败: {str(e)[:100]}")
        
        # ========== 策略2：从 web_output.log 读取（备用方案）==========
        if not result_url:
            try:
                web_log_file = PathManager.get_web_output_file()
                
                if should_log:
                    print(f"[{current_time}] [URL-Source] 📂 尝试备用源: {web_log_file}")
                
                if os.path.exists(web_log_file):
                    with open(web_log_file, 'r', encoding='utf-8', errors='replace') as f:
                        content = f.read()
                    
                    # 匹配 "Public URL: https://..." 格式
                    matches = re.findall(r'Public URL:\s*(https?://[^\s]+)', content)
                    if matches:
                        candidate_url = matches[-1].rstrip('/')
                        
                        if should_log:
                            print(f"[{current_time}] [URL-Source] 📋 从 web_output.log 提取到候选URL: {candidate_url}")
                        
                        should_validate = config['validate_url'] and candidate_url and not skip_validation
                        
                        if should_validate:
                            is_valid = PathManager._validate_url_accessibility(candidate_url, config['url_validation_timeout'])
                            if is_valid:
                                result_url = candidate_url
                                url_source = 'web_output.log (validated)'
                                
                                if should_log:
                                    print(f"[{current_time}] [URL-Source] ✅ 备用源URL验证通过！")
                                    print(f"[{current_time}] [URL-Source] 🎯 最终URL: {result_url}")
                                    print(f"[{current_time}] [URL-Source] 💡 建议: 应将此URL同步到 tunnel_url.txt")
                                    
                                PathManager._sync_url_to_tunnel_file(result_url)
                            else:
                                if should_log:
                                    print(f"[{current_time}] [URL-Source] ⚠️ 备用源URL也不可用")
                        else:
                            result_url = candidate_url
                            url_source = 'web_output.log' + (' (skip_validation)' if skip_validation else ' (no validation)')
                    else:
                        # 回退：匹配任意 hostc.dev URL
                        matches = re.findall(r'(https://[a-zA-Z0-9_-]+\.hostc\.dev)', content)
                        if matches:
                            candidate_url = matches[-1].rstrip('/')
                            
                            should_validate = config['validate_url'] and candidate_url and not skip_validation
                            
                            if should_validate:
                                is_valid = PathManager._validate_url_accessibility(candidate_url, config['url_validation_timeout'])
                                if is_valid:
                                    result_url = candidate_url
                                    url_source = 'web_output.log.fallback (validated)'
                                    
                                    if should_log:
                                        print(f"[{current_time}] [URL-Source] ✅ 回退匹配成功并验证通过")
                                        print(f"[{current_time}] [URL-Source] 🎯 最终URL: {result_url}")
                                        
                                    PathManager._sync_url_to_tunnel_file(result_url)
                else:
                    if should_log:
                        print(f"[{current_time}] [URL-Source] ⚠️ web_output.log 文件不存在")
                        
            except Exception as e:
                if should_log:
                    print(f"[{current_time}] [URL-Source] ❌ 读取 web_output.log 失败: {str(e)[:100]}")
        
        # ========== 记录日志 ==========
        if should_log:
            PathManager._last_url_source_log = {
                'timestamp': current_time,
                'url': result_url,
                'source': url_source,
                'success': result_url is not None
            }
            
            if result_url:
                print(f"[{current_time}] [URL-Source] 🎉 获取成功！来源: {url_source}")
            else:
                print(f"[{current_time}] [URL-Source] ❌ 所有数据源均无法提供有效URL")
        
        return result_url
    
    _url_validation_cache = {}
    _cache_expiry_seconds = 60
    
    @staticmethod
    def _validate_url_accessibility(url, timeout=None, max_retries=2):
        """验证URL是否可访问（增强版）
        
        改进点：
        1. 多种验证方式（GET + HEAD + TCP连接）
        2. 自动重试机制
        3. 缓存机制避免频繁验证
        4. 更详细的错误分类
        """
        if timeout is None:
            timeout = TIMEOUT_CONFIG['http_request']
        
        current_time = time.time()
        cache_key = url
        
        # 检查缓存（避免短时间内重复验证）
        if cache_key in PathManager._url_validation_cache:
            cached_result, cached_time = PathManager._url_validation_cache[cache_key]
            if current_time - cached_time < PathManager._cache_expiry_seconds:
                if PathManager._url_source_config.get('enable_logging'):
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [URL-Validate] 📦 使用缓存结果: {cached_result} (剩余{int(PathManager._cache_expiry_seconds - (current_time - cached_time))}秒)")
                return cached_result
        
        validation_methods = [
            ('GET', PathManager._validate_with_get),
            ('HEAD', PathManager._validate_with_head),
            ('TCP', PathManager._validate_with_tcp)
        ]
        
        last_error = None
        
        for attempt in range(max_retries + 1):
            for method_name, method_func in validation_methods:
                try:
                    is_valid, error_msg = method_func(url, timeout)
                    
                    if is_valid:
                        # 缓存成功结果
                        PathManager._url_validation_cache[cache_key] = (True, current_time)
                        
                        if PathManager._url_source_config.get('enable_logging'):
                            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            print(f"[{log_time}] [URL-Validate] ✅✅✅ URL验证成功!")
                            print(f"[{log_time}] [URL-Validate]   方法: {method_name}")
                            print(f"[{log_time}] [URL-Validate]   URL: {url}")
                            if attempt > 0:
                                print(f"[{log_time}] [URL-Validate]   重试次数: {attempt}")
                        
                        return True
                    else:
                        last_error = f"{method_name}: {error_msg}"
                        
                except Exception as e:
                    last_error = f"{method_name} 异常: {str(e)[:80]}"
            
            # 重试前等待一小段时间
            if attempt < max_retries:
                time.sleep(0.5)
        
        # 所有方法都失败，缓存失败结果（较短时间）
        PathManager._url_validation_cache[cache_key] = (False, current_time)
        
        if PathManager._url_source_config.get('enable_logging'):
            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{log_time}] [URL-Validate] ❌ URL验证失败")
            print(f"[{log_time}] [URL-Validate]   URL: {url}")
            print(f"[{log_time}] [URL-Validate]   最后错误: {last_error}")
            print(f"[{log_time}] [URL-Validate]   重试次数: {max_retries}")
            print(f"[{log_time}] [URL-Validate]   💡 提示: URL可能暂时不可用或网络波动")
        
        return False
    
    @staticmethod
    def _validate_with_get(url, timeout):
        """使用GET请求验证"""
        try:
            req = urllib.request.Request(url, method='GET')
            req.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
            req.add_header('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8')
            req.add_header('Accept-Language', 'zh-CN,zh;q=0.9,en;q=0.8')
            
            response = urllib.request.urlopen(req, timeout=timeout)
            
            if response.status in [200, 301, 302, 303, 307, 308]:
                return (True, None)
            else:
                return (False, f"HTTP状态码: {response.status}")
                
        except urllib.error.HTTPError as e:
            if e.code in [401, 403, 404, 405]:
                return (True, f"HTTP {e.code} (服务存在但受限)")  # 服务存在只是需要认证等
            return (False, f"HTTP错误: {e.code}")
        except urllib.error.URLError as e:
            return (False, f"连接错误: {str(e.reason)[:50]}")
        except socket.timeout:
            return (False, "连接超时")
        except Exception as e:
            return (False, str(e)[:80])
    
    @staticmethod
    def _validate_with_head(url, timeout):
        """使用HEAD请求验证"""
        try:
            req = urllib.request.Request(url, method='HEAD')
            req.add_header('User-Agent', 'Mozilla/5.0 (compatible; URLCheck/1.0)')
            
            response = urllib.request.urlopen(req, timeout=timeout)
            
            if response.status in [200, 301, 302, 303, 307, 308]:
                return (True, None)
            else:
                return (False, f"HTTP状态码: {response.status}")
                
        except urllib.error.HTTPError as e:
            if e.code in [401, 403, 404, 405]:
                return (True, f"HTTP {e.code}")
            return (False, f"HTTP错误: {e.code}")
        except Exception as e:
            return (False, str(e)[:80])
    
    @staticmethod
    def _validate_with_tcp(url, timeout):
        """使用TCP连接验证（最底层）"""
        sock = None
        try:
            parsed = urllib.parse.urlparse(url)
            hostname = parsed.hostname
            port = parsed.port or (443 if parsed.scheme == 'https' else 80)
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            
            result = sock.connect_ex((hostname, port))
            
            if result == 0:
                return (True, None)
            else:
                return (False, f"TCP连接失败 (错误码: {result})")
                
        except socket.gaierror:
            return (False, "DNS解析失败")
        except socket.timeout:
            return (False, "TCP连接超时")
        except Exception as e:
            return (False, str(e)[:80])
        finally:
            if sock:
                try:
                    sock.close()
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
    
    @staticmethod
    def _sync_url_to_tunnel_file(url):
        """将有效的URL同步到 tunnel_url.txt（确保权威源始终有最新可用URL）"""
        try:
            tunnel_file = PathManager.get_tunnel_url_file()
            
            # 读取现有内容
            existing_content = ""
            if os.path.exists(tunnel_file):
                with open(tunnel_file, 'r', encoding='utf-8') as f:
                    existing_content = f.read()
            
            # 检查是否需要更新
            if re.search(re.escape(url), existing_content):
                return False  # 已存在，无需更新
            
            # 更新或创建文件
            with open(tunnel_file, 'w', encoding='utf-8') as f:
                port = 8888
                tunnel_name = url.split('//')[1].split('.')[0] if '//' in url else 'unknown'
                
                f.write(f"Success  Tunnel ready\n")
                f.write(f"  Public URL: {url}\n")
                f.write(f"  Local:      http://localhost:{port}/\n")
                f.write(f"  Tunnel:     {tunnel_name}\n")
                f.write(f"  Channels:   2\n")
                f.write(f"\n# Auto-synced at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{current_time}] [URL-Sync] ✅ 已将URL同步到 tunnel_url.txt: {url}")
            return True
            
        except Exception as e:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{current_time}] [URL-Sync] ❌ 同步失败: {str(e)[:100]}")
            return False
    
    @staticmethod
    def check_url_files_health():
        """检查两个URL文件的健康状态和一致性"""
        
        config = PathManager._url_source_config
        current_time = time.time()
        
        # 控制检查频率（默认每5分钟检查一次）
        if config['enable_health_check']:
            if current_time - PathManager._url_health_check_time < config.get('auto_sync_interval', 300):
                return {'status': 'skipped', 'reason': '检查间隔未到'}
            
            PathManager._url_health_check_time = current_time
        
        health_result = {
            'timestamp': datetime.now().isoformat(),
            'tunnel_file': {'exists': False, 'has_url': False, 'url': None, 'valid': False},
            'weblog_file': {'exists': False, 'has_url': False, 'url': None, 'valid': False},
            'consistent': False,
            'action_taken': None
        }
        
        try:
            # 检查 tunnel_url.txt
            tunnel_file = PathManager.get_tunnel_url_file()
            if os.path.exists(tunnel_file):
                health_result['tunnel_file']['exists'] = True
                with open(tunnel_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                match = re.search(r'Public URL:\s*(https://[^\s]+)', content)
                if match:
                    url = match.group(1).rstrip('/')
                    health_result['tunnel_file']['has_url'] = True
                    health_result['tunnel_file']['url'] = url
                    health_result['tunnel_file']['valid'] = PathManager._validate_url_accessibility(url)
            
            # 检查 web_output.log
            weblog_file = PathManager.get_web_output_file()
            if os.path.exists(weblog_file):
                health_result['weblog_file']['exists'] = True
                with open(weblog_file, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
                
                matches = re.findall(r'Public URL:\s*(https?://[^\s]+)', content)
                if matches:
                    url = matches[-1].rstrip('/')
                    health_result['weblog_file']['has_url'] = True
                    health_result['weblog_file']['url'] = url
                    health_result['weblog_file']['valid'] = PathManager._validate_url_accessibility(url)
            
            # 检查一致性
            tunnel_url = health_result['tunnel_file']['url']
            weblog_url = health_result['weblog_file']['url']
            
            if tunnel_url and weblog_url:
                health_result['consistent'] = (tunnel_url == weblog_url)
            
            # 自动修复不一致
            if not health_result['consistent']:
                if health_result['tunnel_file']['valid'] and not health_result['weblog_file']['valid']:
                    # tunnel_url.txt 有效但 web_output.log 无效或不同步
                    PathManager._sync_url_to_weblog(tunnel_url)
                    health_result['action_taken'] = f'synced_to_weblog:{tunnel_url}'
                elif health_result['weblog_file']['valid'] and not health_result['tunnel_file']['valid']:
                    # web_output.log 有效但 tunnel_url.txt 无效
                    PathManager._sync_url_to_tunnel_file(weblog_url)
                    health_result['action_taken'] = f'synced_to_tunnel:{weblog_url}'
            
            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if config['enable_logging']:
                print(f"[{log_time}] [URL-Health] 📊 文件健康检查完成:")
                print(f"[{log_time}] [URL-Health]   tunnel_url.txt: {'✅' if health_result['tunnel_file']['valid'] else '❌'} {health_result['tunnel_file']['url'] or '无'}")
                print(f"[{log_time}] [URL-Health]   web_output.log: {'✅' if health_result['weblog_file']['valid'] else '❌'} {health_result['weblog_file']['url'] or '无'}")
                print(f"[{log_time}] [URL-Health]   一致性: {'✅' if health_result['consistent'] else '⚠️ 不一致'}")
                if health_result['action_taken']:
                    print(f"[{log_time}] [URL-Health] 🔧 执行操作: {health_result['action_taken']}")
            
        except Exception as e:
            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if config['enable_logging']:
                print(f"[{log_time}] [URL-Health] ❌ 健康检查异常: {str(e)[:100]}")
        
        return health_result
    
    @staticmethod
    def _sync_url_to_weblog(url):
        """将URL同步到 web_output.log"""
        try:
            weblog_file = PathManager.get_web_output_file()
            
            # 读取现有内容
            existing_lines = []
            if os.path.exists(weblog_file):
                with open(weblog_file, 'r', encoding='utf-8', errors='replace') as f:
                    existing_lines = f.readlines()
            
            # 查找是否有 Public URL 行
            updated = False
            new_lines = []
            for line in existing_lines:
                if re.match(r'.*Public URL:.*', line):
                    new_lines.append(f"  Public URL: {url}\n")
                    updated = True
                else:
                    new_lines.append(line)
            
            if not updated:
                # 在文件末尾添加
                new_lines.append(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Auto-Sync] Public URL: {url}\n")
            
            with open(weblog_file, 'w', encoding='utf-8') as f:
                f.writelines(new_lines)
            
            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{log_time}] [URL-Sync] ✅ 已将URL同步到 web_output.log: {url}")
            return True
            
        except Exception as e:
            log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{log_time}] [URL-Sync] ❌ 同步到weblog失败: {str(e)[:100]}")
            return False
    
    @staticmethod
    def configure_url_source(**kwargs):
        """配置URL获取策略
        
        参数:
            primary_source: 主数据源 ('tunnel_url.txt' 或 'web_output.log')
            fallback_source: 备用数据源
            enable_logging: 是否启用日志 (True/False)
            enable_health_check: 是否启用健康检查 (True/False)
            auto_sync_interval: 自动同步间隔（秒）
            validate_url: 是否验证URL可用性 (True/False)
            url_validation_timeout: URL验证超时时间（秒）
        """
        for key, value in kwargs.items():
            if key in PathManager._url_source_config:
                old_value = PathManager._url_source_config[key]
                PathManager._url_source_config[key] = value
                
                log_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print(f"[{log_time}] [URL-Config] ⚙️ 配置更新: {key} = {value} (原值: {old_value})")
        
        return PathManager._url_source_config.copy()
    
    @staticmethod
    def get_url_source_status():
        """获取当前URL获取状态和配置"""
        return {
            'config': PathManager._url_source_config.copy(),
            'last_log': PathManager._last_url_source_log,
            'last_health_check': datetime.fromtimestamp(PathManager._url_health_check_time).isoformat() if PathManager._url_health_check_time > 0 else None
        }
    
    @staticmethod
    def get_lan_ip():
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(2)
            s.connect((os.environ.get('LAN_IP_DETECT_HOST', '8.8.8.8'), int(os.environ.get('LAN_IP_DETECT_PORT', '80'))))
            ip = s.getsockname()[0]
            return ip
        except (socket.error, OSError, ValueError, TypeError) as e:
            print(f"⚠️ 获取局域网IP失败: {e}")
            return ''
        finally:
            if s:
                try:
                    s.close()
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass

    @staticmethod
    def sync_web_output_from_tunnel_url():
        """从 tunnel_url.txt 同步公网地址到 web_output.log（统一入口）"""
        try:
            tunnel_file = PathManager.get_tunnel_url_file()
            web_log_file = PathManager.get_web_output_file()
            if os.path.exists(tunnel_file):
                with open(tunnel_file, 'r', encoding='utf-8') as f:
                    tunnel_content = f.read()
                
                tunnel_match = re.search(r'Public URL:\s*(https://[^\s]+)', tunnel_content)
                if not tunnel_match:
                    # 如果 tunnel_url.txt 没有 Public URL，尝试直接匹配 hostc.dev URL
                    tunnel_match = re.search(r'(https://[a-zA-Z0-9_-]+\.hostc\.dev)', tunnel_content)
                if tunnel_match:
                    new_url = tunnel_match.group(1)
                    
                    try:
                        if os.path.exists(web_log_file):
                            with open(web_log_file, 'r', encoding='utf-8') as f:
                                lines = f.readlines()
                            
                            updated = False
                            for i, line in enumerate(lines):
                                if 'Public URL:' in line and 'hostc.dev' in line:
                                    lines[i] = f"  Public URL: {new_url}\n"
                                    updated = True
                            
                            if updated:
                                with open(web_log_file, 'w', encoding='utf-8') as f:
                                    f.writelines(lines)
                                return True
                    except Exception as e:
                        print(f"[Tunnel] 更新 web_output.log 失败: {e}")
                    
                    try:
                        port = args.port if 'args' in dir() and hasattr(args, 'port') else 8888
                        lan_ip = PathManager.get_lan_ip()
                        header = f"""==================================================
Szwego商品爬虫 - Web服务
==================================================
访问地址: http://localhost:{port}
"""
                        if lan_ip:
                            header += f"局域网地址: http://{lan_ip}:{port}\n"
                        with open(web_log_file, 'w', encoding='utf-8') as f:
                            f.write(header)
                            f.write(f"  Public URL: {new_url}\n")
                    except Exception as e2:
                        handle_exception(e2, 'sync_web_output_from_tunnel_url重建web_output')
                        return False
                    return True
        except Exception as e:
            handle_exception(e, 'sync_web_output_from_tunnel_url同步')
        return False
    
    @staticmethod
    def ensure_dirs_exist():
        """确保所有必要的目录存在"""
        dirs = [PathManager.get_config_dir(), PathManager.get_file_dir()]
        for dir_path in dirs:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)


class EmailNotifier:
    """邮件通知类"""
    
    def __init__(self, config_manager=None):
        self.config_manager = config_manager or ConfigManager()
    
    def get_email_config(self):
        """获取邮件配置"""
        return {
            'enabled': self.config_manager.get('email_notification_enabled', False),
            'smtp_host': self.config_manager.get('email_smtp_host', 'smtp.qq.com'),
            'smtp_port': self.config_manager.get('email_smtp_port', 587),
            'smtp_user': self.config_manager.get('email_smtp_user', ''),
            'smtp_password': self.config_manager.get('email_smtp_password', ''),
            'from_name': self.config_manager.get('email_from_name', '公网IP监控'),
            'to_email': self.config_manager.get('email_to', '980187223@qq.com'),
        }
    
    def send_tunnel_notification(self, tunnel_url, event_type='new'):
        """发送隧道URL变化通知邮件"""
        _current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _thread_id = threading.current_thread().name
        
        config = self.get_email_config()
        
        if not config['enabled']:
            print(f"[{_current_time}] [EmailNotifier-Thread:{_thread_id}] ⚠️ 邮件通知未启用，跳过发送")
            return False
        
        try:
            print(f"[{_current_time}] [EmailNotifier-Thread:{_thread_id}] 🖥️ SMTP服务器: {config['smtp_host']}:{config['smtp_port']}")
            print(f"[{_current_time}] [EmailNotifier-Thread:{_thread_id}] 👤 发送人: {config['smtp_user']}")
            print(f"[{_current_time}] [EmailNotifier-Thread:{_thread_id}] 📬 接收人: {config['to_email']}")
            
            msg = MIMEMultipart('alternative')
            msg['From'] = formataddr((config['from_name'], config['smtp_user']))
            msg['To'] = config['to_email']
            event_titles = {
                'new': '✅ 新公网地址',
                'available': '✅ 公网地址可用',
                'update': '✅ 公网地址已更新',
                'stable_available': '✅ 公网地址已稳定可用',
                'fallback_available': '🔄 备用公网地址可用',
                'unavailable': '🚨 公网地址不可用',
                'restarted': '🔄 隧道已重启'
            }
            event_title = event_titles.get(event_type, f'{"✅ 新" if event_type == "new" else "✅"}公网地址')
            
            msg['Subject'] = f'【{event_title}】{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            status_note = ""
            html_status_note = ""
            
            if event_type == 'stable_available':
                try:
                    _verify_dur = int(time.time() - globals().get('url_first_seen_time', 0)) if globals().get('url_first_seen_time', 0) > 0 else 0
                    _min_confirms = globals().get('stable_url_min_confirms', 3)
                    _confirm_count = globals().get('stable_url_confirm_count', 0)
                except (NameError, TypeError):
                    _verify_dur = 0
                    _min_confirms = 3
                    _confirm_count = 0
                verify_duration = _verify_dur
                status_note = f"""
✅ 稳定性验证通过
验证次数: {_min_confirms}次连续通过
验证耗时: {verify_duration}秒
当前状态: 🎯 确认稳定可用

"""
                html_status_note = f'''
<div style="background-color: #e8f5e9; border-left: 4px solid #4caf50; padding: 15px; margin: 20px 0; border-radius: 4px;">
<div style="color: #2e7d32; font-size: 16px; font-weight: bold; margin-bottom: 10px;">✅ 稳定性验证通过</div>
<table style="width: 100%; color: #333;">
<tr><td style="padding: 3px 0;"><strong>验证次数:</strong></td><td>{_min_confirms} 次连续通过</td></tr>
<tr><td style="padding: 3px 0;"><strong>验证耗时:</strong></td><td>{verify_duration} 秒</td></tr>
<tr><td style="padding: 3px 0;"><strong>当前状态:</strong></td><td style="color: #2e7d32; font-weight: bold;">🎯 确认稳定可用</td></tr>
</table>
</div>
'''
            elif event_type == 'fallback_available':
                status_note = f"""
🔄 原隧道不可用，已切换到备用地址
当前可用地址: {tunnel_url}
当前状态: ✅ 备用地址已验证可用，可放心使用

"""
                html_status_note = f'''
<div style="background-color: #fff3e0; border-left: 4px solid #ff9800; padding: 15px; margin: 20px 0; border-radius: 4px;">
<div style="color: #e65100; font-size: 16px; font-weight: bold; margin-bottom: 10px;">🔄 原隧道不可用，已切换到备用地址</div>
<table style="width: 100%; color: #333;">
<tr><td style="padding: 3px 0;"><strong>当前可用地址:</strong></td><td><a href="{tunnel_url}" target="_blank" style="color: #1976d2;">{tunnel_url}</a></td></tr>
<tr><td style="padding: 3px 0;"><strong>当前状态:</strong></td><td style="color: #2e7d32; font-weight: bold;">✅ 备用地址已验证可用</td></tr>
</table>
</div>
'''
            elif event_type == 'unavailable':
                status_note = f"""
🚨 公网地址不可用
原公网地址: {tunnel_url}
当前状态: ❌ 连续验证失败，正在重启隧道服务器
处理措施: 系统已自动触发隧道重启，重启成功后将发送新地址通知

"""
                html_status_note = f'''
<div style="background-color: #ffebee; border-left: 4px solid #f44336; padding: 15px; margin: 20px 0; border-radius: 4px;">
<div style="color: #c62828; font-size: 16px; font-weight: bold; margin-bottom: 10px;">🚨 公网地址不可用</div>
<table style="width: 100%; color: #333;">
<tr><td style="padding: 3px 0;"><strong>原公网地址:</strong></td><td style="color: #c62828;">{tunnel_url}</td></tr>
<tr><td style="padding: 3px 0;"><strong>当前状态:</strong></td><td style="color: #c62828; font-weight: bold;">❌ 连续验证失败，正在重启隧道服务器</td></tr>
<tr><td style="padding: 3px 0;"><strong>处理措施:</strong></td><td>系统已自动触发隧道重启，重启成功后将发送新地址通知</td></tr>
</table>
</div>
'''
            elif event_type == 'restarted':
                status_note = f"""
🔄 隧道已重启
新公网地址: {tunnel_url}
当前状态: ✅ 隧道重启成功，新地址已写入 tunnel_url.txt 和 web_output.log

"""
                html_status_note = f'''
<div style="background-color: #e3f2fd; border-left: 4px solid #2196f3; padding: 15px; margin: 20px 0; border-radius: 4px;">
<div style="color: #1565c0; font-size: 16px; font-weight: bold; margin-bottom: 10px;">🔄 隧道已重启</div>
<table style="width: 100%; color: #333;">
<tr><td style="padding: 3px 0;"><strong>新公网地址:</strong></td><td><a href="{tunnel_url}" target="_blank" style="color: #1976d2;">{tunnel_url}</a></td></tr>
<tr><td style="padding: 3px 0;"><strong>当前状态:</strong></td><td style="color: #2e7d32; font-weight: bold;">✅ 隧道重启成功</td></tr>
<tr><td style="padding: 3px 0;"><strong>数据同步:</strong></td><td>新地址已写入 tunnel_url.txt 和 web_output.log</td></tr>
</table>
</div>
'''
            
            if event_type == 'unavailable':
                body = f"""{event_title}

时间: {current_time}
原公网地址: {tunnel_url}
{status_note}
系统正在自动重启隧道服务器，重启成功后将发送新地址通知。
"""
            elif event_type == 'restarted':
                body = f"""{event_title}

时间: {current_time}
新公网地址: {tunnel_url}
{status_note}
请使用新地址访问服务。
"""
            elif event_type == 'fallback_available':
                body = f"""{event_title}

时间: {current_time}
当前可用地址: {tunnel_url}
{status_note}
备用地址已验证可用，可放心使用。
"""
            else:
                body = f"""{event_title}

时间: {current_time}
公网地址: {tunnel_url}
{status_note}
请妥善保管此地址。
"""
            
            _header_gradient = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)'
            if event_type == 'unavailable':
                _header_gradient = 'linear-gradient(135deg, #e53935 0%, #c62828 100%)'
            elif event_type == 'restarted':
                _header_gradient = 'linear-gradient(135deg, #1e88e5 0%, #1565c0 100%)'
            elif event_type == 'fallback_available':
                _header_gradient = 'linear-gradient(135deg, #ff9800 0%, #f57c00 100%)'
            
            _url_label = '原公网地址' if event_type == 'unavailable' else '新公网地址' if event_type == 'restarted' else '当前可用地址' if event_type == 'fallback_available' else '公网地址'
            _footer_text = '系统正在自动重启隧道服务器，重启成功后将发送新地址通知。' if event_type == 'unavailable' else '请使用新地址访问服务。' if event_type == 'restarted' else '备用地址已验证可用，可放心使用。' if event_type == 'fallback_available' else '请妥善保管此地址。'
            
            html_body = f"""
<html>
<body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">

<div style="background: {_header_gradient}; color: white; padding: 30px; border-radius: 12px; margin-bottom: 25px; text-align: center;">
<h1 style="margin: 0; font-size: 28px; font-weight: bold;">{event_title}</h1>
<p style="margin: 10px 0 0 0; opacity: 0.9; font-size: 16px;">隧道服务通知</p>
</div>

<div style="background-color: #ffffff; border: 1px solid #e0e0e0; border-radius: 8px; padding: 25px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
<table style="width: 100%; border-collapse: collapse;">
<tr style="border-bottom: 1px solid #f0f0f0;">
<td style="padding: 12px 0; font-weight: bold; color: #555; width: 100px;">时间:</td>
<td style="padding: 12px 0; color: #333;">{current_time}</td>
</tr>
<tr>
<td style="padding: 12px 0; font-weight: bold; color: #555;">{_url_label}:</td>
<td style="padding: 12px 0;">
<a href="{tunnel_url}" target="_blank" style="color: #1976d2; text-decoration: none; word-break: break-all;">{tunnel_url}</a>
<button onclick="window.open('{tunnel_url}', '_blank')" style="margin-left: 10px; background-color: #1976d2; color: white; border: none; padding: 6px 16px; border-radius: 4px; cursor: pointer; font-size: 14px;">点击访问</button>
</td>
</tr>
</table>

{html_status_note}

<div style="margin-top: 25px; padding-top: 20px; border-top: 1px solid #f0f0f0; text-align: center; color: #666; font-size: 14px;">
{_footer_text}
</div>
</div>

<div style="text-align: center; margin-top: 20px; color: #999; font-size: 12px;">
<p>此邮件由系统自动发送，请勿直接回复</p>
<p>发送时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
</div>

</body>
</html>
"""
            
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            msg.attach(MIMEText(html_body, 'html', 'utf-8'))
            
            _connect_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{_connect_time}] [EmailNotifier-Thread:{_thread_id}] 🔌 正在连接SMTP服务器 (超时: 30秒)...")
            timeout = 30
            
            _connect_start = datetime.now()
            if config['smtp_port'] == 465:
                server = smtplib.SMTP_SSL(config['smtp_host'], config['smtp_port'], timeout=timeout)
            else:
                server = smtplib.SMTP(config['smtp_host'], config['smtp_port'], timeout=timeout)
                server.starttls()
            _connect_end = datetime.now()
            _connect_duration = (_connect_end - _connect_start).total_seconds()
            
            _login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{_login_time}] [EmailNotifier-Thread:{_thread_id}] ✅ SMTP连接成功 (耗时: {_connect_duration:.2f}秒)")
            print(f"[{_login_time}] [EmailNotifier-Thread:{_thread_id}] 🔐 正在登录SMTP服务器...")
            
            _login_start = datetime.now()
            server.login(config['smtp_user'], config['smtp_password'])
            _login_end = datetime.now()
            _login_duration = (_login_end - _login_start).total_seconds()
            
            _send_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{_send_time}] [EmailNotifier-Thread:{_thread_id}] ✅ SMTP登录成功 (耗时: {_login_duration:.2f}秒)")
            print(f"[{_send_time}] [EmailNotifier-Thread:{_thread_id}] 📤 正在发送邮件...")
            
            _send_start = datetime.now()
            server.sendmail(config['smtp_user'], config['to_email'], msg.as_string())
            server.quit()
            _send_end = datetime.now()
            _send_duration = (_send_end - _send_start).total_seconds()
            
            _success_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{_success_time}] [EmailNotifier-Thread:{_thread_id}] ✅✅✅ 邮件发送成功！")
            print(f"[{_success_time}] [EmailNotifier-Thread:{_thread_id}] 📬 收件人: {config['to_email']}")
            print(f"[{_success_time}] [EmailNotifier-Thread:{_thread_id}] ⏱️ 发送耗时: {_send_duration:.2f}秒")
            print(f"[{_success_time}] [EmailNotifier-Thread:{_thread_id}] ✅ SMTP连接已关闭")
            return True
        except smtplib.SMTPServerDisconnected as e:
            handle_exception(e, 'Email SMTP连接断开')
            raise AppException.email_error(
                f"SMTP连接断开: {e}。请检查网络连接或SMTP服务器配置",
                smtp_host=config['smtp_host'],
                recipient=config['to_email']
            )
        except smtplib.SMTPAuthenticationError as e:
            handle_exception(e, 'Email SMTP认证')
            raise AppException.email_error(
                f"SMTP认证失败: {e}。请检查邮箱账号和授权码",
                smtp_host=config['smtp_host'],
                recipient=config['to_email']
            )
        except smtplib.SMTPException as e:
            handle_exception(e, 'Email SMTP错误')
            raise AppException.email_error(
                f"SMTP错误: {e}",
                smtp_host=config['smtp_host'],
                recipient=config['to_email']
            )
        except Exception as e:
            handle_exception(e, 'Email发送失败')
            raise AppException.email_error(
                f"发送邮件失败: {e}",
                smtp_host=config['smtp_host'],
                recipient=config['to_email']
            )
    
    def save_email_config(self, smtp_host, smtp_port, smtp_user, smtp_password, from_name, to_email):
        """保存邮件配置"""
        self.config_manager.set('email_notification_enabled', True)
        self.config_manager.set('email_smtp_host', smtp_host)
        self.config_manager.set('email_smtp_port', smtp_port)
        self.config_manager.set('email_smtp_user', smtp_user)
        self.config_manager.set('email_smtp_password', smtp_password)
        self.config_manager.set('email_from_name', from_name)
        self.config_manager.set('email_to', to_email)
        return True


class ConfigManager:
    def __init__(self, config_path=None):
        self.config_path = config_path or PathManager.get_config_file()
        self._config = None

    @property
    def config(self):
        if self._config is None:
            self._config = self._load_config()
        return self._config

    def _load_config(self):
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            return {}
        except json.JSONDecodeError as e:
            handle_exception(e, 'ConfigManager JSON解析')
            raise AppException.config_error(
                f"配置文件格式错误: {e}",
                config_key=self.config_path
            )
        except Exception as e:
            handle_exception(e, 'ConfigManager加载配置')
            raise AppException.config_error(
                f"加载配置文件失败: {e}",
                config_key=self.config_path
            )

    def save_config(self):
        if self._config:
            try:
                with open(self.config_path, 'w', encoding='utf-8') as f:
                    json.dump(self._config, f, ensure_ascii=False, indent=2)
                return True
            except PermissionError as e:
                handle_exception(e, 'ConfigManager保存配置权限')
                raise AppException.permission_error(
                    f"保存配置文件失败（权限不足）: {e}",
                    path=self.config_path,
                    operation='write'
                )
            except Exception as e:
                handle_exception(e, 'ConfigManager保存配置')
                raise AppException.config_error(
                    f"保存配置文件失败: {e}",
                    config_key=self.config_path
                )
        return False

    def get(self, key, default=None):
        return self.config.get(key, default)

    def set(self, key, value):
        if self._config is not None:
            self._config[key] = value
            self.save_config()

    def get_cookie_file(self):
        return self.config.get('cookie_file', PathManager.get_cookie_file())

    def get_output_file(self):
        return self.config.get('output_file', PathManager.get_output_file())

    def get_excel_file(self):
        excel_files = self.config.get('excel_files', [])
        if excel_files:
            for path in excel_files:
                expanded_path = os.path.expanduser(path)
                if FileManager.file_exists(expanded_path):
                    return expanded_path
        return self.config.get('excel_file', '')

    def get_all_excel_files(self):
        excel_files = self.config.get('excel_files', [])
        existing_files = []
        if excel_files:
            for path in excel_files:
                expanded_path = os.path.expanduser(path)
                if FileManager.file_exists(expanded_path):
                    existing_files.append(expanded_path)
        return existing_files

    def get_target_url(self):
        return self.config.get('target_url', '')

    def get_user_agent(self):
        return self.config.get('user_agent', WegoScraper.get_user_agent())


# ============================================================
# API Key 认证初始化 (v3.8.89.29) - 必须在 ConfigManager 定义之后
# ============================================================
_web_config_mgr = ConfigManager()
WEB_API_KEY = _web_config_mgr.get('web_api_key')
if not WEB_API_KEY:
    WEB_API_KEY = secrets.token_urlsafe(32)
    _web_config_mgr.set('web_api_key', WEB_API_KEY)


class CookieValidator:
    """Cookie验证器 - 提供完整的cookie验证和友好提示"""
    
    @staticmethod
    def validate_and_prompt(cookie_file):
        """验证cookie并给出友好提示，返回: (is_valid, cookies_or_None)"""
        print_separator()
        print('[验证] Cookie状态检查...')
        print_separator()
        
        # 1. 检查文件是否存在
        if not os.path.exists(cookie_file):
            CookieValidator._show_prompt('Cookie文件不存在', cookie_file, 
                reasons=['首次使用程序，尚未获取Cookie', 'Cookie文件被误删除', '配置文件路径错误'],
                solutions=['返回主菜单选择"4. 更新Cookie"', '浏览器将自动打开并跳转到登录页面', '手动登录账号后关闭浏览器', 'Cookie将自动保存并可以正常使用'],
                tip='Cookie有效期为30天，建议定期更新')
            return False, None
        
        # 2. 检查文件是否可读
        try:
            with open(cookie_file, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
        except json.JSONDecodeError:
            CookieValidator._show_prompt('Cookie文件格式错误', cookie_file,
                reasons=['文件被意外修改', '文件保存时出错', '文件传输过程中损坏'],
                solutions=['删除当前的Cookie文件', '运行"4. 更新Cookie"功能', '重新获取有效的Cookie'],
                tip='建议定期备份Cookie文件')
            return False, None
        except Exception as e:
            CookieValidator._show_prompt('Cookie文件读取失败', cookie_file,
                extra_info=f'❌ 错误信息: {str(e)}',
                reasons=['文件权限不足', '文件被其他程序占用', '磁盘空间不足'],
                solutions=['检查文件权限设置', '关闭可能占用文件的其他程序', '检查磁盘空间', '如果问题持续，请重新运行"更新Cookie"功能'])
            return False, None
        
        # 3. 检查cookie是否为空
        if not cookies:
            CookieValidator._show_prompt('Cookie为空', cookie_file,
                reasons=['Cookie文件被清空', '获取Cookie时出错', 'Cookie保存失败'],
                solutions=['运行"4. 更新Cookie"功能', '重新登录账号', '确认Cookie已正确保存'])
            return False, None
        
        print(f'[OK] Cookie文件存在，共 {len(cookies)} 个Cookie')
        
        # 4. 检查是否存在token
        token_cookie = next((c for c in cookies if 'token' in c.get('name', '').lower()), None)
        if not token_cookie:
            CookieValidator._show_prompt('未找到Token Cookie', cookie_file,
                reasons=['未登录或登录已失效', 'Token Cookie被清除', 'Cookie保存不完整'],
                solutions=['运行"4. 更新Cookie"功能', '确保使用正确的账号登录', '登录成功后再关闭浏览器'],
                tip='Token是保持登录状态的关键Cookie')
            return False, None

        logger.debug('[CookieValidator] 找到Token Cookie')

        # 5. 检查token是否过期
        expires = token_cookie.get('expires')
        if expires and expires < time.time():
            expired_time = datetime.fromtimestamp(expires).strftime('%Y-%m-%d %H:%M:%S')
            current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            days_expired = int((time.time() - expires) / 86400)
            CookieValidator._show_prompt('Token已过期', '',
                extra_info=f'📅 过期时间: {expired_time}\n📅 当前时间: {current_time_str}\n⏱️  已过期: {days_expired}天',
                reasons=['无法获取商品数据', '登录状态失效', '需要重新登录'],
                solutions=['选择"4. 更新Cookie"功能', '使用您的账号重新登录', '更新完成后即可继续使用'],
                tip='建议在Cookie过期前一周进行更新',
                impact=True)
            return False, None
        
        expires_time = datetime.fromtimestamp(expires).strftime('%Y-%m-%d %H:%M:%S') if expires else '未知'
        print(f'[OK] Token有效期至: {expires_time}')
        
        # 6. 检查token值是否有效
        token_value = token_cookie.get('value', '')
        if not token_value or len(token_value) < 10:
            CookieValidator._show_prompt('Token值无效', cookie_file,
                reasons=['Token值为空', 'Token值过短或格式错误', 'Token被意外修改'],
                solutions=['运行"4. 更新Cookie"功能', '重新登录账号', '确认Token已正确保存'],
                tip='正常的Token应该是一长串加密字符串')
            return False, None

        logger.debug('[CookieValidator] Token值验证通过')

        # 7. 检查cookie是否即将过期（7天内）
        if expires:
            days_until_expiry = int((expires - time.time()) / 86400)
            if days_until_expiry <= 7:
                CookieValidator._show_expiry_warning(days_until_expiry)
        
        print_separator()
        print('✅ Cookie验证通过！')
        print_separator()
        
        return True, cookies
    
    @staticmethod
    def _show_prompt(title, file_path, extra_info=None, reasons=None, solutions=None, tip=None, impact=False):
        """显示统一的友好提示"""
        print('\n' + '─'*60)
        print(f'⚠️  检测到{title}')
        print('─'*60)
        if file_path:
            print(f'📂 文件位置: {file_path}')
        if extra_info:
            print(extra_info)
        print()
        if reasons:
            print('📝 可能原因:' if not impact else '📝 影响范围:')
            for reason in reasons:
                print(f'   • {reason}')
            print()
        if solutions:
            print('✅ 解决方案:')
            for i, solution in enumerate(solutions, 1):
                print(f'   {i}. {solution}')
            print()
        if tip:
            print(f'💡 提示: {tip}')
        print('─'*60)
    
    @staticmethod
    def _show_expiry_warning(days_until_expiry):
        """显示即将过期的警告"""
        print('\n' + '─'*60)
        print('⏰  Cookie即将过期提醒')
        print('─'*60)
        print(f'⏱️  剩余有效期: {days_until_expiry}天')
        print()
        if days_until_expiry <= 3:
            print('🔴 状态: 即将过期（3天内）')
            print('⚠️  建议: 立即更新Cookie')
        else:
            print('🟡 状态: 即将过期（7天内）')
            print('💡 建议: 近期更新Cookie')
        print()
        print('✅ 操作: 返回主菜单选择"4. 更新Cookie"')
        print('─'*60)


class FileManager:
    @staticmethod
    def read_json(file_path):
        with ExceptionContext(f"FileManager.read_json({file_path})", default=None) as ctx:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)

    @staticmethod
    def write_json(file_path, data, indent=2):
        with ExceptionContext(f"FileManager.write_json({file_path})", default=False) as ctx:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=indent)
            return True

    @staticmethod
    def read_text(file_path):
        with ExceptionContext(f"FileManager.read_text({file_path})", default=None) as ctx:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()

    @staticmethod
    def write_text(file_path, content):
        with ExceptionContext(f"FileManager.write_text({file_path})", default=False) as ctx:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return True

    @staticmethod
    def file_exists(file_path):
        return os.path.exists(file_path)

    @staticmethod
    def get_latest_json_file(directory=None, pattern='微购相册'):
        with ExceptionContext(f"FileManager.get_latest_json_file", default=None) as ctx:
            directory = directory or PathManager.get_file_dir()
            if not os.path.exists(directory):
                print(f'目录 {directory} 不存在')
                return None
            
            json_files = []
            for file in os.listdir(directory):
                if file.endswith('.json') and pattern in file:
                    file_path = os.path.join(directory, file)
                    json_files.append((file_path, os.path.getmtime(file_path)))
            
            if not json_files:
                print(f'未找到包含"{pattern}"的JSON文件')
                return None
            
            json_files.sort(key=lambda x: x[1], reverse=True)
            latest_file = json_files[0][0]
            print(f'找到最新的JSON文件: {latest_file}')
            return latest_file

    @staticmethod
    def get_today_json_files(directory=None, pattern='微购相册'):
        """
        获取用于对比的两个JSON文件
        优先级：
        1. 当天的缓存文件和最新文件
        2. 当天的最新文件和前一天的文件
        3. 最新的两个文件
        """
        with ExceptionContext("FileManager.get_today_json_files", default=(None, None)) as ctx:
            directory = directory or PathManager.get_file_dir()
            if not os.path.exists(directory):
                print(f'目录 {directory} 不存在')
                return None, None
            
            today = datetime.now().strftime('%Y%m%d')
            
            all_json_files = []
            for file in os.listdir(directory):
                if file.endswith('.json') and pattern in file and '_cache' not in file:
                    file_path = os.path.join(directory, file)
                    all_json_files.append((file_path, os.path.getmtime(file_path)))
            
            if len(all_json_files) < 1:
                print(f'未找到包含"{pattern}"的JSON文件')
                return None, None
            
            all_json_files.sort(key=lambda x: x[1], reverse=True)
            latest_file = all_json_files[0][0]
            
            cache_file = PathManager.get_cache_file_path(today)
            if os.path.exists(cache_file):
                print(f'找到当天缓存文件: {cache_file}')
                print(f'找到当天最新文件: {latest_file}')
                return latest_file, cache_file
            
            today_files = []
            for file in os.listdir(directory):
                if file.endswith('.json') and pattern in file and today in file and '_cache' not in file:
                    file_path = os.path.join(directory, file)
                    today_files.append((file_path, os.path.getmtime(file_path)))
            
            if len(today_files) >= 2:
                today_files.sort(key=lambda x: x[1], reverse=True)
                print(f'找到当天最新文件: {today_files[0][0]}')
                print(f'找到当天次新文件: {today_files[1][0]}')
                return today_files[0][0], today_files[1][0]
            
            if len(all_json_files) >= 2:
                print(f'当天只有一个文件，使用最新文件和次新文件对比')
                print(f'最新文件: {latest_file}')
                print(f'次新文件: {all_json_files[1][0]}')
                return latest_file, all_json_files[1][0]
            
            print(f'只找到一个文件: {latest_file}')
            return latest_file, None

    @staticmethod
    def list_files(directory=None, pattern=None):
        with ExceptionContext("FileManager.list_files", default=[]) as ctx:
            directory = directory or PathManager.get_file_dir()
            if not os.path.exists(directory):
                return []
            
            files = os.listdir(directory)
            if pattern:
                files = [f for f in files if pattern in f]
            return files

    @staticmethod
    def safe_read_excel(excel_file, max_retries=3, retry_delay=0.5):
        """
        安全读取Excel文件，处理Windows共享违规问题
        通过复制到临时文件再读取，确保原文件不被锁定
        
        Args:
            excel_file: Excel文件路径
            max_retries: 最大重试次数
            retry_delay: 重试间隔（秒）
            
        Returns:
            dict: {sheet_name: DataFrame} 或 None
        """
        if pd is None:
            return None
        
        temp_file = None
        try:
            for attempt in range(max_retries):
                if temp_file and os.path.exists(temp_file):
                    safe_execute_func(lambda: os.remove(temp_file), context='safe_read_excel重试清理旧临时文件')
                    temp_file = None
                try:
                    temp_dir = os.path.join(PROJECT_DIR, 'temp')
                    os.makedirs(temp_dir, exist_ok=True)
                    temp_file = os.path.join(temp_dir, f'_temp_excel_{uuid.uuid4().hex}.xlsx')
                    shutil.copy2(excel_file, temp_file)
                    
                    with file_write_lock:
                        xls = pd.ExcelFile(temp_file, engine='openpyxl')
                        dfs = {}
                        for sheet in xls.sheet_names:
                            dfs[sheet] = xls.parse(sheet)
                        xls.close()
                    return dfs
                except PermissionError as e:
                    if "sharing violation" in str(e).lower() or "另一个程序" in str(e) or "正在使用" in str(e) or "Permission" in str(e):
                        if attempt < max_retries - 1:
                            print(f'Excel文件被占用，正在等待重试 ({attempt + 1}/{max_retries}): {excel_file}')
                            time.sleep(retry_delay)
                        else:
                            raise AppException.excel_error(f'Excel文件读取失败（共享违规）', excel_file=excel_file)
                    else:
                        raise
                except Exception as e:
                    raise AppException.parse_error(f'读取Excel文件失败', data_type='excel', raw_data=str(e))
        finally:
            if temp_file and os.path.exists(temp_file):
                safe_execute_func(
                    lambda: os.remove(temp_file),
                    context='清理临时Excel文件'
                )
            auto_clean_temp_dir()
        
        return None


class WegoScraper:
    def __init__(self, config_path=None):
        if config_path is None:
            config_path = PathManager.get_config_file()
        self.config_manager = ConfigManager(config_path)

    @staticmethod
    def get_system_info():
        return Environment.get_system_info()
    
    @staticmethod
    def get_chrome_path():
        return Environment.get_chrome_path()
    
    @staticmethod
    def get_browser_args():
        return Environment.get_browser_args()
    
    @staticmethod
    def get_user_agent():
        return Environment.get_user_agent()

    @staticmethod
    def clean_product_name(text):
        if not text:
            return None
        
        text = re.sub(r'¥\d+', '', text)
        
        for pattern in [r'删除下载刷新编辑分享商品属性标签', r'售价：', r'昨天分享']:
            text = re.sub(pattern, '', text)
        
        text = re.sub(r'\s+', ' ', text).strip()
        return text if text else None

    async def close_popups(self, page, close_limit=3, wait_time=0.3):
        popup_selectors = [
            '[class*="close"]', '[class*="modal-close"]', '[class*="dialog-close"]',
            'button:has-text("关闭")', 'button:has-text("×")', 'button:has-text("✕")',
            '.ant-modal-close', '.el-dialog__close',
        ]
        
        for selector in popup_selectors[:close_limit]:
            safe_execute_func(
                lambda: self._close_popup_impl(page, selector, wait_time),
                context=f'close_popups({selector})'
            )
    
    async def _close_popup_impl(self, page, selector, wait_time):
        close_button = await page.query_selector(selector, timeout=1000)
        if close_button:
            await close_button.click(timeout=1000)
            print(f'关闭了弹窗: {selector}')
            await asyncio.sleep(wait_time)

    async def scroll_to_load_all(self, page):
        print('开始滚动加载所有商品...')
        
        config = self.config_manager.get('scroll_config', {})
        max_attempts = config.get('max_attempts', 30)
        same_height_limit = config.get('same_height_limit', 8)
        scroll_wait_time = config.get('scroll_wait_time', 0.8)
        popup_close_interval = config.get('popup_close_interval', 5)
        popup_close_limit = config.get('popup_close_limit', 3)
        popup_close_wait = config.get('popup_close_wait', 0.3)
        
        print(f'滚动配置: 最大尝试{max_attempts}次, 高度不变限制{same_height_limit}次, 初始等待时间{scroll_wait_time}秒')
        
        last_height = 0
        no_change_count = 0
        height_history = []
        dynamic_adjust = config.get('dynamic_adjust', True)
        
        for scroll_attempts in range(max_attempts):
            with ExceptionContext(f"scroll_to_load_all第{scroll_attempts + 1}次滚动", default=None) as ctx:
                start_time = time.time()
                
                current_height = await asyncio.wait_for(
                    page.evaluate('document.body.scrollHeight'),
                    timeout=5.0
                ) if scroll_attempts > 0 else 0
                
                load_time = time.time() - start_time
                
                height_history.append(current_height)
                if len(height_history) > 10:
                    height_history.pop(0)
                
                if current_height == last_height:
                    no_change_count += 1
                    if no_change_count >= same_height_limit:
                        print(f'页面已滚动到底部（高度连续{same_height_limit}次不变: {current_height}），停止滚动')
                        break
                else:
                    no_change_count = 0
                    last_height = current_height
                
                scroll_distance = current_height * 0.3 if scroll_attempts < 10 else current_height
                safe_execute_func(
                    lambda: asyncio.wait_for(
                        page.evaluate(f'window.scrollBy(0, {scroll_distance})' if scroll_attempts < 10 else 'window.scrollTo(0, document.body.scrollHeight)'),
                        timeout=3.0
                    ),
                    context='scroll操作'
                )
                
                await asyncio.sleep(scroll_wait_time)
                
                progress_percent = min(100, int((scroll_attempts + 1) / max_attempts * 100))
                print(f'滚动 {scroll_attempts + 1}/{max_attempts} ({progress_percent}%) - 当前高度: {current_height} - 加载耗时: {load_time:.2f}秒')
                
                if dynamic_adjust and len(height_history) >= 5:
                    height_changes = [abs(height_history[i] - height_history[i-1]) for i in range(1, len(height_history))]
                    avg_change = sum(height_changes) / len(height_changes)
                    
                    if avg_change < 50 and scroll_wait_time < 2.0:
                        scroll_wait_time = min(2.0, scroll_wait_time + 0.1)
                        print(f'  ⚠️  页面加载较慢，增加等待时间至 {scroll_wait_time:.1f}秒')
                    elif avg_change > 300 and scroll_wait_time > 0.5:
                        scroll_wait_time = max(0.5, scroll_wait_time - 0.1)
                        print(f'  ✅ 页面加载较快，减少等待时间至 {scroll_wait_time:.1f}秒')
                
                if (scroll_attempts + 1) % popup_close_interval == 0:
                    await self.close_popups(page, popup_close_limit, popup_close_wait)
                
                if ctx.error:
                    break
        
        print('滚动完成')

    @staticmethod
    def extract_product_info(element_text, html_content):
        try:
            stock_match = re.search(r'货号[：:]\s*(\d+)', element_text)
            stock_number = stock_match.group(1) if stock_match else ''
            
            def extract_price(text):
                try:
                    price_match = re.search(r'售价[：:]\s*¥?\s*([\d,]+)', text)
                    if not price_match:
                        price_match = re.search(r'¥\s*([\d,]+)(?![0-9])', text)
                    if price_match:
                        try:
                            price_value = int(price_match.group(1).replace(',', ''))
                            if 100 <= price_value <= 50000:
                                return '¥' + price_match.group(1)
                        except (ValueError, TypeError):
                            pass
                    return None
                except Exception as e:
                    logger.warning(f'Exception extracting price: {e}')
                    return None

            price = extract_price(element_text)
            
            def extract_cost_price(text, html):
                try:
                    cost_match = re.search(r'拿货价[：:]\s*¥?\s*([\d,]+)', text)
                    if cost_match:
                        try:
                            cost_value = int(cost_match.group(1).replace(',', ''))
                            if cost_value > 50:
                                return '¥' + cost_match.group(1)
                        except (ValueError, TypeError):
                            pass
                    
                    cost_match2 = re.search(r'拿货价\s*[：:]\s*([\d,]+)', text)
                    if cost_match2:
                        try:
                            cost_value = int(cost_match2.group(1).replace(',', ''))
                            if cost_value > 50:
                                return '¥' + cost_match2.group(1)
                        except (ValueError, TypeError):
                            pass
                    
                    if html:
                        html_cost_match = re.search(r'拿货价[：:]\s*¥?\s*([\d,]+)', html)
                        if html_cost_match:
                            try:
                                cost_value = int(html_cost_match.group(1).replace(',', ''))
                                if cost_value > 50:
                                    return '¥' + html_cost_match.group(1)
                            except (ValueError, TypeError):
                                pass
                        
                        html_cost_match2 = re.search(r'拿货价\s*[：:]\s*([\d,]+)', html)
                        if html_cost_match2:
                            try:
                                cost_value = int(html_cost_match2.group(1).replace(',', ''))
                                if cost_value > 50:
                                    return '¥' + html_cost_match2.group(1)
                            except (ValueError, TypeError):
                                pass
                    
                    return None
                except Exception as e:
                    logger.warning(f'Exception extracting cost price: {e}')
                    return None

            cost_price = extract_cost_price(element_text, html_content)
            
            employee_match = re.search(r'员工[：:]\s*(.+)', element_text)
            employee = employee_match.group(1).strip() if employee_match else None
            
            def extract_remark(text, emp_match):
                if emp_match:
                    emp_pos = text.find('员工')
                    price_match = re.search(r'售价[：:]', text)
                    if price_match:
                        price_pos = price_match.start()
                        between_text = text[price_pos:emp_pos]
                        if between_text and len(between_text.strip()) > 0:
                            between_text = between_text.replace('售价：', '').replace('售价:', '')
                            between_text = re.sub(r'¥\s*[\d,]+', '', between_text)
                            between_text = re.sub(r'\s+', ' ', between_text).strip()
                            if between_text and len(between_text) > 0:
                                return between_text
                
                remark_match = re.search(r'备注[：:]\s*(.+?)(?:\s*员工[：:]|$)', text, re.DOTALL)
                if remark_match:
                    return re.sub(r'\s+', ' ', remark_match.group(1).strip())
                return None
            
            remark = extract_remark(element_text, employee_match)
            
            cut_pos = min(
                len(element_text),
                *(pos for pos in [element_text.find('¥'), element_text.find('删除'), element_text.find('货号')] if pos > 0)
            )
            
            name_part = element_text[:cut_pos] if cut_pos < len(element_text) and cut_pos > 10 else element_text
            name = WegoScraper.clean_product_name(re.sub(r'\s+', ' ', name_part.strip()))
            
            if name:
                return {
                    '商品名称': name,
                    '售价': price if price else '',
                    '拿货价': cost_price if cost_price else '',
                    '货号': stock_number,
                    '备注': remark if remark else '',
                    '员工': employee if employee else '',
                    '图片': ''
                }
            return None
        except Exception as e:
            print(f'提取商品信息时出错: {e}')
            return None

    async def process_elements_concurrently(self, page, elements):
        print(f'开始并发处理 {len(elements)} 个商品元素...')
        
        elements_data = []
        for i, element in enumerate(elements):
            try:
                element_text = await asyncio.wait_for(element.text_content(), timeout=2.0)
                html_content = await asyncio.wait_for(element.inner_html(), timeout=2.0)
                
                # 尝试获取商品ID - 尝试多种属性
                element_id = None
                try:
                    element_id = await element.get_attribute('data-id')
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass
                if not element_id:
                    try:
                        href = await element.get_attribute('href')
                        if href:
                            href_match = re.search(r'/(\d+)(?:\?|$)', href)
                            if href_match:
                                element_id = href_match.group(1)
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
                if not element_id:
                    try:
                        element_id = await element.get_attribute('data-goods-id')
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass

                if not element_text or not element_text.strip():
                    continue
                if '试试批量' in element_text or '暂无搭配' in element_text:
                    continue
                if re.match(r'^\d{2}月 \d{4}$', element_text.strip()):
                    continue
                if len(element_text.strip()) < 30:
                    continue

                elements_data.append((element_text, html_content, element_id))
            except asyncio.TimeoutError:
                continue
            except Exception as e:
                print(f'收集元素 {i} 数据时出错: {e}')

        print(f'收集了 {len(elements_data)} 个有效商品数据')

        # 第一轮：提取商品基本信息
        products = []
        seen_products = set()
        products_need_api = []  # 需要通过API获取拿货价的商品

        with ThreadPoolExecutor(max_workers=15) as executor:
            futures = [executor.submit(self.extract_product_info, text, html) for text, html, _ in elements_data]

            for i, future in enumerate(futures):
                try:
                    result = future.result(timeout=2)
                    if result:
                        product_key = result['货号'] or result['商品名称']
                        if product_key not in seen_products:
                            seen_products.add(product_key)
                            
                            # 如果没有拿货价，记录下来后面用API获取
                            if not result.get('拿货价'):
                                products_need_api.append((result, elements_data[i][2]))  # result和element_id
                            else:
                                products.append(result)
                            
                            if len(products) + len(products_need_api) <= 10:
                                print(f'商品 {len(products) + len(products_need_api)}: {result["商品名称"][:50]}...')
                                print(f'  售价: {result["售价"]}')
                                print(f'  拿货价: {result["拿货价"]}')
                                print(f'  货号: {result["货号"]}')
                                print(f'  data-id: {elements_data[i][2]}\n')
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass
        
        # 第二轮：通过API获取缺失拿货价的商品
        if products_need_api:
            print(f'\n通过API获取缺失的拿货价，需要处理 {len(products_need_api)} 个商品...')
            await self.fetch_cost_prices_via_api(page, products_need_api, products)
        
        return products
    
    async def fetch_cost_prices_via_api(self, page, products_need_api, products):
        """通过API获取缺失拿货价的商品详情"""
        print(f'开始通过API获取 {len(products_need_api)} 个商品的拿货价...')
        
        cookie_file = os.path.join(os.path.dirname(__file__), 'config', 'cookies.json')
        cookies = []
        if os.path.exists(cookie_file):
            try:
                with open(cookie_file, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                print(f'读取到 {len(cookies)} 个cookie')
            except Exception as e:
                handle_exception(e, 'fetch_cost_prices_via_api读取Cookie')
                return
        
        try:
            cookie_str = '; '.join([f'{c.get("name", "")}={c.get("value", "")}' for c in cookies if c.get("name") and c.get("value")])
        except (AttributeError, TypeError) as e:
            print(f'  ⚠️ Cookie格式错误: {e}')
            return
        
        current_url = page.url
        album_id = '_du7mJco53PgiClrX_onUY7Hs5F3Mez8q5_nMrFQ'
        
        try:
            album_id_match = re.search(r'albumId=([^&/]+)|/shop_detail/([^/?#]+)', current_url)
            if album_id_match:
                album_id = album_id_match.group(1) or album_id_match.group(2) or album_id
        except Exception as e:
            print(f'  ⚠️ URL解析失败，使用默认albumId: {e}')
        
        print(f'Album ID: {album_id}')
        
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'User-Agent': Environment.get_user_agent(),
            'x-wg-language': 'zh'
        }
        
        # 使用正确的API获取商品列表
        api_url = 'https://www.szwego.com/album/personal/all'
        
        all_goods_data = []
        
        # 获取所有商品（处理分页）
        page_timestamp = ''
        for page_num in range(20):
            params = {
                'albumId': album_id,
                'searchValue': '',
                'searchImg': '',
                'startDate': '',
                'endDate': '',
                'sourceId': ''
            }
            # 只有第一页不需要timestamp，后续需要
            if page_timestamp:
                params['timestamp'] = page_timestamp
            
            try:
                headers_with_cookie = dict(headers)
                headers_with_cookie['Cookie'] = cookie_str
                
                try:
                    response = await page.request.get(api_url, params=params, headers=headers_with_cookie)
                except Exception as req_error:
                    print(f'  ⚠️ API请求异常: {req_error}')
                    if 'pattern' in str(req_error).lower():
                        print(f'  💡 可能原因: URL格式错误或网络问题')
                    break
                
                if response.status == 200:
                    try:
                        text = await response.text()
                    except Exception as text_error:
                        print(f'  ⚠️ 响应内容读取失败: {text_error}')
                        break
                    
                    # 检查是否返回了HTML而非JSON（常见问题：Cookie过期、反爬等）
                    if text.strip().startswith('<'):
                        print(f'  ⚠️  错误: API返回了HTML而非JSON（可能原因：Cookie过期/失效、触发反爬机制、服务器错误）')
                        print(f'  📄 响应内容前200字符: {text[:200]}...')
                        
                        # 尝试检测具体的错误类型
                        if '登录' in text or 'login' in text.lower():
                            print(f'  🔒 检测到: 需要重新登录（Cookie已过期）')
                        elif '验证码' in text or 'captcha' in text.lower():
                            print(f'  🛡️ 检测到: 触发了验证码验证')
                        elif '403' in text or 'forbidden' in text.lower():
                            print(f'  🚫 检测到: 访问被禁止（403 Forbidden）')
                        elif '404' in text:
                            print(f'  ❌ 检测到: API端点不存在（404 Not Found）')
                        else:
                            print(f'  ⚠️  未知错误类型，请检查网络连接和Cookie有效性')
                        
                        break
                    
                    try:
                        data = json.loads(text)
                        
                        if not isinstance(data, dict):
                            print(f'  ⚠️ API返回数据格式错误（非字典类型）')
                            break
                        
                        if data.get('code') and data.get('code') != 0:
                            print(f'  ❌ API业务错误: code={data.get("code")}, message={data.get("message", "未知错误")}')
                            break
                        
                        result = data.get('result')
                        if not result or not isinstance(result, dict):
                            print(f'  ⚠️ API返回数据缺少result字段')
                            break
                        
                        items = result.get('items', [])
                        if not isinstance(items, list):
                            print(f'  ⚠️ items字段格式错误')
                            items = []
                        
                        pagination = result.get('pagination', {})
                        if not isinstance(pagination, dict):
                            pagination = {}
                        
                        if items:
                            all_goods_data.extend(items)
                            print(f'  第{page_num+1}页: 获取 {len(items)} 个商品')
                            
                            if page_num == 0 and items:
                                try:
                                    first_item = items[0] if items else {}
                                    title = str(first_item.get('title', ''))[:30]
                                    goods_num = str(first_item.get('goodsNum', ''))
                                    print(f'    调试: 第一个商品 title={title}... goodsNum={goods_num}')
                                except Exception as debug_error:
                                    print(f'    调试: 商品信息读取失败: {debug_error}')
                            
                            is_load_more = pagination.get('isLoadMore', False)
                            try:
                                page_timestamp = str(pagination.get('pageTimestamp', ''))
                            except (ValueError, TypeError):
                                page_timestamp = ''
                            
                            if is_load_more and page_timestamp:
                                params['timestamp'] = page_timestamp
                            else:
                                break
                        else:
                            break
                    except json.JSONDecodeError as e:
                        print(f'  ❌ JSON解析失败: {e}')
                        print(f'  📄 响应内容前500字符: {text[:500]}...')
                        break
                    except Exception as e:
                        handle_exception(e, 'fetch_cost_prices_via_api解析响应')
                        break
                else:
                    print(f'  请求失败: HTTP {response.status}')
                    
                    # 打印错误响应内容以帮助调试
                    error_text = await response.text()
                    if error_text:
                        print(f'  📄 错误响应内容: {error_text[:300]}...')
                    
                    # 根据状态码给出具体建议
                    if response.status == 401:
                        print(f'  💡 建议: Cookie已过期或无效，请重新获取Cookie')
                    elif response.status == 403:
                        print(f'  💡 建议: 访问被拒绝，可能触发了反爬机制')
                    elif response.status == 429:
                        print(f'  💡 建议: 请求过于频繁，请稍后重试')
                    elif response.status >= 500:
                        print(f'  💡 建议: 服务器内部错误，请稍后重试或联系管理员')
                    
                    break
            except Exception as e:
                handle_exception(e, 'fetch_cost_prices_via_api请求')
                break
        
        print(f'共获取 {len(all_goods_data)} 个商品数据')
        
        # 调试：打印API获取到的goodsNum列表
        api_goods_nums = [item.get('goodsNum', '') for item in all_goods_data if item.get('goodsNum')]
        print(f'  调试: API goodsNums: {api_goods_nums[:10]}')
        
        # 调试：打印products_need_api中的货号和名称
        print(f'  调试: 需要API的商品数: {len(products_need_api)}')
        if products_need_api:
            sample = products_need_api[0][0]
            print(f'  调试: 第一个商品的货号={sample.get("货号", "")}, 名称={sample.get("商品名称", "")[:30]}...')
        
        goods_by_num = {}
        goods_by_title = {}
        for item in all_goods_data:
            try:
                if not isinstance(item, dict):
                    continue
                goods_num = str(item.get('goodsNum', ''))
                title = str(item.get('title', ''))
                if goods_num:
                    goods_by_num[goods_num] = item
                if title:
                    goods_by_title[title] = item
            except Exception as e:
                logger.debug(f'Exception in loop: {e}')
                continue
        
        success_count = 0
        for product, element_id in products_need_api:
            try:
                if not isinstance(product, dict):
                    continue
                
                goods_num = str(product.get('货号', ''))
                title = str(product.get('商品名称', ''))
                
                matched_item = None
                if goods_num and goods_num in goods_by_num:
                    matched_item = goods_by_num[goods_num]
                elif title and title in goods_by_title:
                    matched_item = goods_by_title[title]
                
                if matched_item and isinstance(matched_item, dict):
                    price_arr = matched_item.get('priceArr', [])
                    if not isinstance(price_arr, list):
                        price_arr = []
                    
                    cost_price = None
                    for price_item in price_arr:
                        if not isinstance(price_item, dict):
                            continue
                        try:
                            if price_item.get('priceType') == 1:
                                cost_price = price_item.get('value')
                                break
                        except Exception as e:
                            logger.debug(f'Exception in loop: {e}')
                            continue
                    
                    if cost_price is not None:
                        try:
                            cost_price_int = int(cost_price)
                            product['拿货价'] = f'¥{cost_price_int:,}'
                            success_count += 1
                            product_name = str(product.get('商品名称', ''))[:30]
                            print(f'  [OK] 获取拿货价: {product_name}... -> {product["拿货价"]}')
                        except (ValueError, TypeError):
                            product_name = str(product.get('商品名称', ''))[:30]
                            print(f'  ⚠ 拿货价格式错误: {product_name}...')
                    else:
                        product_name = str(product.get('商品名称', ''))[:30]
                        print(f'  ⚠ 无拿货价: {product_name}...')
                else:
                    product_name = str(product.get('商品名称', ''))[:30]
                    print(f'  ⚠ 未匹配到: {product_name}...')
            except Exception as e:
                print(f'  ⚠ 处理商品匹配失败: {e}')
                continue
        
        print(f'API获取完成，成功获取 {success_count}/{len(products_need_api)} 个拿货价')

    async def get_data_with_playwright(self, page):
        try:
            target_url = self.config_manager.get_target_url()
            print(f'正在访问目标页面: {target_url}')
            print(f'当前系统: {self.get_system_info()}')
            
            # 页面导航重试机制
            max_retries = 3
            page_loaded = False
            
            for retry in range(max_retries):
                goto_start = time.time()
                try:
                    print(f'尝试加载页面 (第{retry + 1}/{max_retries}次)...')
                    await page.goto(target_url, timeout=30000, wait_until='domcontentloaded')
                    print('页面DOM已加载')
                    page_loaded = True
                    break
                except Exception as e:
                    print(f'页面导航出错: {e}')
                    if retry < max_retries - 1:
                        print(f'等待3秒后重试...')
                        await asyncio.sleep(3)
                    else:
                        print('所有重试都失败，尝试继续执行...')
                        await asyncio.sleep(2)
                print(f'页面导航耗时: {time.time() - goto_start:.2f}秒')
            
            if not page_loaded:
                print('警告: 页面可能未完全加载，继续执行...')
            
            await asyncio.sleep(2)
            
            # 直接通过API获取所有商品数据
            print('直接通过API获取所有商品数据...')
            products = await self.fetch_all_products_via_api(page)
            
            if products:
                print(f'通过API成功获取 {len(products)} 个商品')
                print(f'get_data_with_playwright 返回: {len(products)} 个商品')
                return products
            
            # 如果API失败，返回空列表
            print('API获取失败，返回空列表')
            return []
        except Exception as e:
            print(f'获取数据失败: {e}')
            traceback.print_exc()
            return []

    async def fetch_all_products_via_api(self, page):
        """通过API获取所有商品数据"""
        cookie_file = os.path.join(os.path.dirname(__file__), 'config', 'cookies.json')
        cookies = []
        if os.path.exists(cookie_file):
            try:
                with open(cookie_file, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                print(f'读取到 {len(cookies)} 个cookie')
            except Exception as e:
                print(f'读取cookie失败: {e}')
                return None
        
        try:
            cookie_str = '; '.join([f'{c.get("name", "")}={c.get("value", "")}' for c in cookies if c.get("name") and c.get("value")])
        except (AttributeError, TypeError) as e:
            print(f'  ⚠️ Cookie格式错误: {e}')
            return None
        
        current_url = page.url
        album_id = '_du7mJco53PgiClrX_onUY7Hs5F3Mez8q5_nMrFQ'
        
        try:
            album_id_match = re.search(r'albumId=([^&/]+)|/shop_detail/([^/?#]+)', current_url)
            if album_id_match:
                album_id = album_id_match.group(1) or album_id_match.group(2) or album_id
        except Exception as e:
            print(f'  ⚠️ URL解析失败，使用默认albumId: {e}')
        
        print(f'Album ID: {album_id}')
        
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'User-Agent': Environment.get_user_agent(),
            'x-wg-language': 'zh'
        }
        
        api_url = 'https://www.szwego.com/album/personal/all'
        
        all_goods_data = []
        page_timestamp = ''
        
        print('开始通过API获取所有商品...')
        for page_num in range(20):
            params = {
                'albumId': album_id,
                'searchValue': '',
                'searchImg': '',
                'startDate': '',
                'endDate': '',
                'sourceId': ''
            }
            if page_timestamp:
                params['timestamp'] = page_timestamp
            
            try:
                headers_with_cookie = dict(headers)
                headers_with_cookie['Cookie'] = cookie_str
                
                try:
                    response = await page.request.get(api_url, params=params, headers=headers_with_cookie)
                except Exception as req_error:
                    print(f'  ⚠️ API请求异常: {req_error}')
                    if 'pattern' in str(req_error).lower():
                        print(f'  💡 可能原因: URL格式错误或网络问题')
                    break
                
                if response.status == 200:
                    try:
                        text = await response.text()
                    except Exception as text_error:
                        print(f'  ⚠️ 响应内容读取失败: {text_error}')
                        break
                    
                    # 检查是否返回了HTML而非JSON（常见问题：Cookie过期、反爬等）
                    if text.strip().startswith('<'):
                        print(f'  ⚠️  错误: API返回了HTML而非JSON（可能原因：Cookie过期/失效、触发反爬机制、服务器错误）')
                        print(f'  📄 响应内容前200字符: {text[:200]}...')
                        
                        # 尝试检测具体的错误类型
                        if '登录' in text or 'login' in text.lower():
                            print(f'  🔒 检测到: 需要重新登录（Cookie已过期）')
                        elif '验证码' in text or 'captcha' in text.lower():
                            print(f'  🛡️ 检测到: 触发了验证码验证')
                        elif '403' in text or 'forbidden' in text.lower():
                            print(f'  🚫 检测到: 访问被禁止（403 Forbidden）')
                        elif '404' in text:
                            print(f'  ❌ 检测到: API端点不存在（404 Not Found）')
                        else:
                            print(f'  ⚠️  未知错误类型，请检查网络连接和Cookie有效性')
                        
                        break
                    
                    try:
                        data = json.loads(text)
                        
                        if not isinstance(data, dict):
                            print(f'  ⚠️ API返回数据格式错误（非字典类型）')
                            break
                        
                        if data.get('code') and data.get('code') != 0:
                            print(f'  ❌ API业务错误: code={data.get("code")}, message={data.get("message", "未知错误")}')
                            break
                        
                        result = data.get('result')
                        if not result or not isinstance(result, dict):
                            print(f'  ⚠️ API返回数据缺少result字段')
                            break
                        
                        items = result.get('items', [])
                        if not isinstance(items, list):
                            print(f'  ⚠️ items字段格式错误')
                            items = []
                        
                        pagination = result.get('pagination', {})
                        if not isinstance(pagination, dict):
                            pagination = {}
                        
                        if items:
                            all_goods_data.extend(items)
                            print(f'  第{page_num+1}页: 获取 {len(items)} 个商品')
                            
                            is_load_more = pagination.get('isLoadMore', False)
                            try:
                                page_timestamp = str(pagination.get('pageTimestamp', ''))
                            except (ValueError, TypeError):
                                page_timestamp = ''
                            
                            if is_load_more and page_timestamp:
                                params['timestamp'] = page_timestamp
                            else:
                                break
                        else:
                            break
                    except json.JSONDecodeError as e:
                        print(f'  ❌ JSON解析失败: {e}')
                        print(f'  📄 响应内容前500字符: {text[:500]}...')
                        break
                    except Exception as e:
                        print(f'  解析失败: {e}')
                        break
                else:
                    print(f'  请求失败: HTTP {response.status}')
                    
                    # 打印错误响应内容以帮助调试
                    error_text = await response.text()
                    if error_text:
                        print(f'  📄 错误响应内容: {error_text[:300]}...')
                    
                    # 根据状态码给出具体建议
                    if response.status == 401:
                        print(f'  💡 建议: Cookie已过期或无效，请重新获取Cookie')
                    elif response.status == 403:
                        print(f'  💡 建议: 访问被拒绝，可能触发了反爬机制')
                    elif response.status == 429:
                        print(f'  💡 建议: 请求过于频繁，请稍后重试')
                    elif response.status >= 500:
                        print(f'  💡 建议: 服务器内部错误，请稍后重试或联系管理员')
                    
                    break
            except Exception as e:
                print(f'  请求异常: {e}')
                break
        
        print(f'共获取 {len(all_goods_data)} 个商品数据')
        
        if not all_goods_data:
            return None
        
        products = []
        for item in all_goods_data:
            try:
                if not isinstance(item, dict):
                    continue
                
                title = str(item.get('title', ''))
                goods_num = str(item.get('goodsNum', ''))
                
                price_arr = item.get('priceArr', [])
                if not isinstance(price_arr, list):
                    price_arr = []
                
                sale_price = None
                cost_price = None
                for price_item in price_arr:
                    if not isinstance(price_item, dict):
                        continue
                    try:
                        if price_item.get('priceType') == 1:
                            cost_price = price_item.get('value')
                        elif price_item.get('priceType') == 2:
                            sale_price = price_item.get('value')
                    except Exception as e:
                        logger.debug(f'Exception in loop: {e}')
                        continue
                
                note_arr = item.get('noteArr', [])
                remark = ''
                if isinstance(note_arr, list) and note_arr:
                    try:
                        remark = str(note_arr[0].get('value', '')) if isinstance(note_arr[0], dict) else ''
                    except (IndexError, AttributeError, TypeError):
                        remark = ''
                
                staff_info = item.get('staffInfo', {})
                staff_nick = str(staff_info.get('staffNick', '')) if isinstance(staff_info, dict) else ''
                
                media_b64_list = []
                imgs_src = item.get('imgsSrc', [])
                if isinstance(imgs_src, list):
                    for url in imgs_src:
                        try:
                            media_b64_list.append(base64.b64encode(str(url).encode('utf-8')).decode('utf-8'))
                        except Exception as e:
                            logger.debug(f'Exception in loop: {e}')
                            continue
                
                video_url = item.get('videoUrl', '')
                if video_url:
                    try:
                        media_b64_list.append(base64.b64encode(str(video_url).encode('utf-8')).decode('utf-8'))
                    except Exception as e:                        logger.debug(f"Silent exception: {e}")
                
                media_b64 = media_b64_list[0] if len(media_b64_list) == 1 else media_b64_list
                
                time_stamp = item.get('time_stamp', 0)
                old_time = str(item.get('old_time', ''))
                
                created_time = None
                if time_stamp:
                    try:
                        created_time = datetime.fromtimestamp(int(time_stamp) / 1000).strftime('%Y-%m-%d %H:%M:%S')
                    except (ValueError, TypeError, OSError):
                        pass
            
                try:
                    sale_price_int = int(sale_price) if sale_price is not None else None
                    cost_price_int = int(cost_price) if cost_price is not None else None
                except (ValueError, TypeError):
                    sale_price_int = None
                    cost_price_int = None
                
                product = {
                    '商品描述': title,
                    'name': title,
                    '售价': f'¥{sale_price_int:,}' if sale_price_int is not None else '',
                    'price': f'¥{sale_price_int:,}' if sale_price_int is not None else '',
                    '拿货价': f'¥{cost_price_int:,}' if cost_price_int is not None else '',
                    'cost_price': f'¥{cost_price_int:,}' if cost_price_int is not None else '',
                    '货号': goods_num,
                    'stock_number': goods_num,
                    '备注': remark,
                    'remark': remark,
                    '员工': staff_nick,
                    'staff': staff_nick,
                    '图片': media_b64,
                    'image': media_b64,
                    '入库时间': old_time,
                    'created_time': old_time,
                    '入库时间戳': created_time,
                    'timestamp': created_time
                }
                products.append(product)
            except Exception as e:
                print(f'  ⚠️ 处理商品数据失败: {e}')
                continue
        
        print(f'fetch_all_products_via_api 返回: {len(products)} 个商品')
        return products

    @staticmethod
    def parse_price(price_str):
        """解析价格字符串，返回整数价格或None"""
        if not price_str:
            return None        # 移除常见的价格符号
        price_clean = str(price_str).replace('¥', '').replace(',', '').replace('元', '').strip()
        # 使用正则提取数字部分
        match = re.search(r'(\d+)', price_clean)
        if match:
            return int(match.group(1))
        return None

    def filter_high_price_products(self, data, min_price=599):
        """筛选高价商品"""
        return [p for p in data if self.parse_price(p.get('售价', '') or p.get('price', '')) and self.parse_price(p.get('售价', '') or p.get('price', '')) >= min_price]

    def analyze_data_changes(self, data, previous_file):
        """分析数据变化"""
        if not previous_file:
            return ""
        
        try:
            old_data = FileManager.read_json(previous_file)
            if not old_data:
                return ""
            
            old_items = old_data.get('商品列表', [])
            old_nums = {item.get('货号', '') for item in old_items if item.get('货号')}
            current_nums = {item.get('货号', '') for item in data if item.get('货号')}
            
            added = current_nums - old_nums
            removed = old_nums - current_nums
            
            if not added and not removed:
                return "数据无变化"
            
            def get_product_detail(item):
                return {
                    "商品描述": item.get('商品描述', '') or item.get('name', '') or item.get('商品名称', ''),
                    "售价": item.get('售价', '') or item.get('price', ''),
                    "货号": item.get('货号', '') or item.get('stock_number', ''),
                    "备注": item.get('备注', '') or item.get('remark', ''),
                    "员工": item.get('员工', '') or item.get('staff', '')
                }
            
            def format_json_array(items):
                if not items:
                    return "[]"
                lines = ["["]
                for i, item in enumerate(items):
                    lines.append('  {')
                    for j, (k, v) in enumerate(item.items()):
                        lines.append(f'    "{k}": "{v}"' + (',' if j < len(item) - 1 else ''))
                    lines.append('  }' + (',' if i < len(items) - 1 else ''))
                lines.append("]")
                return '\n'.join(lines)
            
            added_details = [get_product_detail(item) for item in data if item.get('货号') in added]
            removed_details = [get_product_detail(item) for item in old_items if item.get('货号') in removed]
            
            return f"对比 {old_data.get('生成日期', 'N/A')} 新增 {len(added)} 个，删除 {len(removed)} 个\n【新增商品】({len(added)}个):\n{format_json_array(added_details)}\n【删除商品】({len(removed)}个):\n{format_json_array(removed_details)}"
        except Exception as e:
                handle_exception(e, 'analyze_data_changes对比分析')
                return f"对比分析失败: {str(e)}"

    def save_data(self, data, filename=None):
        today = datetime.now().strftime('%Y%m%d')
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_filename = PathManager.get_json_file_path(today)
        cache_filename = PathManager.get_cache_file_path(today)
        
        # 如果当天的JSON文件已存在，先保存为缓存文件
        existing_summary = None
        cache_products_map = {}
        if os.path.exists(new_filename):
            try:
                shutil.copy2(new_filename, cache_filename)
                print(f'已将旧数据保存为缓存文件: {cache_filename}')
                
                # 读取现有的"小计"字段和商品列表（用于合并价格）
                existing_data = FileManager.read_json(new_filename)
                if existing_data and '小计' in existing_data:
                    existing_summary = existing_data['小计']
                    print(f'已保留 {len(existing_summary) if isinstance(existing_summary, list) else 1} 条对比记录')
                
                # 构建cache中的货号->商品映射（用于价格合并）
                if existing_data and '商品列表' in existing_data:
                    cache_products = existing_data['商品列表']
                    for cache_p in cache_products:
                        sku = cache_p.get('货号', '') or cache_p.get('stock_number', '')
                        if sku:
                            cache_products_map[sku] = cache_p
                    if cache_products_map:
                        print(f'已加载 {len(cache_products_map)} 个cache商品用于价格合并')
            except Exception as e:
                print(f'创建缓存文件失败: {e}')
        
        # 价格合并逻辑：如果新数据的价格为空，则从cache中获取
        merged_count = 0
        for product in data:
            sku = product.get('货号', '') or product.get('stock_number', '')
            if not sku:
                continue
            
            # 检查售价是否为空
            if (not product.get('售价', '').strip() or not product.get('price', '').strip()) and sku in cache_products_map:
                cache_product = cache_products_map[sku]
                cache_price = cache_product.get('售价', '') or cache_product.get('price', '')
                if cache_price.strip():
                    product['售价'] = cache_price
                    product['price'] = cache_price
                    merged_count += 1
            
            # 检查拿货价是否为空
            if (not product.get('拿货价', '').strip() or not product.get('cost_price', '').strip()) and sku in cache_products_map:
                cache_product = cache_products_map[sku]
                cache_cost = cache_product.get('拿货价', '') or cache_product.get('cost_price', '')
                if cache_cost.strip():
                    product['拿货价'] = cache_cost
                    product['cost_price'] = cache_cost
        
        if merged_count > 0:
            print(f'[OK] 已从cache合并 {merged_count} 个商品的价格信息')
        
        total_count = len(data)
        high_price_products = self.filter_high_price_products(data)
        high_price_count = len(high_price_products)
        
        # 计算累计值
        total_sell_price = 0.0
        total_platform_fee = 0.0
        total_cost_price = 0.0  # 累计成本
        
        for product in data:
            sell_price = 0.0
            cost_price = 0.0
            
            if '售价' in product:
                try:
                    sell_price = float(str(product['售价']).replace('¥', '').replace(',', '').strip())
                    total_sell_price += sell_price
                except (ValueError, TypeError):
                    pass
            
            if '拿货价' in product:
                try:
                    cost_price_str = str(product['拿货价']).replace('¥', '').replace(',', '').strip()
                    if cost_price_str:  # 确保不是空字符串
                        cost_price = float(cost_price_str)
                        total_cost_price += cost_price
                except (ValueError, TypeError):
                    pass
            
            # 计算闲鱼平台手续费（售价 * 1.6%）
            if sell_price > 0:
                platform_fee = sell_price * 0.016
                total_platform_fee += platform_fee
        
        # 计算平均每个设备售出均价
        avg_sell_price = total_sell_price / total_count if total_count > 0 else 0.0
        
        # 计算预计毛利
        estimated_gross_profit = total_sell_price - total_cost_price
        
        # 计算毛利率
        gross_profit_margin = estimated_gross_profit / total_cost_price if total_cost_price > 0 else 0.0
        
        # 计算所有设备卖到闲鱼平台的毛利
        idle_fish_gross_profit = total_sell_price - total_platform_fee
        
        # 计算单个设备的平均回收价格（累计成本 / 商品数量）
        avg_cost_price = total_cost_price / total_count if total_count > 0 else 0.0
        
        existing_files = sorted(FileManager.list_files(PathManager.get_file_dir(), '微购相册'), reverse=True)
        
        previous_file = None
        for f in existing_files:
            if f != PathManager.get_json_filename(today):
                previous_file = os.path.join(PathManager.get_file_dir(), f)
                break
        
        change_summary = self.analyze_data_changes(data, previous_file)
        
        output_data = {
            "生成日期": today,
            "时间戳": current_time,
            "成功获取": f"{total_count} 个商品",
            "商品列表": data,
            "单个设备的平均回收价格": f"{avg_cost_price:,.2f}",
            "平均每个设备售出均价": f"{avg_sell_price:,.2f}",
            "累计成本": f"{total_cost_price:,.2f}",
            "预计毛利": f"{estimated_gross_profit:,.2f}",
            "毛利率": f"{gross_profit_margin:.2%}",
            "预计售出价格累计": f"{total_sell_price:,.2f}",
            "闲鱼平台手续费累计": f"{total_platform_fee:,.2f}",
            "所有设备卖到闲鱼平台的毛利": f"{idle_fish_gross_profit:,.2f}",
            "统计": f"共计获取到 {total_count} 个商品"
        }
        
        # 保留现有的"小计"字段（如果有）
        if existing_summary:
            output_data["小计"] = existing_summary
        
        # "小计"字段将由compare_json_files方法管理，用于存储多次对比的差异记录
        
        output_data["高价商品统计"] = {
            "筛选条件": "售价 >= 599",
            "数量": high_price_count,
            "商品列表": high_price_products
        }
        
        FileManager.write_json(new_filename, output_data)
        print(f'数据已保存到 {new_filename}')
        print(f'成功获取 {total_count} 个商品')
        print(f'售价 >= 599 的商品: {high_price_count} 个')
        print(f'预计售出价格累计: ¥{total_sell_price:,.2f}')
        print(f'平均每个设备售出均价: ¥{avg_sell_price:,.2f}')
        print(f'闲鱼平台手续费累计: ¥{total_platform_fee:,.2f}')
        if change_summary:
            print(f'{change_summary}')
            
        

    async def run(self):
        start_time = time.time()
        start_datetime = datetime.now()
        
        try:
            print('='*50)
            print(f'Szwego商品爬虫 - v{VERSION}')
            print(f'当前系统: {self.get_system_info()}')
            print(f'Python版本: {platform.python_version()}')
            print(f'开始时间: {start_datetime.strftime("%Y-%m-%d %H:%M:%S")}')
            print('='*50)
            print('开始运行...')
            
            browser_start = time.time()
            async with async_playwright() as p:
                print('正在启动浏览器...')
                
                system = self.get_system_info()
                browser_args = self.get_browser_args()
                chrome_path = self.get_chrome_path()
                
                print(f'检测到系统: {system}')
                if chrome_path:
                    print(f'使用系统Chrome: {chrome_path}')
                else:
                    print(f'使用Playwright内置Chromium')
                
                try:
                    browser = await p.chromium.launch(headless=False, args=browser_args, executable_path=chrome_path)
                except Exception as launch_err:
                    if 'Executable doesn\'t exist' in str(launch_err) or 'executable doesn\'t exist' in str(launch_err).lower():
                        print('Playwright内置Chromium不存在，正在自动安装浏览器...')
                        install_playwright_cdn()
                        try:
                            browser = await p.chromium.launch(headless=False, args=browser_args, executable_path=None)
                            print('浏览器安装完成，使用Playwright内置Chromium启动成功')
                        except Exception as retry_err:
                            fallback_chrome = Environment._find_system_chrome()
                            if fallback_chrome:
                                print(f'Playwright内置Chromium仍不可用，回退到系统Chrome: {fallback_chrome}')
                                browser = await p.chromium.launch(headless=False, args=browser_args, executable_path=fallback_chrome)
                            else:
                                print(f'自动安装失败且无系统Chrome可用: {retry_err}')
                                print('请手动运行: python -m playwright install chromium')
                                raise
                    else:
                        raise
                print(f'浏览器启动耗时: {time.time() - browser_start:.2f}秒')
                
                context_start = time.time()
                context = await browser.new_context(
                    viewport=Environment.get_default_viewport(),
                    user_agent=self.get_user_agent()
                )
                print(f'上下文创建耗时: {time.time() - context_start:.2f}秒')
                
                cookie_start = time.time()
                cookie_file = self.config_manager.get_cookie_file()
                if FileManager.file_exists(cookie_file):
                    cookies = FileManager.read_json(cookie_file)
                    if cookies:
                        print(f'已加载 {len(cookies)} 个Cookie')
                        await context.add_cookies(cookies)
                print(f'Cookie加载耗时: {time.time() - cookie_start:.2f}秒')
                
                page_start = time.time()
                page = await context.new_page()
                print(f'页面创建耗时: {time.time() - page_start:.2f}秒')
                
                data_start = time.time()
                products = await self.get_data_with_playwright(page)
                print(f'数据获取耗时: {time.time() - data_start:.2f}秒')
                
                if products:
                    save_start = time.time()
                    self.save_data(products)
                    print(f'数据保存耗时: {time.time() - save_start:.2f}秒')
                    
                    compare_start = time.time()
                    print('\n开始自动对比当天JSON文件...')
                    comparator = StockNumberComparator()
                    comparator.compare_json_files()
                    print(f'对比耗时: {time.time() - compare_start:.2f}秒')
                
                save_cookie_start = time.time()
                try:
                    cookies = await context.cookies()
                    # 统一domain格式，将所有cookie的domain改为.szwego.com
                    szwego_cookies = [cookie for cookie in cookies if 'szwego.com' in cookie['domain']]
                    for cookie in szwego_cookies:
                        if cookie['domain'] == 'www.szwego.com':
                            cookie['domain'] = '.szwego.com'
                    FileManager.write_json(cookie_file, szwego_cookies)
                    print(f'Cookie已保存到 {cookie_file}')
                    print(f'Cookie保存耗时: {time.time() - save_cookie_start:.2f}秒')
                except Exception as e:
                    print(f'⚠️  Cookie保存失败: {e}')
                    print('继续执行，不影响数据获取...')
                
                close_start = time.time()
                try:
                    await browser.close()
                    print(f'浏览器关闭耗时: {time.time() - close_start:.2f}秒')
                except Exception as e:
                    print(f'⚠️  浏览器关闭失败: {e}')
                    print('浏览器可能已经关闭，继续执行...')
                
        except Exception as e:
            print(f'运行失败: {e}')
            traceback.print_exc()
        finally:
            end_time = time.time()
            end_datetime = datetime.now()
            total_time = end_time - start_time
            
            print('='*50)
            print(f'结束时间: {end_datetime.strftime("%Y-%m-%d %H:%M:%S")}')
            print(f'总运行时间: {total_time:.2f} 秒 ({total_time/60:.2f} 分钟)')
            print('='*50)


class StockNumberComparator:
    def __init__(self, output_file=None, input_file=None, config_path=None):
        self.json_file = output_file or FileManager.get_latest_json_file()
        self.input_file = input_file or PathManager.get_input_file()
        self.config_manager = ConfigManager(config_path)
        self.excel_file = self._get_excel_file()

    def _get_excel_file(self):
        excel_file = self.config_manager.get_excel_file()
        if excel_file and FileManager.file_exists(excel_file):
            return excel_file
        return None

    def load_json_data(self):
        if not self.json_file:
            print('[StockNumberComparator] Warning: No JSON file found')
            return []
        
        json_data = FileManager.read_json(self.json_file)
        if not json_data:
            print('[StockNumberComparator] Warning: Cannot read file')
            return []
        
        if isinstance(json_data, dict) and '商品列表' in json_data:
            products = json_data.get('商品列表', [])
            print(f'[StockNumberComparator] Loaded {len(products)} products from {os.path.basename(self.json_file)}')
            return products
        elif isinstance(json_data, list):
            print(f'[StockNumberComparator] Loaded {len(json_data)} products (list format)')
            return json_data
        else:
            print('[StockNumberComparator] Warning: Invalid JSON format')
            return []

    @staticmethod
    def extract_stock_numbers(data):
        return {item.get('货号') for item in data if item.get('货号') and item['货号'] != 'N/A'}

    @staticmethod
    def parse_input_string(input_str):
        if not input_str:
            return []
        
        cleaned = re.sub(r'序列号', '', input_str)
        numbers = re.split(r'[,，\s;；\n\t]+', cleaned)
        return [num.strip() for num in numbers if num.strip()]

    def load_excel_data(self, excel_file=None, remove_duplicates=True):
        if excel_file is None:
            excel_file = self.excel_file
        
        if not excel_file:
            return None
        
        temp_file = None
        try:
            if not FileManager.file_exists(excel_file):
                return None
            
            temp_dir = os.path.join(PROJECT_DIR, 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            temp_file = os.path.join(temp_dir, f'_temp_excel_{uuid.uuid4().hex}.xlsx')
            shutil.copy2(excel_file, temp_file)
            
            print(f'正在读取Excel文件: {excel_file}')
            workbook = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)
            
            sheet = next((workbook[sheet_name] for sheet_name in workbook.sheetnames if '闲鱼' in sheet_name), None)
            
            if sheet is None:
                print('未找到"闲鱼"工作表，使用第一个工作表')
                sheet = workbook.active
            else:
                print(f'使用工作表: {sheet.title}')
            
            stock_numbers = []
            for row in sheet.iter_rows(min_col=5, max_col=5, values_only=True):
                cell_value = row[0]
                if cell_value:
                    cell_str = str(cell_value).strip()
                    number_match = re.match(r'^([A-Za-z0-9]{3,10})$', cell_str)
                    if number_match:
                        stock_numbers.append(cell_str)
            
            stock_numbers = list(set(stock_numbers)) if remove_duplicates else stock_numbers
            print(f'从Excel文件的E列中读取到 {len(stock_numbers)} 个货号')
            workbook.close()
            return stock_numbers
        except Exception as e:
                handle_exception(e, 'load_excel_data读取Excel')
                return None
        finally:
            if temp_file and os.path.exists(temp_file):
                safe_execute_func(lambda: os.remove(temp_file), context='load_excel_data清理临时文件')
            auto_clean_temp_dir()

    def load_all_excel_data(self, remove_duplicates=True):
        all_stock_numbers = []
        excel_files = self.config_manager.get_all_excel_files()
        
        if not excel_files:
            print('未找到任何Excel文件')
            return None
        
        excel_files = list(dict.fromkeys(os.path.abspath(f) for f in excel_files))
        
        print(f'找到 {len(excel_files)} 个Excel文件')
        
        for excel_file in excel_files:
            temp_file = None
            try:
                if not FileManager.file_exists(excel_file):
                    continue
                
                temp_dir = os.path.join(PROJECT_DIR, 'temp')
                os.makedirs(temp_dir, exist_ok=True)
                temp_file = os.path.join(temp_dir, f'_temp_excel_{uuid.uuid4().hex}.xlsx')
                shutil.copy2(excel_file, temp_file)
                
                print(f'正在读取Excel文件: {excel_file}')
                workbook = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)
                
                sheet = next((workbook[sheet_name] for sheet_name in workbook.sheetnames if '闲鱼' in sheet_name), None)
                
                if sheet is None:
                    print('未找到"闲鱼"工作表，使用第一个工作表')
                    sheet = workbook.active
                else:
                    print(f'使用工作表: {sheet.title}')
                
                file_stock_numbers = []
                for row in sheet.iter_rows(min_col=5, max_col=5, values_only=True):
                    cell_value = row[0]
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        number_match = re.match(r'^([A-Za-z0-9]{3,10})$', cell_str)
                        if number_match:
                            file_stock_numbers.append(cell_str)
                
                print(f'从 {os.path.basename(excel_file)} 的E列中读取到 {len(file_stock_numbers)} 个货号')
                all_stock_numbers.extend(file_stock_numbers)
                workbook.close()
                
            except Exception as e:
                handle_exception(e, f'load_all_excel_data读取Excel {excel_file}')
                continue
            finally:
                if temp_file and os.path.exists(temp_file):
                    safe_execute_func(lambda: os.remove(temp_file), context='load_all_excel_data清理临时文件')
                auto_clean_temp_dir()
        
        if remove_duplicates:
            all_stock_numbers = list(set(all_stock_numbers))
        
        print(f'总共读取到 {len(all_stock_numbers)} 个货号（已去重）')
        return all_stock_numbers

    def save_input_to_file(self, input_str):
        if FileManager.write_text(self.input_file, input_str):
            print(f'输入已保存到 {self.input_file}')
            return True
        return False

    def load_input_from_file(self):
        return FileManager.read_text(self.input_file)

    @staticmethod
    def compare_stock_numbers(json_stock_numbers, input_stock_numbers, high_price_stock_numbers=None):
        json_set = set(json_stock_numbers)
        input_set = set(input_stock_numbers)
        
        result = {
            'missing': sorted(list(input_set - json_set)),
            'existing': sorted(list(input_set & json_set)),
            'extra_in_json': sorted(list(json_set - input_set)),
            'total_input': len(input_set),
            'total_json': len(json_set),
            'missing_count': len(input_set - json_set),
            'existing_count': len(input_set & json_set),
            'extra_in_json_count': len(json_set - input_set)
        }
        
        if high_price_stock_numbers:
            result['high_price_stock_numbers'] = sorted(list(set(high_price_stock_numbers)))
            result['high_price_count'] = len(result['high_price_stock_numbers'])
        
        return result

    @staticmethod
    def find_duplicate_stock_numbers(input_stock_numbers):
        seen = {}
        for num in input_stock_numbers:
            seen[num] = seen.get(num, 0) + 1
        
        return [{'货号': num, 'count': count, 'positions': count} 
                for num, count in seen.items() if count > 1]

    def save_duplicate_log(self, duplicates, log_file=None):
        try:
            log_file = log_file or PathManager.get_duplicate_log_file()
            log_data = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_duplicates': len(duplicates),
                'duplicates': duplicates
            }
            
            if FileManager.file_exists(log_file):
                existing_data = FileManager.read_json(log_file)
                if isinstance(existing_data, list):
                    log_data['history'] = existing_data
                elif isinstance(existing_data, dict) and 'history' in existing_data:
                    log_data['history'] = existing_data['history']
            
            FileManager.write_json(log_file, log_data)
            print(f'重复序列号日志已保存到 {log_file}')
            return True
        except Exception as e:
                handle_exception(e, 'save_duplicate_log保存重复日志')
                return False

    def compare_json_files(self):
        """
        对比当天最新的两个JSON文件，将差异写进最新的JSON文件中
        如果使用缓存文件，对比完成后会删除缓存文件
        """
        try:
            print_separator()
            print('当天JSON文件对比工具')
            print_separator()
            
            # 获取用于对比的两个JSON文件
            latest_json_file, second_latest_json_file = FileManager.get_today_json_files()
            if not latest_json_file:
                print('无法获取最新的JSON文件')
                return False
            
            if not second_latest_json_file:
                print('只找到一个JSON文件，无法进行对比')
                print(f'当前文件: {latest_json_file}')
                print('提示：运行爬虫后再次运行此功能即可进行对比')
                return True
            
            # 检查是否使用缓存文件
            is_cache_used = '_cache' in second_latest_json_file
            
            # 读取最新的JSON文件
            latest_json_data = FileManager.read_json(latest_json_file)
            if not latest_json_data:
                print('无法读取最新的JSON文件')
                return False
            
            # 读取次新的JSON文件
            second_json_data = FileManager.read_json(second_latest_json_file)
            if not second_json_data:
                print('无法读取次新的JSON文件')
                return False
            
            # 提取商品列表
            latest_products = latest_json_data.get('商品列表', [])
            second_products = second_json_data.get('商品列表', [])
            
            if not latest_products or not second_products:
                print('JSON文件中没有商品列表')
                return False
            
            # 提取货号
            latest_stock_numbers = {item.get('货号', '') for item in latest_products if item.get('货号')}
            second_stock_numbers = {item.get('货号', '') for item in second_products if item.get('货号')}
            
            print(f'从最新JSON文件中读取到 {len(latest_stock_numbers)} 个货号')
            print(f'从次新JSON文件中读取到 {len(second_stock_numbers)} 个货号\n')
            
            # 计算差异
            added = latest_stock_numbers - second_stock_numbers
            removed = second_stock_numbers - latest_stock_numbers
            
            # 找出售价>=599的商品货号
            high_price_stock_numbers = []
            for product in latest_products:
                price = WegoScraper.parse_price(product.get('售价', '') or product.get('price',''))
                if price and price >= 599:
                    stock_num = product.get('货号', '')
                    if stock_num:
                        high_price_stock_numbers.append(stock_num)
            
            # 筛选新增的高价商品
            high_price_added = []
            if high_price_stock_numbers and added:
                for stock_num in high_price_stock_numbers:
                    if stock_num in added:
                        high_price_added.append(stock_num)
            
            # 生成差异报告
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            date_str = datetime.now().strftime('%Y%m%d')
            
            diff_data = {
                'timestamp': timestamp,
                'date': date_str,
                'latest_file': os.path.basename(latest_json_file),
                'second_file': os.path.basename(second_latest_json_file),
                'added_count': len(added),
                'removed_count': len(removed),
                'added': sorted(list(added)),
                'removed': sorted(list(removed)),
                'high_price_added': sorted(list(high_price_added)),
                'high_price_added_count': len(high_price_added),
                'high_price_description': '新增的售价>=599的商品'
            }
            
            # 将差异信息追加到"小计"字段中
            if '小计' not in latest_json_data:
                latest_json_data['小计'] = []
            
            # 处理"小计"字段的不同格式
            if isinstance(latest_json_data['小计'], str):
                # 如果是字符串，先保存为字典
                old_summary = latest_json_data['小计']
                latest_json_data['小计'] = []
                # 尝试解析字符串中的信息（如果有）
                if old_summary and old_summary != "数据无变化":
                    # 创建一个基础记录
                    base_record = {
                        'timestamp': current_time,
                        'date': date_str,
                        'description': old_summary
                    }
                    latest_json_data['小计'].append(base_record)
            elif isinstance(latest_json_data['小计'], dict):
                # 如果是字典，转换为列表
                latest_json_data['小计'] = [latest_json_data['小计']]
            elif not isinstance(latest_json_data['小计'], list):
                # 其他类型，初始化为列表
                latest_json_data['小计'] = []
            
            # 追加新的差异记录
            latest_json_data['小计'].append(diff_data)
            
            # 按时间戳排序
            latest_json_data['小计'].sort(key=lambda x: x['timestamp'])
            
            FileManager.write_json(latest_json_file, latest_json_data)
            print(f'\n对比差异已追加到 {latest_json_file}')
            print(f'当前共有 {len(latest_json_data["小计"])} 条对比记录')
            
            # 打印对比结果
            print_separator()
            print('对比结果')
            print_separator()
            print(f'对比文件: {os.path.basename(second_latest_json_file)} -> {os.path.basename(latest_json_file)}')
            print(f'新增商品数: {len(added)}')
            print(f'删除商品数: {len(removed)}')
            print(f'新增高价商品数: {len(high_price_added)}')
            print('='*60)
            
            if added:
                print('\n新增的商品:')
                for i, num in enumerate(added, 1):
                    print(f'  {i}. {num}')
            
            if removed:
                print('\n删除的商品:')
                for i, num in enumerate(removed, 1):
                    print(f'  {i}. {num}')
            
            if high_price_added:
                print(f'\n新增的售价>=599的商品:')
                for i, num in enumerate(high_price_added, 1):
                    print(f'  {i}. {num}')
            
            print('='*60 + '\n')
            
            # 不删除缓存文件，保留用于后续对比
            # 缓存文件会在下一次运行爬虫时被覆盖
            if is_cache_used:
                print(f'注意：缓存文件 {second_latest_json_file} 已保留，用于后续对比')
                print(f'提示：下次运行爬虫时会自动更新缓存文件')
            
            return True
        except Exception as e:
            handle_exception(e, 'compare_json_files对比JSON文件')
            return False

    def compare_excel_with_json(self):
        try:
            print_separator()
            print('Excel与JSON数据对比工具')
            print_separator()
            
            latest_json_file = FileManager.get_latest_json_file()
            if not latest_json_file:
                print('无法获取最新的JSON文件')
                return False
            
            json_data = FileManager.read_json(latest_json_file)
            if not json_data:
                print('无法读取JSON文件')
                return False
            
            if isinstance(json_data, dict) and '商品列表' in json_data:
                products = json_data['商品列表']
            else:
                products = json_data if isinstance(json_data, list) else []
            
            json_stock_numbers = self.extract_stock_numbers(products)
            print(f'从JSON文件中读取到 {len(json_stock_numbers)} 个货号\n')
            
            high_price_stock_numbers = [
                p.get('货号', '') for p in products 
                if WegoScraper.parse_price(p.get('售价', '') or p.get('price','')) is not None
                and WegoScraper.parse_price(p.get('售价', '') or p.get('price','')) >= 599
            ]
            
            excel_stock_numbers = self.load_all_excel_data(remove_duplicates=False)
            if not excel_stock_numbers:
                print('无法从Excel文件读取货号')
                return False
            
            json_set, excel_set = set(json_stock_numbers), set(excel_stock_numbers)
            
            # 高价商品与已存在货号的对比
            high_price_set = set(high_price_stock_numbers)
            existing_set = json_set & excel_set  # 已存在的货号
            
            # 高价商品中已存在于Excel的货号
            high_price_existing = sorted(list(high_price_set & existing_set))
            # 高价商品中不在Excel的货号（多余的）
            high_price_extra = sorted(list(high_price_set - excel_set))
            
            # 保存对比结果
            result = self.compare_stock_numbers(json_stock_numbers, excel_stock_numbers, high_price_extra)
            result['high_price_existing'] = high_price_existing
            result['high_price_existing_count'] = len(high_price_existing)
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            date_str = datetime.now().strftime('%Y%m%d')
            
            data_change = "数据无变化" if result['missing_count'] == 0 and result['extra_in_json_count'] == 0 else f"数据有变化：缺失 {result['missing_count']} 个货号，多余 {result['extra_in_json_count']} 个货号"
            
            added_products = sorted([p.get('货号') for p in products if p.get('货号') and p.get('货号') not in excel_set]) if products else []
            
            removed_products = []
            prev_file = FileManager.get_latest_json_file('微购相册')
            if prev_file and prev_file != latest_json_file:
                old_data = FileManager.read_json(prev_file)
                if old_data and isinstance(old_data, dict):
                    removed_products = sorted([item.get('货号') for item in old_data.get('商品列表', []) if item.get('货号') and item.get('货号') not in json_set])
            
            duplicates = self.find_duplicate_stock_numbers(excel_stock_numbers)
            
            diff_data = {
                'timestamp': timestamp,
                'date': date_str,
                'json_file': os.path.basename(latest_json_file),
                'excel_file': os.path.basename(self.excel_file) if self.excel_file else 'None',
                'comparison': {
                    'missing_description': '微购相册比本地表格多出的序列号仅供参考',
                    'existing_description': '本地表格比微购相册上多的序列号，请仔细核对后删除多出的地方',
                    'extra_in_json_description': '微购相册比本地表格多出的序列号',
                    'high_price_extra_in_json': high_price_extra,
                    'high_price_extra_in_json_count': len(high_price_extra),
                    'high_price_extra_in_json_description': '只在JSON中存在但不在Excel中的售价>=599的货号',
                    **result
                },
                'result_message': self.get_result_message(result, duplicates),
                'data_change': data_change,
                'duplicates': duplicates,
                'duplicates_count': len(duplicates),
                '新增商品': added_products,
                '新增商品数量': len(added_products),
                '删除商品': removed_products,
                '删除商品数量': len(removed_products)
            }
            
            self._save_diff_log(date_str, diff_data)
            self._add_high_price_info_to_json(latest_json_file, json_data, high_price_extra)
            self._add_diff_to_json_summary(latest_json_file, json_data, diff_data)
            
            self.print_comparison_result(result, duplicates)
            return True
        except Exception as e:
            print(f'对比失败: {e}')
            traceback.print_exc()
            return False

    def _get_product_detail(self, item):
        return {
            "商品名称": item.get('商品名称', ''),
            "售价": item.get('售价', ''),
            "货号": item.get('货号', ''),
            "备注": item.get('备注', ''),
            "员工": item.get('员工', '')
        }
    
    def _save_diff_log(self, date_str, diff_data):
        diff_log_file = PathManager.get_diff_log_file(date_str)
        existing_data = FileManager.read_json(diff_log_file) if FileManager.file_exists(diff_log_file) else {'logs': []}
        
        if isinstance(existing_data, list):
            existing_data.append(diff_data)
        elif isinstance(existing_data, dict) and 'logs' in existing_data:
            existing_data['logs'].append(diff_data)
        else:
            existing_data = {'logs': [existing_data, diff_data]}
        
        FileManager.write_json(diff_log_file, existing_data)
        print(f'\n差异日志已保存到 {diff_log_file}')
    
    def _add_high_price_info_to_json(self, json_file_path, json_data, high_price_stock_numbers):
        if not json_data or not isinstance(json_data, dict):
            return
        
        products = json_data.get('商品列表', [])
        if not products:
            return
        
        high_price_count = 0
        for product in products:
            stock_num = product.get('货号', '')
            if stock_num in high_price_stock_numbers:
                product['备注'] = f'高价商品(≥599) - 只在JSON中存在但不在Excel中'
                high_price_count += 1
        
        if '统计信息' not in json_data:
            json_data['统计信息'] = {}
        
        json_data['统计信息']['高价商品数量'] = high_price_count
        json_data['统计信息']['高价商品货号'] = high_price_stock_numbers
        json_data['统计信息']['高价商品描述'] = '只在JSON中存在但不在Excel中的售价>=599的货号'
        
        FileManager.write_json(json_file_path, json_data)
        print(f'已为 {high_price_count} 个高价商品添加备注，并更新统计信息到 {json_file_path}')
    
    def _add_diff_to_json_summary(self, json_file_path, json_data, diff_data):
        if not json_data or not isinstance(json_data, dict):
            return
        
        if '小计' not in json_data:
            json_data['小计'] = []
        elif not isinstance(json_data['小计'], list):
            json_data['小计'] = [json_data['小计']] if json_data['小计'] else []
        
        excel_diff_record = {
            'timestamp': diff_data['timestamp'],
            'date': diff_data['date'],
            'json_file': diff_data['json_file'],
            'excel_file': diff_data['excel_file'],
            'comparison_type': 'Excel与JSON对比',
            'missing_count': diff_data['comparison']['missing_count'],
            'extra_in_json_count': diff_data['comparison']['extra_in_json_count'],
            'high_price_extra_in_json_count': diff_data['comparison']['high_price_extra_in_json_count'],
            'added': diff_data.get('新增商品', []),
            'removed': diff_data.get('删除商品', []),
            'added_products_count': diff_data['新增商品数量'],
            'removed_products_count': diff_data['删除商品数量'],
            'data_change': diff_data['data_change'],
            'result_message': diff_data['result_message']
        }
        
        json_data['小计'].append(excel_diff_record)
        json_data['小计'].sort(key=lambda x: x['timestamp'])
        
        FileManager.write_json(json_file_path, json_data)
        print(f'Excel对比记录已追加到 {json_file_path} 的"小计"字段')

    @staticmethod
    def get_result_message(result, duplicates):
        if result['missing_count'] == 0 and len(duplicates) == 0:
            return '对比结果: 成功 - 所有货号都存在且无重复'
        elif result['missing_count'] > 0 and len(duplicates) == 0:
            return f'对比结果: 部分成功 - 缺失 {result["missing_count"]} 个货号'
        elif result['missing_count'] == 0 and len(duplicates) > 0:
            return f'对比结果: 部分成功 - 发现 {len(duplicates)} 个重复序列号'
        else:
            return f'对比结果: 失败 - 缺失 {result["missing_count"]} 个货号，发现 {len(duplicates)} 个重复序列号'

    @staticmethod
    def print_comparison_result(result, duplicates):
        print('\n' + '='*60)
        print('货号对比结果')
        print('='*60)
        print(f'输入货号总数: {result["total_input"]}')
        print(f'JSON中货号总数: {result["total_json"]}')
        print(f'已存在货号数: {result["existing_count"]}')
        print(f'缺失货号数: {result["missing_count"]}')
        print(f'JSON中多余货号数: {result.get("extra_in_json_count", 0)}')
        print(f'重复序列号数: {len(duplicates)}')
        if result.get('high_price_count'):
            print(f'只在JSON中存在但不在Excel中的售价>=599货号数: {result["high_price_count"]}')
        
        print('='*60)
        
        if result['existing']:
            print('\n已存在的货号:')
            for i, num in enumerate(result['existing'], 1):
                print(f'  {i}. {num}')
        
        if result['missing']:
            print('\n缺失的货号:')
            for i, num in enumerate(result['missing'], 1):
                print(f'  {i}. {num}')
        else:
            print('\n所有输入货号都已存在！')
        
        # 显示高价商品(≥599)与已存在货号的对比结果
        print('\n=== 高价商品(≥599)与已存在货号对比 ===')
        print(f"高价商品总数: {len(result.get('high_price_stock_numbers', []))}")
        print(f"已存在于Excel的高价商品: {result.get('high_price_existing_count', 0)}")
        print(f"不在Excel的高价商品(多余): {result.get('high_price_count', 0)}")
        
        if result.get('high_price_existing'):
            print('\n高价商品中已存在于Excel的货号:')
            for i, num in enumerate(result['high_price_existing'], 1):
                print(f'  {num}', end=', ' if i % 5 != 0 else '\n')
            print()
        
        if result.get('high_price_stock_numbers'):
            print('\nJSON多余货号列表(高价商品):')
            for i, num in enumerate(result['high_price_stock_numbers'], 1):
                print(f'  {num}', end=', ' if i % 5 != 0 else '\n')
            print()
        else:
            print('\nJSON多余货号列表(高价商品): 无')
        
        if duplicates:
            print('\n重复的序列号:')
            for i, dup in enumerate(duplicates, 1):
                print(f'  {i}. 序列号: {dup["货号"]} (重复次数: {dup["count"]})')
        
        print('='*60)
        
        print('\n' + '='*60)
        print(StockNumberComparator.get_result_message(result, duplicates))
        print('='*60 + '\n')

    def run_comparison(self):
        print('='*60)
        print('货号对比工具 (TXT文件对比JSON)')
        print('='*60)
        
        data = self.load_json_data()
        json_stock_numbers = self.extract_stock_numbers(data)
        print(f'已从JSON加载 {len(json_stock_numbers)} 个货号\n')
        
        input_stock_numbers = None
        input_source = None
        
        if FileManager.file_exists(self.input_file):
            print(f'从文件 {self.input_file} 读取输入...')
            input_str = FileManager.read_text(self.input_file)
            if input_str:
                input_stock_numbers = self.parse_input_string(input_str)
                if input_stock_numbers:
                    input_source = 'TXT'
                    print(f'从TXT文件解析出 {len(input_stock_numbers)} 个货号\n')
        
        if not input_stock_numbers:
            print('未找到自动输入源，进入交互模式')
            print('功能说明：')
            print('  1. 直接输入货号字符串')
            print('  2. 输入 "load" 从本地文件读取')
            print('  3. 输入 "quit" 或 "exit" 退出程序')
            print('\n输入格式支持：')
            print('  - 逗号分隔: 12345, 67890, 11111')
            print('  - 空格分隔: 12345 67890 11111')
            print('  - 混合分隔: 12345, 67890 11111; 22222')
            print('='*60 + '\n')
            
            while True:
                try:
                    user_input = input('请输入货号字符串 (输入 "help" 查看帮助): ').strip()
                except (EOFError, KeyboardInterrupt):
                    print('\n程序已退出')
                    return
                
                if user_input.lower() in ['quit', 'exit', 'q', '退出']:
                    print('\n程序已退出')
                    return
                
                if user_input.lower() in ['help', 'h', '帮助']:
                    print('\n帮助信息：')
                    print('  load    - 从本地文件读取')
                    print('  quit    - 退出程序')
                    print('  help    - 显示此帮助信息')
                    print('\n输入格式：')
                    print('  - 直接输入货号字符串')
                    print('  - 支持多种分隔符')
                    print('  - 自动检测重复序列号')
                    print('='*60 + '\n')
                    continue
                
                if user_input.lower() in ['load', '读取']:
                    file_content = self.load_input_from_file()
                    if file_content:
                        input_stock_numbers = self.parse_input_string(file_content)
                        if input_stock_numbers:
                            input_source = 'TXT'
                            print(f'从TXT文件读取到 {len(input_stock_numbers)} 个货号\n')
                            break
                    else:
                        print('未找到TXT输入文件\n')
                
                if user_input:
                    input_stock_numbers = self.parse_input_string(user_input)
                    if input_stock_numbers:
                        input_source = '手动输入'
                        print(f'解析出 {len(input_stock_numbers)} 个货号\n')
                        break
        
        if input_stock_numbers:
            duplicates = self.find_duplicate_stock_numbers(input_stock_numbers)
            if duplicates:
                self.save_duplicate_log(duplicates)
            
            result = self.compare_stock_numbers(json_stock_numbers, input_stock_numbers)
            print(f'\n对比来源: {input_source}')
            self.print_comparison_result(result, duplicates)
        else:
            print('未找到有效的货号输入')


def main():
    while True:
        print_separator()
        print(f'Szwego商品爬虫和货号对比工具 - v{VERSION}')
        print_separator()
        
        config_file = PathManager.get_config_file()
        if not FileManager.file_exists(config_file):
            print(f'⚠️  警告: 配置文件不存在 ({config_file})')
            print(f'请先配置 {config_file} 文件')
            print_separator()
            input('按回车键退出...')
            return
        
        cookie_file = PathManager.get_cookie_file()
        is_valid, _ = CookieValidator.validate_and_prompt(cookie_file)
        
        if not is_valid:
            print('\n⚠️  Cookie状态异常，建议先更新Cookie')
            print_separator()
        
        print('请选择功能：')
        print('1. 运行爬虫（自动对比当天JSON文件）')
        print('2. 货号对比')
        print('3. Excel与JSON对比（自动保存差异日志）')
        print('4. 更新Cookie（自动更新）')
        print('5. 启动Web服务（可视化界面）')
        print('6. 文件清理工具')
        print('0. 退出')
        print_separator()
        
        try:
            choice = input('请输入选项 (0-6): ').strip()
        except (EOFError, KeyboardInterrupt):
            print('\n程序已退出')
            return
        
        def start_web():
            print('\n正在启动Web服务...')
            print(f'访问地址: http://localhost:{args.port if "args" in dir() and hasattr(args, "port") else 8888} (默认端口)')
            print('按 Ctrl+C 停止服务\n')

            subprocess.Popen([VENV_PYTHON, 'main.py', '--web'])
        
        actions = {
            '1': lambda: run_scraper() or True,
            '2': lambda: StockNumberComparator().run_comparison() or True,
            '3': lambda: StockNumberComparator().compare_excel_with_json() or True,
            '4': lambda: update_cookie() or True,
            '5': lambda: start_web() or True,
            '6': lambda: run_cleaner() or True,
        }
        
        if choice == '0':
            print('程序已退出')
            break
        elif choice in actions:
            actions[choice]()
        else:
            print('无效的选项')
            input('按回车键继续...')


def run_scraper():
    """运行爬虫"""
    try:
        # 直接运行爬虫，使用现有的cookie
        print('准备启动爬虫，将使用现有Cookie...')
        
        scraper = WegoScraper()
        asyncio.run(scraper.run())
    except Exception as e:
        handle_exception(e, 'run_scraper运行爬虫')
        input('按回车键继续...')


def update_cookie():
    """自动更新Cookie功能"""
    print_separator()
    print('Cookie自动更新工具')
    print_separator()
    print('说明：')
    print('  - 自动打开浏览器并获取最新Cookie')
    print('  - 用户手动登录后关闭浏览器')
    print('  - 完成后自动保存到配置文件')
    print_separator()
    
    if async_playwright is None:
        print("错误: 请先安装playwright")
        print("运行: pip install playwright && playwright install chromium")
        return
    
    cookie_file = PathManager.get_cookie_file()
    if FileManager.file_exists(cookie_file):
        print(f'清空现有Cookie文件: {cookie_file}')
        FileManager.write_json(cookie_file, [])
        print('[OK] Cookie文件已清空')
    
    async def get_cookie():
        async with async_playwright() as p:
            browser_args = WegoScraper.get_browser_args()
            chrome_path = WegoScraper.get_chrome_path()
            
            print(f'检测到系统: {Environment.SYSTEM}')
            if chrome_path:
                print(f'使用系统Chrome: {chrome_path}')
            else:
                print(f'使用Playwright内置Chromium')
            
            try:
                browser = await p.chromium.launch(
                    headless=False, 
                    args=browser_args, 
                    executable_path=chrome_path
                )
            except Exception as launch_err:
                if "Executable doesn't exist" in str(launch_err) or "executable doesn't exist" in str(launch_err).lower():
                    print('Playwright内置Chromium不存在，正在自动安装浏览器...')
                    install_playwright_cdn()
                    try:
                        browser = await p.chromium.launch(
                            headless=False, 
                            args=browser_args, 
                            executable_path=None
                        )
                        print('浏览器安装完成，使用Playwright内置Chromium启动成功')
                    except Exception as retry_err:
                        fallback_chrome = Environment._find_system_chrome()
                        if fallback_chrome:
                            print(f'Playwright内置Chromium仍不可用，回退到系统Chrome: {fallback_chrome}')
                            browser = await p.chromium.launch(
                                headless=False, 
                                args=browser_args, 
                                executable_path=fallback_chrome
                            )
                        else:
                            print(f'自动安装失败且无系统Chrome可用: {retry_err}')
                            print('请手动运行: python -m playwright install chromium')
                            raise
                else:
                    raise
            
            context = await browser.new_context(
                viewport=Environment.get_default_viewport(),
                user_agent=WegoScraper.get_user_agent()
            )
            
            existing_cookies = []
            cookie_file = PathManager.get_cookie_file()
            if FileManager.file_exists(cookie_file):
                existing_cookies = FileManager.read_json(cookie_file)
                if existing_cookies:
                    print(f'已加载 {len(existing_cookies)} 个现有Cookie')
                    await context.add_cookies(existing_cookies)
            
            page = await context.new_page()
            await page.goto('https://www.szwego.com', wait_until='networkidle')
            
            print('浏览器已打开，正在获取Cookie...')
            print('请稍候，系统会自动处理...')
            
            try:
                await page.wait_for_load_state('networkidle', timeout=10000)
                print('页面加载完成')
            except Exception as e:
                print(f'页面加载超时，继续获取Cookie: {e}')
            
            print_separator()
            print('浏览器已打开')
            print('请在浏览器中完成以下操作：')
            print('1. 如果需要登录，请完成登录')
            print('2. 登录后刷新一下页面')
            print('3. 程序会自动检测登录状态并关闭浏览器')
            print_separator()
            print('自动检测登录状态...')
            print_separator()
            
            start_time = time.time()
            timeout = 300
            login_detected = False
            
            while time.time() - start_time < timeout:
                try:
                    cookies = await context.cookies()
                    auth_cookies = [c for c in cookies if 'token' in c['name'].lower() or 'session' in c['name'].lower() or 'auth' in c['name'].lower()]
                    
                    if auth_cookies:
                        print('[OK] 检测到登录成功，自动关闭浏览器...')
                        login_detected = True
                        break
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass
                
                await asyncio.sleep(5)
                elapsed = int(time.time() - start_time)
                print(f'等待登录中... ({elapsed}秒)')
            
            if not login_detected:
                print('⚠️ 登录超时，尝试获取当前Cookie')
            
            cookies = await context.cookies()
            szwego_cookies = [cookie for cookie in cookies if 'szwego.com' in cookie['domain']]
            
            for cookie in szwego_cookies:
                if cookie['domain'] == 'www.szwego.com':
                    cookie['domain'] = '.szwego.com'
            
            FileManager.write_json(cookie_file, szwego_cookies)
            
            print(f'[OK] Cookie已保存到 {cookie_file}')
            print(f'[OK] 共保存 {len(szwego_cookies)} 个Cookie')
            
            print_separator()
            print('Cookie有效期信息：')
            token_cookie = next((c for c in szwego_cookies if c['name'] == 'token'), None)
            if token_cookie and 'expires' in token_cookie and token_cookie['expires']:
                expiry_time = datetime.fromtimestamp(token_cookie['expires'])
                expiry_str = expiry_time.strftime('%Y-%m-%d')
                print(f'Token有效期: {expiry_str}')
            else:
                print('未找到Token Cookie')
            print_separator()
            
            config_file = PathManager.get_config_file()
            if FileManager.file_exists(config_file):
                config_data = FileManager.read_json(config_file)
                
                cookie_header = '; '.join([f'{c["name"]}={c["value"]}' for c in szwego_cookies])
                
                if 'headers' not in config_data:
                    config_data['headers'] = {}
                config_data['headers']['cookie'] = cookie_header
                config_data['cookies'] = szwego_cookies
                
                FileManager.write_json(config_file, config_data)
                print('[OK] config.json中的Cookie已更新')
            
            await browser.close()
            print('[OK] 浏览器已自动关闭')
            
            return True
    
    try:
        asyncio.run(get_cookie())
        print('\n[OK] Cookie更新完成')
    except Exception as e:
        handle_exception(e, 'update_cookie更新Cookie')


def select_pip_mirror(venv_path: str):
    """pip镜像智能测速+写入配置"""
    MIRRORS = [
        ("阿里云", "https://mirrors.aliyun.com/pypi/simple/", "mirrors.aliyun.com"),
        ("清华", "https://pypi.tuna.tsinghua.edu.cn/simple/", "pypi.tuna.tsinghua.edu.cn"),
        ("腾讯云", "https://mirrors.cloud.tencent.com/pypi/simple/", "mirrors.cloud.tencent.com"),
        ("中科大", "https://mirrors.ustc.edu.cn/pypi/simple/", "mirrors.ustc.edu.cn"),
        ("豆瓣", "https://pypi.douban.com/simple/", "pypi.douban.com"),
    ]

    def test_mirror(name, url):
        try:
            start = time.time()
            urllib.request.urlopen(url, timeout=3)
            return round(time.time() - start, 3)
        except Exception as e:
            logger.warning(f'Mirror test failed for {name}: {e}')
            return None

    print("[*] 检测到未配置pip镜像源，正在测试镜像源速度...")
    os.makedirs(os.path.join(venv_path, "pip_config"), exist_ok=True)

    results = []
    for i, (name, url, host) in enumerate(MIRRORS, 1):
        print(f"[*] 测试镜像源 {i}/5: {name}...", end="", flush=True)
        elapsed = test_mirror(name, url)
        if elapsed is not None:
            print(f" {elapsed}秒")
            results.append((name, url, host, elapsed))
        else:
            print(" 失败")

    if results:
        results.sort(key=lambda x: x[3])
        fastest_name, fastest_mirror, fastest_host, fastest_time = results[0]
        print(f"[*] 最终选择最快镜像源: {fastest_name} ({fastest_time}秒)")
    else:
        fastest_name, fastest_mirror, fastest_host = "阿里云", "https://mirrors.aliyun.com/pypi/simple/", "mirrors.aliyun.com"
        print(f"[WARNING] 所有镜像源均失败，使用默认阿里云")

    conf_path = os.path.join(venv_path, "pip_config", "pip.conf" if platform.system() != "Windows" else "pip.ini")
    with open(conf_path, "w", encoding="utf-8") as f:
        if platform.system() != "Windows":
            f.write("[global]\n")
            f.write(f"index-url = {fastest_mirror}\n")
            f.write("[install]\n")
            f.write(f"trusted-host = {fastest_host}\n")
        else:
            f.write("[global]\r\n")
            f.write(f"index-url = {fastest_mirror}\r\n")
            f.write("[install]\r\n")
            f.write(f"trusted-host = {fastest_host}\r\n")
    print(f"[*] pip配置已写入: {conf_path}")


def check_deps_satisfied(requirements_file="requirements.txt"):
    def ver_tuple(v):
        return tuple(int(x) for x in re.split(r'[.\-]', v) if x.isdigit())

    try:
        with open(requirements_file, encoding="utf-8") as f:
            lines = [l.strip() for l in f if l.strip() and not l.strip().startswith("#")]
    except FileNotFoundError:
        print(f"[!] {requirements_file} not found")
        sys.exit(1)

    missing = []
    for line in lines:
        parts = re.split(r"[><=!~\s;]+", line)
        pkg_name = parts[0].replace("-", "_")
        req_ver = parts[1] if len(parts) > 1 else None

        installed_ver = None
        try:
            installed_ver = im.version(parts[0])
        except im.PackageNotFoundError:
            pass

        if installed_ver is None:
            missing.append(line)
        elif req_ver:
            try:
                if ver_tuple(installed_ver) < ver_tuple(req_ver):
                    missing.append(line)
            except (ValueError, IndexError):
                missing.append(line)

    if missing:
        print(f"[!] Missing/outdated: {missing}")
        sys.exit(1)
    else:
        print("[OK] All dependencies satisfied")
        sys.exit(0)


def install_playwright_cdn():
    """Playwright CDN智能测速+安装"""
    CDNS = [
        ("npmmirror", "https://npmmirror.com/mirrors/playwright"),
        ("azureedge", "https://playwright.azureedge.net/builds"),
        ("cdn", "https://cdn.playwright.dev"),
    ]

    def test_cdn(name, url):
        try:
            start = time.time()
            urllib.request.urlopen(url, timeout=3)
            return round(time.time() - start, 3)
        except Exception as e:
            print(f'[CDN] Test failed {name}: {e}')
            return None

    print("[*] 测试Playwright CDN速度...")
    results = []
    for name, url in CDNS:
        elapsed = test_cdn(name, url)
        if elapsed is not None:
            print(f"    {name}: {elapsed}秒")
            results.append((name, url, elapsed))
        else:
            print(f"    {name}: 失败")

    if results:
        results.sort(key=lambda x: x[2])
        fastest_name, fastest_url, fastest_time = results[0]
        print(f"[*] 最终选择最快Playwright CDN: {fastest_name} ({fastest_time}秒)")
        download_order = [(n, u) for n, u, _ in results]
    else:
        print("[WARNING] 所有CDN连通性测试失败，使用官方CDN下载")
        download_order = [("cdn", "https://cdn.playwright.dev")]

    print("[*] 安装Playwright浏览器...")
    installed = False
    for name, url in download_order:
        print(f"    尝试从 {name} 下载...", flush=True)
        env = os.environ.copy()
        env["PLAYWRIGHT_DOWNLOAD_HOST"] = url.rstrip("/")
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True, text=True, env=env
        )
        if result.returncode == 0:
            print(f"[*] Playwright浏览器安装成功 (来源: {name})")
            installed = True
            break
        if result.stderr:
            for err_line in result.stderr.split('\n'):
                if 'Error:' in err_line or '404' in err_line:
                    print(f"    {err_line.strip()}")
        print(f"    {name} 下载失败，尝试下一个...")

    if not installed:
        print("[WARNING] Playwright浏览器安装失败，将尝试使用系统Chrome")


if __name__ == '__main__':
    check_python_version((3, 0))  # Python 版本兼容性检查 (>=3.0)
    
    parser = argparse.ArgumentParser(description='Szwego商品爬虫')
    parser.add_argument('--web', action='store_true', help='启动Web服务模式')
    parser.add_argument('--port', type=int, default=int(os.environ.get('WEB_PORT', '8888')), help=f'Web服务端口 (默认{os.environ.get("WEB_PORT", "8888")})')
    parser.add_argument('--setup', action='store_true', help='运行配置初始化向导')
    parser.add_argument('--username', '-u', help='登录用户名')
    parser.add_argument('--password', '-p', help='登录密码')
    parser.add_argument('--url', '-l', help='目标店铺URL')
    parser.add_argument('--excel', '-e', help='Excel文件路径')
    parser.add_argument('--task', type=int, choices=[1, 2, 3, 4, 6], help='直接执行指定任务后退出 (1:爬虫, 2:货号对比, 3:Excel对比, 4:更新Cookie, 6:文件清理)')
    parser.add_argument('--install-playwright', action='store_true', help='Playwright CDN智能测速+安装浏览器')
    parser.add_argument('--check-deps', action='store_true', help='检查requirements.txt依赖是否已满足')
    parser.add_argument('--select-pip-mirror', action='store_true', help='pip镜像智能测速并写入配置')
    args = parser.parse_args()

    if args.select_pip_mirror:
        venv_path = os.environ.get('VIRTUAL_ENV') or os.path.join(PROJECT_DIR, '.venv')
        select_pip_mirror(venv_path)
        sys.exit(0)

    if args.check_deps:
        check_deps_satisfied()
        sys.exit(0)

    if args.install_playwright:
        install_playwright_cdn()
        sys.exit(0)

    if args.setup or (args.username and args.password and args.url):
        print("=" * 50)
        print("配置初始化向导")
        print("=" * 50)

        username = args.username
        password = args.password
        url = args.url
        excel = args.excel or ""

        if not username:
            username = input("请输入用户名: ")
        if not password:
            password = input("请输入密码: ")
        if not url:
            url = input("请输入目标URL: ")

        print("\n正在启动浏览器进行登录...")
        print("请在浏览器中登录您的账号")

        os.makedirs("config", exist_ok=True)

        async def get_cookie():
            async with async_playwright() as p:
                try:
                    browser = await p.chromium.launch(headless=False)
                except Exception as launch_err:
                    if "Executable doesn't exist" in str(launch_err) or "executable doesn't exist" in str(launch_err).lower():
                        print('Playwright内置Chromium不存在，正在自动安装浏览器...')
                        install_playwright_cdn()
                        try:
                            browser = await p.chromium.launch(headless=False)
                            print('浏览器安装完成，使用Playwright内置Chromium启动成功')
                        except Exception as retry_err:
                            fallback_chrome = Environment._find_system_chrome()
                            if fallback_chrome:
                                print(f'Playwright内置Chromium仍不可用，回退到系统Chrome: {fallback_chrome}')
                                browser = await p.chromium.launch(headless=False, executable_path=fallback_chrome)
                            else:
                                print(f'自动安装失败且无系统Chrome可用: {retry_err}')
                                print('请手动运行: python -m playwright install chromium')
                                raise
                    else:
                        raise
                context = await browser.new_context(
                    viewport=Environment.get_default_viewport(),
                    user_agent=WegoScraper.get_user_agent()
                )
                page = await context.new_page()
                await page.goto('https://www.szwego.com', wait_until='networkidle')

                print('\n请在浏览器中登录您的账号...')

                start_time = time.time()
                timeout = 300
                login_detected = False

                while time.time() - start_time < timeout:
                    try:
                        cookies = await context.cookies()
                        token_cookie = next((c for c in cookies if c['name'] == 'token'), None)
                        if token_cookie and token_cookie['value']:
                            print('\n[OK] 检测到登录成功！正在获取Cookie...')
                            login_detected = True
                            break
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
                    await asyncio.sleep(3)
                    elapsed = int(time.time() - start_time)
                    print(f'等待登录中... ({elapsed}秒)', end='\r')

                if not login_detected:
                    print('\n⚠️ 登录超时，将尝试获取当前Cookie')

                cookies = await context.cookies()
                szwego_cookies = [cookie for cookie in cookies if 'szwego.com' in cookie.get('domain', '')]
                for cookie in szwego_cookies:
                    if cookie['domain'] == 'www.szwego.com':
                        cookie['domain'] = '.szwego.com'

                cookie_file = "config/cookies.json"
                with open(cookie_file, "w", encoding="utf-8") as f:
                    json.dump(szwego_cookies, f, ensure_ascii=False, indent=2)
                print(f'[OK] Cookie已保存 ({len(szwego_cookies)}个)')

                await browser.close()
                return szwego_cookies

        if async_playwright is None:
            print("错误: 请先安装playwright")
            print("运行: pip install playwright && playwright install chromium")
            sys.exit(1)

        cookies = asyncio.run(get_cookie())

        cookie_header = '; '.join([f'{c["name"]}={c["value"]}' for c in cookies])

        config = {
            "login": {
                "username": username,
                "password": password,
                "login_type": "phone"
            },
            "target_url": url,
            "scroll_config": {
                "max_attempts": 30,
                "same_height_limit": 8,
                "scroll_wait_time": 0.8,
                "popup_close_interval": 5,
                "popup_close_limit": 3,
                "popup_close_wait": 0.3,
                "dynamic_adjust": True
            },
            "headers": {
                "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                "accept-encoding": "gzip, deflate, br, zstd",
                "accept-language": "zh-CN,zh;q=0.9",
                "cache-control": "max-age=0",
                "connection": "keep-alive",
                "cookie": cookie_header,
                "host": "www.szwego.com",
                "sec-ch-ua": "\"Chromium\";v=\"146\", \"Not-A.Brand\";v=\"24\", \"Google Chrome\";v=\"146\"",
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": f"\"{Environment.SYSTEM}\"",
                "sec-fetch-dest": "document",
                "sec-fetch-mode": "navigate",
                "sec-fetch-site": "none",
                "sec-fetch-user": "?1",
                "upgrade-insecure-requests": "1",
                "user-agent": Environment.get_user_agent()
            },
            "cookies": cookies,
            "output_file": "file/output.json",
            "cookie_file": "config/cookies.json",
            "excel_files": [excel] if excel else []
        }

        config_path = "config/config.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        print(f'\n[OK] 配置文件创建成功！')
        print("\n========================================")
        print("   初始化完成！")
        print("========================================")
        print(f"\n配置信息：")
        print(f"  - 用户名: {username}")
        print(f"  - 目标URL: {url}")
        print(f"  - Cookie数量: {len(cookies)}")
        sys.exit(0)

    if args.task:
        print(f'执行任务 {args.task}...')
        try:
            if args.task == 1:
                run_scraper()
            elif args.task == 2:
                StockNumberComparator().run_comparison()
            elif args.task == 3:
                StockNumberComparator().compare_excel_with_json()
            elif args.task == 4:
                update_cookie()
            elif args.task == 6:
                run_cleaner()
            print('任务完成')
        except Exception as e:
            handle_exception(e, f'任务{args.task}执行')
        sys.exit(0)
    
    if args.web:
        # ============================================================
        # 请求日志中间件 (v3.8.70)
        # ============================================================
        _request_logger = logging.getLogger('api_requests')
        _request_logger.setLevel(logging.INFO)
        if not _request_logger.handlers:
            _rh = logging.StreamHandler()
            _rh.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
            _request_logger.addHandler(_rh)

        @app.middleware("http")
        async def _log_and_security_middleware(request: Request, call_next):
            start_time = time.time()

            path = request.url.path

            if not (path.startswith('/static') or path.startswith('/favicon')):
                client_ip = request.client.host if request.client else "unknown"
                _request_logger.info(f'[{request.method}] {path} | IP: {client_ip}')
                if request.method in ['POST', 'PUT', 'PATCH']:
                    content_length = request.headers.get('content-length', 0)
                    cl = int(content_length) if content_length else 0
                    if cl > 1024 * 1024:
                        _request_logger.warning(f'大请求体: {cl / 1024:.1f}KB')

            response = await call_next(request)

            if not path.startswith('/static'):
                if response.status_code >= 400:
                    _request_logger.warning(f'[{response.status_code}] {path}')

                duration = (time.time() - start_time) * 1000
                response.headers['X-Response-Time'] = f'{duration:.2f}ms'

                if REQUEST_LATENCY is not None:
                    try:
                        REQUEST_LATENCY.labels(request.method, path).observe(time.time() - start_time)
                    except Exception as e:                        logger.debug(f"Silent exception: {e}")

                if REQUEST_COUNT is not None:
                    try:
                        endpoint = getattr(request.state, 'endpoint', None) or path
                        REQUEST_COUNT.labels(request.method, endpoint, response.status_code).inc()
                    except Exception as e:                        logger.debug(f"Silent exception: {e}")

            response.headers['X-Content-Type-Options'] = 'nosniff'
            if path.startswith('/dist/'):
                response.headers['X-Frame-Options'] = 'SAMEORIGIN'
            else:
                response.headers['X-Frame-Options'] = 'DENY'
            response.headers['X-XSS-Protection'] = '1; mode=block'
            response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
            response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
            response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=(), payment=()'

            origin = request.headers.get('origin', '')
            allowed_origins = [
                'http://localhost:5000',
                'http://127.0.0.1:5000',
                'http://localhost:8080',
                'http://127.0.0.1:8080',
            ]

            if path.startswith('/api/') or path in ['/run', '/input', '/kill']:
                if origin and origin in allowed_origins:
                    response.headers['Access-Control-Allow-Origin'] = origin
                    response.headers['Access-Control-Allow-Credentials'] = 'true'
                elif not origin:
                    response.headers['Access-Control-Allow-Origin'] = '*'
                else:
                    response.headers['Access-Control-Allow-Origin'] = 'null'

                response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
                response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
                response.headers['Access-Control-Max-Age'] = '86400'

            if path == '/docs/':
                response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; img-src 'self' data: https:; media-src 'self' https:; font-src 'self' data:;"
            elif path == '/':
                response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://code.jquery.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; font-src 'self' data: https: cdn.jsdelivr.net; img-src 'self' data: https:; media-src 'self' https:;"
            elif path.startswith('/dist/'):
                response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:; media-src 'self' https:; font-src 'self' data:; connect-src 'self' https://api.bigdatacloud.net https://api.open-meteo.com https://air-quality-api.open-meteo.com;"
            elif not path.startswith('/api/'):
                response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:; media-src 'self' https:; font-src 'self' data:;"

            return response



        @app.get('/api/bootstrap')
        async def api_bootstrap(request: Request):
            return JSONResponse(content={'api_key': WEB_API_KEY})

        @app.get('/health')
        async def health_check():
            health_data = {
                'status': 'healthy',
                'version': '3.8.73',
                'timestamp': datetime.now().isoformat(),
            }
            status_code = 200
            if PSUTIL_AVAILABLE:
                try:
                    cpu_percent = psutil.cpu_percent(interval=0.1)
                    memory = psutil.virtual_memory()
                    disk = psutil.disk_usage('/')
                    health_data['cpu_percent'] = cpu_percent
                    health_data['memory_percent'] = memory.percent
                    health_data['memory_used_gb'] = round(memory.used / (1024**3), 2)
                    health_data['memory_total_gb'] = round(memory.total / (1024**3), 2)
                    health_data['disk_percent'] = disk.percent
                    if cpu_percent > 90 or memory.percent > 90 or disk.percent > 95:
                        health_data['status'] = 'degraded'
                    if cpu_percent > 95 or memory.percent > 95:
                        health_data['status'] = 'unhealthy'
                        status_code = 503
                except Exception as e:
                    health_data['system_check_error'] = str(e)
            else:
                health_data['system_check'] = 'psutil未安装'
            if json_cache is not None:
                try:
                    cache_stats = json_cache.get_stats()
                    health_data['cache'] = cache_stats
                except Exception as e:                    logger.debug(f"Silent exception: {e}")
            health_data['active_tasks'] = len(tasks) if 'tasks' in dir() else 0
            return JSONResponse(content=health_data, status_code=status_code)

        @app.get('/ready')
        async def readiness_check():
            try:
                return JSONResponse(content={'ready': True, 'timestamp': datetime.now().isoformat()}, status_code=200)
            except Exception as e:
                logger.error(f'Readiness check failed: {e}')
                return JSONResponse(content={'ready': False}, status_code=503)

        @app.get('/metrics')
        async def metrics_endpoint():
            if not PROMETHEUS_AVAILABLE:
                return JSONResponse(
                    status_code=404,
                    content={'error': 'prometheus_client未安装', 'hint': 'pip install prometheus_client'}
                )
            try:
                if ACTIVE_TASKS_GAUGE is not None:
                    try:
                        ACTIVE_TASKS_GAUGE.set(len(tasks) if 'tasks' in dir() else 0)
                    except Exception as e:                        logger.debug(f"Silent exception: {e}")
                return Response(content=generate_latest(), media_type=CONTENT_TYPE_LATEST)
            except Exception as e:
                return JSONResponse(status_code=500, content={'error': str(e)})

        @app.get('/api/security/check')
        async def security_check_endpoint():
            try:
                checker=SecurityChecker()
                report=checker.run_all_checks()
                return JSONResponse(content=report)
            except Exception as e:
                return JSONResponse(status_code=500,content={'error':f'安全检查失败: {type(e).__name__}'})

        @app.get('/api/security/audit')
        async def security_audit_endpoint():
            try:
                auditor=DependencyAuditor()
                report=auditor.audit()
                return JSONResponse(content=report)
            except Exception as e:
                return JSONResponse(status_code=500,content={'error':f'依赖审计失败: {type(e).__name__}'})

        @app.post('/api/security/encrypt-init')
        async def security_encrypt_init(request):
            try:
                body=await request.json()
                password=body.get('password','')
                success,msg=SecureConfigManager.initialize_encryption(password)
                return JSONResponse(content={'success':success,'message':msg})
            except Exception as e:
                return JSONResponse(status_code=500,content={'error':f'加密初始化失败: {type(e).__name__}'})

        @app.get('/api/swagger.json')
        async def swagger_spec():
            spec = {
                'openapi': '3.0.0',
                'info': {'title': 'Szwego商品爬虫API', 'version': '3.8.73', 'description': 'Szwego商品爬虫Web服务API文档'},
                'servers': [{'url': '/api'}],
                'paths': {
                    '/version': {'get': {'summary': '获取版本信息', 'tags': ['系统'], 'responses': {'200': {'description': '版本信息'}}}},
                    '/health': {'get': {'summary': '健康检查', 'description': 'CPU/内存/磁盘/缓存状态', 'tags': ['系统'], 'responses': {'200': {'description': '健康'}, '503': {'description': '异常'}}}},
                    '/ready': {'get': {'summary': '就绪检查', 'description': 'Kubernetes就绪探针', 'tags': ['系统'], 'responses': {'200': {'description': '就绪'}}}},
                    '/server/info': {'get': {'summary': '获取服务器信息', 'tags': ['系统'], 'responses': {'200': {'description': '服务器信息'}}}},
                    '/cookie': {'get': {'summary': '获取Cookie状态', 'tags': ['系统'], 'responses': {'200': {'description': 'Cookie状态'}}}},
                    '/changelog': {'get': {'summary': '获取更新日志', 'tags': ['系统'], 'responses': {'200': {'description': '更新日志'}}}},
                    '/run': {'post': {'summary': '执行命令', 'tags': ['任务'], 'requestBody': {'content': {'application/json': {'schema': {'type': 'object', 'required': ['command'], 'properties': {'command': {'type': 'string', 'description': '要执行的命令'}}}}}}, 'responses': {'200': {'description': '执行结果'}}}},
                    '/products': {'get': {'summary': '获取商品列表', 'tags': ['商品'], 'parameters': [{'name': 'date', 'in': 'query', 'schema': {'type': 'string'}, 'description': '日期'}], 'responses': {'200': {'description': '商品列表'}}}},
                    '/product/search': {'get': {'summary': '搜索商品', 'tags': ['商品'], 'parameters': [{'name': 'keyword', 'in': 'query', 'schema': {'type': 'string'}, 'description': '搜索关键词'}], 'responses': {'200': {'description': '搜索结果'}}}},
                    '/daily-profit': {'get': {'summary': '获取每日利润', 'tags': ['商品'], 'responses': {'200': {'description': '利润数据'}}}},
                    '/sku/compare': {'get': {'summary': 'SKU对比', 'tags': ['商品'], 'responses': {'200': {'description': '对比结果'}}}},
                    '/tunnel/status': {'get': {'summary': '获取隧道状态', 'tags': ['隧道'], 'responses': {'200': {'description': '隧道状态'}}}},
                    '/tunnel/start': {'post': {'summary': '启动隧道', 'tags': ['隧道'], 'responses': {'200': {'description': '启动结果'}}}},
                    '/tunnel/stop': {'post': {'summary': '停止隧道', 'tags': ['隧道'], 'responses': {'200': {'description': '停止结果'}}}},
                    '/email/config': {'get': {'summary': '获取邮件配置', 'tags': ['邮件'], 'responses': {'200': {'description': '邮件配置'}}}, 'post': {'summary': '保存邮件配置', 'tags': ['邮件'], 'responses': {'200': {'description': '保存结果'}}}},
                    '/email/test': {'post': {'summary': '测试邮件发送', 'tags': ['邮件'], 'responses': {'200': {'description': '测试结果'}}}},
                },
                'tags': [
                    {'name': '系统', 'description': '系统管理相关接口'},
                    {'name': '商品', 'description': '商品管理相关接口'},
                    {'name': '任务', 'description': '任务管理相关接口'},
                    {'name': '隧道', 'description': '隧道管理相关接口'},
                    {'name': '邮件', 'description': '邮件管理相关接口'},
                ]
            }
            return JSONResponse(content=spec)

        @app.get('/docs/')
        async def swagger_ui():
            html_content = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="default">
    <title>Szwego商品爬虫API文档</title>
    <link rel="stylesheet" type="text/css" href="https://cdn.jsdelivr.net/npm/swagger-ui-dist@5/swagger-ui.css">
    <style>
        /* 移动端适配 */
        @media screen and (max-width: 768px) {
            .swagger-ui .topbar {
                padding: 8px 10px;
            }
            .swagger-ui .topbar-wrapper img {
                max-width: 150px;
            }
            .swagger-ui .information-container {
                padding: 10px;
            }
            .swagger-ui .scheme-container {
                padding: 10px 0;
            }
            .swagger-ui .opblock-tag {
                padding: 10px;
            }
            .swagger-ui .opblock .opblock-summary {
                padding: 10px;
            }
            .swagger-ui .opblock-body {
                padding: 10px;
            }
            .swagger-ui .parameters-col_description {
                width: 100%;
                padding: 5px 0;
            }
            .swagger-ui .parameters-col_name {
                width: 100%;
                padding: 5px 0;
            }
            .swagger-ui .response-col_description {
                width: 100%;
                padding: 5px 0;
            }
            .swagger-ui .execute-wrapper {
                padding: 10px;
            }
            .swagger-ui .btn {
                padding: 8px 12px;
                font-size: 14px;
            }
            .swagger-ui input[type=text],
            .swagger-ui input[type=password],
            .swagger-ui input[type=email],
            .swagger-ui input[type=file],
            .swagger-ui textarea {
                font-size: 16px;
            }
            .swagger-ui .opblock .opblock-section {
                padding: 10px;
            }
            .swagger-ui .opblock .opblock-section-header {
                padding: 8px 10px;
            }
            .swagger-ui table thead tr th,
            .swagger-ui table thead tr td,
            .swagger-ui table tbody tr th,
            .swagger-ui table tbody tr td {
                padding: 8px;
            }
            .swagger-ui .response-col_status {
                width: auto;
                min-width: 80px;
            }
            .swagger-ui .col.response-headers {
                width: auto;
                min-width: 80px;
            }
            .swagger-ui .wrapper {
                padding: 10px;
            }
            .swagger-ui .opblock-tag-section {
                margin: 0 0 10px 0;
            }
            .swagger-ui .opblock.is-open .opblock-summary {
                border-bottom: 1px solid #ebeef2;
            }
            body {
                margin: 0;
                padding: 0;
            }
        }

        /* 超小屏幕优化 */
        @media screen and (max-width: 480px) {
            .swagger-ui .topbar {
                padding: 5px;
            }
            .swagger-ui .information-container {
                padding: 5px;
            }
            .swagger-ui .opblock-tag {
                font-size: 18px;
                padding: 8px;
            }
            .swagger-ui .opblock .opblock-summary {
                padding: 8px;
            }
            .swagger-ui .opblock .opblock-summary-description {
                font-size: 13px;
            }
            .swagger-ui .btn {
                padding: 6px 10px;
                font-size: 13px;
            }
            .swagger-ui .wrapper {
                padding: 5px;
            }
        }

        /* 横屏优化 */
        @media screen and (max-width: 768px) and (orientation: landscape) {
            .swagger-ui .opblock .opblock-summary {
                display: flex;
                align-items: center;
            }
            .swagger-ui .opblock .opblock-summary-method {
                margin-right: 10px;
            }
        }
    </style>
</head>
<body>
    <div id="swagger-ui"></div>
    <script src="https://cdn.jsdelivr.net/npm/swagger-ui-dist@5/swagger-ui-bundle.js"></script>
    <script>
    SwaggerUIBundle({
        url: "/api/swagger.json",
        dom_id: '#swagger-ui',
        presets: [SwaggerUIBundle.presets.apis, SwaggerUIBundle.SwaggerUIStandalonePreset],
        layout: "BaseLayout",
        deepLinking: true,
        displayOperationId: false,
        defaultModelsExpandDepth: 1,
        defaultModelExpandDepth: 1,
        docExpansion: 'list',
        filter: true
    })
    </script>
</body>
</html>'''
            return HTMLResponse(content=html_content)

        @app.api_route('/', methods=['GET', 'HEAD'])
        async def index():
            current_version = get_version_from_readme()
            with open(os.path.join(PROJECT_DIR, 'index.html'), 'r', encoding='utf-8') as f:
                content = f.read()
            content = re.sub(r'版本:\s*[\d.]+', f'版本: {current_version}', content)
            return HTMLResponse(
                content=content,
                headers={
                    'Cache-Control': 'no-cache, no-store, must-revalidate',
                    'Pragma': 'no-cache',
                    'Expires': '0'
                }
            )

        @app.get('/dist/{filename:path}')
        async def dist_files(filename: str, request: Request):
            safe_path, err = sec_sp(os.path.join(PROJECT_DIR, 'dist'), filename)
            if not safe_path:
                raise HTTPException(status_code=403, detail=f"Path traversal blocked: {err}")
            file_path = safe_path

            if not os.path.isfile(file_path):
                raise HTTPException(status_code=404, detail="File not found")

            mimetype_map = {
                '.js': 'application/javascript',
                '.css': 'text/css',
                '.html': 'text/html',
                '.json': 'application/json',
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.svg': 'image/svg+xml',
                '.ico': 'image/x-icon',
                '.woff': 'font/woff',
                '.woff2': 'font/woff2',
            }
            ext = os.path.splitext(filename)[1].lower()
            mimetype = mimetype_map.get(ext, 'application/octet-stream')

            accept_encoding = request.headers.get('accept-encoding', '')
            gzip_enabled = 'gzip' in accept_encoding.lower()

            if gzip_enabled and ext in ['.js', '.css', '.html', '.json', '.svg']:
                with open(file_path, 'rb') as f:
                    content = f.read()

                gzip_content = gzip.compress(content, compresslevel=6)
                return Response(
                    content=gzip_content,
                    media_type=mimetype,
                    headers={
                        'Content-Encoding': 'gzip',
                        'Vary': 'Accept-Encoding',
                        'Cache-Control': 'no-cache, no-store, must-revalidate',
                        'Pragma': 'no-cache',
                        'Expires': '0'
                    }
                )
            else:
                return FileResponse(
                    path=file_path,
                    media_type=mimetype,
                    headers={'Cache-Control': 'no-cache, no-store, must-revalidate', 'Pragma': 'no-cache', 'Expires': '0'}
                )

        @app.post('/run')
        async def run_command(req: RunCommandRequest, request: Request):
            client_ip = request.client.host if request.client else 'unknown'
            if not api_rate_limiter.is_allowed(client_ip):
                return JSONResponse(status_code=429, content={'error': '请求过于频繁', 'retry_after': api_rate_limiter.get_retry_after(client_ip)}, headers={'Retry-After': str(api_rate_limiter.get_retry_after(client_ip))})
            command = req.command

            # 命令长度限制
            if len(command) > 10000:
                raise HTTPException(status_code=400, detail='命令长度超过限制（最大10000字符）')

            try:
                cmd_list = shlex.split(command)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=f'命令格式错误: {e}')

            if len(cmd_list) < 2:
                raise HTTPException(status_code=400, detail='命令格式无效: 至少需要可执行文件和参数')

            exe_base = os.path.basename(cmd_list[0]).lower()
            allowed_exe = {os.path.basename(sys.executable).lower()}
            allowed_exe.update({'python', 'python3', 'py'})
            if Environment.EXE_SUFFIX:
                allowed_exe.update({n + Environment.EXE_SUFFIX for n in allowed_exe if not n.endswith(Environment.EXE_SUFFIX)})
            if exe_base not in allowed_exe and not exe_base.startswith('python'):
                raise HTTPException(status_code=400, detail=f'不允许的可执行文件: {cmd_list[0]}')

            if not any(os.path.basename(arg) == 'main.py' for arg in cmd_list[1:]):
                raise HTTPException(status_code=400, detail='命令必须包含 main.py')

            if exe_base in allowed_exe:
                cmd_list[0] = VENV_PYTHON

            task_id = str(uuid.uuid4())[:8]
            with _tasks_lock:
                tasks[task_id] = {'command': ' '.join(cmd_list), 'status': 'starting', 'output': '', 'returncode': None, 'error': None}
            thread = threading.Thread(target=run_command_background, args=(task_id, cmd_list))
            thread.start()
            return JSONResponse(content={'success': True, 'task_id': task_id, 'message': f'命令已启动 (系统: {Environment.SYSTEM})'})

        @app.post('/input')
        async def send_input(req: TaskInputRequest):
            task_id, user_input = req.task_id, req.user_input
            with _processes_lock:
                if task_id not in processes:
                    raise HTTPException(status_code=404, detail='没有正在运行的进程')
                process = processes[task_id]
            try:
                process.stdin.write(user_input + '\n')
                process.stdin.flush()
                return JSONResponse(content={'success': True, 'message': '输入已发送'})
            except Exception as e:
                raise HTTPException(status_code=500, detail='Internal server error')

        @app.post('/kill')
        async def kill_task(req: KillTaskRequest):
            task_id = req.task_id
            with _processes_lock:
                if task_id not in processes:
                    return JSONResponse(content={'success': True, 'message': '进程已结束'})
                process = processes[task_id]
            try:
                try:
                    process.terminate()
                    try:
                        process.wait(timeout=TIMEOUT_CONFIG['subprocess_kill'])
                    except subprocess.TimeoutExpired:
                        process.kill()
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass
                with _tasks_lock:
                    if task_id in tasks:
                        tasks[task_id]['status'] = 'killed'
                with _processes_lock:
                    if task_id in processes:
                        del processes[task_id]
                return JSONResponse(content={'success': True, 'message': '进程已终止'})
            except Exception as e:
                logger.warning(f'Kill task completed with exception: {e}')
                return JSONResponse(content={'success': True, 'message': '操作完成'})

        @app.get('/output/{task_id}')
        async def get_output(task_id: str):
            with _tasks_lock:
                if task_id not in tasks:
                    raise HTTPException(status_code=404, detail='任务不存在')
                task = tasks[task_id]
            return JSONResponse(content={'status': task['status'], 'output': task.get('output', ''), 'returncode': task.get('returncode'), 'error': task.get('error')})

        @app.get('/api/cookie')
        async def get_cookie_status():
            cookie_file = os.path.join(PROJECT_DIR, 'config', 'cookies.json')
            if not os.path.exists(cookie_file):
                raise HTTPException(status_code=404, detail='Cookie文件不存在')
            try:
                with open(cookie_file, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)

                if not cookies or len(cookies) == 0:
                    raise HTTPException(status_code=404, detail='Cookie文件为空')

                # 查找token cookie
                token_cookie = None
                for c in cookies:
                    if c.get('name') == 'token':
                        token_cookie = c
                        break

                if not token_cookie:
                    raise HTTPException(status_code=404, detail='未找到Token Cookie')

                expires = token_cookie.get('expires')
                if not expires or not isinstance(expires, (int, float)) or expires <= 0:
                    raise HTTPException(status_code=404, detail='Token Cookie无效')

                expires_time = datetime.fromtimestamp(expires)
                hours_remaining = (expires_time - datetime.now()).total_seconds() / 3600

                return JSONResponse(content={
                    'valid': True,
                    'expires': expires_time.strftime('%Y-%m-%d'),
                    'hours_remaining': round(hours_remaining, 1),
                    'expired': hours_remaining <= 0,
                    'system': Environment.SYSTEM,
                    'cookie_name': 'Token'
                })
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=500, detail='Internal server error')

        @app.get('/api/sku/compare')
        async def compare_sku():
            try:
                json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
                if not json_files:
                    raise HTTPException(status_code=404, detail='没有找到JSON文件')
                latest_json = max(json_files, key=os.path.getmtime)

                with open(latest_json, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                json_stock_numbers = sorted([p.get('货号', '') for p in products if p.get('货号')])

                input_file = os.path.join(PROJECT_DIR, 'config', 'input_stock_numbers.txt')
                txt_stock_numbers = []
                if os.path.exists(input_file):
                    with open(input_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        txt_stock_numbers = sorted(set(re.findall(r'\d+', content)))

                json_set = set(json_stock_numbers)
                txt_set = set(txt_stock_numbers)

                result = {
                    'json_file': os.path.basename(latest_json),
                    'json_count': len(json_set),
                    'txt_count': len(txt_set),
                    'json_skus': json_stock_numbers,
                    'txt_skus': txt_stock_numbers,
                    'missing_in_json': sorted(list(txt_set - json_set)),
                    'extra_in_json': sorted(list(json_set - txt_set)),
                    'common': sorted(list(txt_set & json_set)),
                    'missing_count': len(txt_set - json_set),
                    'extra_count': len(json_set - txt_set),
                    'common_count': len(txt_set & json_set)
                }
                return JSONResponse(content=result)
            except HTTPException:
                raise
            except Exception as e:
                error_detail = 'Internal server error'
                print(f'get_daily_profit错误: {error_detail}')
                raise HTTPException(status_code=500, detail='Internal server error')

        @app.api_route('/api/sku/compare/txt', methods=['GET', 'POST'])
        async def compare_sku_txt(request: Request):
            try:
                json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
                if not json_files:
                    return jsonify({'error': '没有找到JSON文件'}, status_code=404)
                latest_json = max(json_files, key=os.path.getmtime)

                with open(latest_json, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                json_stock_numbers = sorted([p.get('货号', '') for p in products if p.get('货号')])

                txt_stock_numbers_raw = []
                if request.method == 'POST':
                    req_data = await request.json()
                    input_skus = req_data.get('skus', '')

                    if len(input_skus) > 50000:
                        return jsonify({'error': '输入数据过长（最大50000字符）'}, status_code=400)

                    txt_stock_numbers_raw = [s.strip() for s in re.split(r'[\s,\n\r\t]+', input_skus) if s.strip()]

                    if len(txt_stock_numbers_raw) > 10000:
                        return jsonify({'error': '货号数量过多（最大10000个）'}, status_code=400)
                else:
                    input_file = os.path.join(PROJECT_DIR, 'config', 'input_stock_numbers.txt')
                    if os.path.exists(input_file):
                        with open(input_file, 'r', encoding='utf-8') as f:
                            content = f.read()
                            txt_stock_numbers_raw = [s.strip() for s in content.split() if s.strip()]
                
                txt_stock_numbers = sorted(set(txt_stock_numbers_raw))
                duplicates = StockNumberComparator.find_duplicate_stock_numbers(txt_stock_numbers_raw)

                today = datetime.now().strftime('%Y%m%d')
                diff_log_file = os.path.join(PROJECT_DIR, 'file', f'diff_log_{today}.json')
                
                json_set = set(json_stock_numbers)
                txt_set = set(txt_stock_numbers)

                high_price_count = 0
                high_price_stock_numbers = []
                for p in products:
                    price = p.get('售价', '') or p.get('price','')
                    if price:
                        try:
                            price_val = float(price.replace('¥', '').replace(',', ''))
                            if price_val >= 599:
                                high_price_count += 1
                                sku = p.get('货号', '')
                                if sku:
                                    high_price_stock_numbers.append(str(sku))
                        except Exception as e:
                            handle_exception(e, '/api/sku/compare/txt解析商品价格')
                            pass

                high_price_set = set(high_price_stock_numbers)
                high_price_existing = sorted(list(high_price_set & txt_set))
                high_price_extra_in_json = sorted(list(high_price_set - txt_set))

                xiaoji_records = data.get('小计', []) if isinstance(data, dict) else []
                today_xiaoji = None
                for record in reversed(xiaoji_records):
                    if record.get('timestamp', '').startswith(today):
                        today_xiaoji = record
                        break

                added_products_all = []
                removed_products = []
                added_high_price = []

                if os.path.exists(diff_log_file):
                    with open(diff_log_file, 'r', encoding='utf-8') as f:
                        diff_data = json.load(f)
                    if diff_data.get('logs'):
                        last_log = diff_data['logs'][-1]
                        added_products_all = last_log.get('added', [])
                        removed_products = last_log.get('removed', [])

                if today_xiaoji:
                    added_products_all = today_xiaoji.get('added', [])
                    removed_products = today_xiaoji.get('removed', [])
                    for sku in added_products_all:
                        for p in products:
                            if str(p.get('货号', '')) == str(sku):
                                price = p.get('售价', '') or p.get('price','')
                                try:
                                    price_val = float(price.replace('¥', '').replace(',', '')) if price else 0
                                    if price_val >= 599:
                                        added_high_price.append(sku)
                                except Exception as e:
                                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                                    pass
                                break

                added_high_price = sorted(list(set(added_high_price)))

                result = {
                    'type': 'txt',
                    'json_file': os.path.basename(latest_json),
                    'json_count': len(json_set),
                    'txt_count': len(txt_set),
                    'missing_in_json': sorted(list(txt_set - json_set)),
                    'extra_in_json': sorted(list(json_set - txt_set)),
                    'common': sorted(list(txt_set & json_set)),
                    'missing_count': len(txt_set - json_set),
                    'extra_count': len(json_set - txt_set),
                    'common_count': len(txt_set & json_set),
                    'duplicate_count': len(duplicates),
                    'duplicates': duplicates,
                    'high_price_count': high_price_count,
                    'high_price_extra_in_json': high_price_extra_in_json,
                    'high_price_existing': high_price_existing,
                    'added_high_price': added_high_price,
                    'added_high_price_count': len(added_high_price),
                    'added_products': sorted(added_products_all)[:100],
                    'added_products_count': len(added_products_all),
                    'removed_products': removed_products[:100],
                    'removed_products_count': len(removed_products)
                }
                return jsonify(result)
            except Exception as e:
                logger.error(f'[compare_sku_txt] 对比失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'error': 'SKU对比失败'}, status_code=500)

        @app.get('/api/sku/compare/excel')
        async def compare_sku_excel():
            try:
                if pd is None:
                    return jsonify({'error': 'pandas未安装，Excel对比功能不可用'}, status_code=500)
                
                today = datetime.now().strftime('%Y%m%d')
                diff_log_file = os.path.join(PROJECT_DIR, 'file', f'diff_log_{today}.json')
                
                json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
                json_files = [f for f in json_files if '_cache' not in f]
                if not json_files:
                    return jsonify({'error': '没有找到JSON文件'}, status_code=404)
                latest_json = max(json_files, key=os.path.getmtime)
                
                with open(latest_json, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                json_stock_numbers = sorted([p.get('货号', '') for p in products if p.get('货号')])
                
                excel_files_list, daily_profit_report = get_excel_files_with_report()
                
                excel_stock_numbers = []
                
                for excel_file in excel_files_list:
                    if os.path.exists(excel_file):
                        try:
                            
                            excel_dfs = FileManager.safe_read_excel(excel_file, max_retries=3, retry_delay=1.0)
                            if excel_dfs is None:
                                print(f'无法读取Excel文件: {excel_file}')
                                continue
                            
                            df = None
                            sheet_name = None
                            sku_column = None
                            
                            for sheet, temp_df in excel_dfs.items():
                                if '货号' in temp_df.columns:
                                    df = temp_df
                                    sheet_name = sheet
                                    sku_column = '货号'
                                    break
                                elif '序列号' in temp_df.columns:
                                    df = temp_df
                                    sheet_name = sheet
                                    sku_column = '序列号'
                                    break
                                elif '闲鱼' in sheet:
                                    if len(temp_df.columns) > 4:
                                        second_row = temp_df.iloc[1].tolist()
                                        if '序列号' in second_row:
                                            col_idx = second_row.index('序列号')
                                            temp_df.columns = second_row
                                            temp_df = temp_df.drop([0, 1]).reset_index(drop=True)
                                            df = temp_df
                                            sheet_name = sheet
                                            sku_column = '序列号'
                                            break
                            
                            if df is not None and sku_column is not None:
                                file_stock_numbers = [str(int(x)) if isinstance(x, float) and x == int(x) else str(x).strip() 
                                                             for x in df[sku_column].dropna() 
                                                             if str(x).strip() and str(x).strip() != 'nan' and str(x).strip() != '序列号']
                                excel_stock_numbers.extend(file_stock_numbers)
                        except PermissionError as e:
                            if "sharing violation" in str(e).lower() or "另一个程序" in str(e) or "正在使用" in str(e):
                                return jsonify({
                                    'error': f'Excel文件被其他程序占用，请关闭后再试',
                                    'detail': f'文件: {os.path.basename(excel_file)}',
                                    'path': excel_file
                                }), 423
                            raise
                        except Exception as e:
                            print(f'读取Excel文件失败: {excel_file} - {e}')
                            continue
                
                if not excel_stock_numbers:
                    return jsonify({'error': f'Excel文件中未找到"货号"或"序列号"列'}, status_code=404)
                
                json_set = set(json_stock_numbers)
                excel_set = set(excel_stock_numbers)
                
                json_sku_counts = {}
                for sku in json_stock_numbers:
                    json_sku_counts[sku] = json_sku_counts.get(sku, 0) + 1
                duplicate_skus_json = {sku: count for sku, count in json_sku_counts.items() if count > 1}
                
                excel_sku_counts = {}
                for sku in excel_stock_numbers:
                    excel_sku_counts[sku] = excel_sku_counts.get(sku, 0) + 1
                duplicate_skus = {sku: count for sku, count in excel_sku_counts.items() if count > 1}
                
                high_price_count = 0
                high_price_stock_numbers = []
                for p in products:
                    price = p.get('售价', '') or p.get('price','')
                    if price:
                        try:
                            price_val = float(price.replace('¥', '').replace(',', ''))
                            if price_val >= 599:
                                high_price_count += 1
                                sku = p.get('货号', '')
                                if sku:
                                    high_price_stock_numbers.append(str(sku))
                        except Exception as e:
                            handle_exception(e, '/api/compare解析商品价格')
                            pass
                
                # 高价商品与已存在货号的对比
                high_price_set = set(high_price_stock_numbers)
                high_price_existing = sorted(list(high_price_set & excel_set))
                high_price_extra_in_json = sorted(list(high_price_set - excel_set))
                
                # 判断今天是否运行过爬虫：从 JSON 的"小计"中查找今天的记录
                xiaoji_records = data.get('小计', []) if isinstance(data, dict) else []
                today_xiaoji = None
                for record in reversed(xiaoji_records):
                    if record.get('timestamp', '').startswith(today):
                        today_xiaoji = record
                        break
                
                # 从差异日志读取今天之前的数据（昨天 vs 今天）
                added_products_all = []
                removed_products = []
                added_high_price = []
                
                if os.path.exists(diff_log_file):
                    with open(diff_log_file, 'r', encoding='utf-8') as f:
                        diff_data = json.load(f)
                    if diff_data.get('logs'):
                        last_log = diff_data['logs'][-1]
                        added_products_all = last_log.get('added', [])
                        removed_products = last_log.get('removed', [])
                
                # 如果今天运行过爬虫（有小计记录），使用小计的数据
                if today_xiaoji:
                    added_products_all = today_xiaoji.get('added', [])
                    removed_products = today_xiaoji.get('removed', [])
                    # 高价新增从新增商品中筛选
                    for sku in added_products_all:
                        for p in products:
                            if str(p.get('货号', '')) == str(sku):
                                price = p.get('售价', '') or p.get('price','')
                                try:
                                    price_val = float(price.replace('¥', '').replace(',', '')) if price else 0
                                    if price_val >= 599:
                                        added_high_price.append(sku)
                                except Exception as e:
                                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                                    pass
                                break
                
                added_high_price = sorted(list(set(added_high_price)))
                
                result = {
                    'type': 'excel',
                    'json_file': os.path.basename(latest_json),
                    'excel_files': [os.path.basename(f) for f in excel_files_list],
                    'excel_files_count': len(excel_files_list),
                    'input_count': len(excel_set),
                    'extra_count': len(json_set - excel_set),
                    'json_count': len(json_set),
                    'common_count': len(excel_set & json_set),
                    'duplicate_count': len(duplicate_skus),
                    'duplicate_count_json': len(duplicate_skus_json),
                    'duplicate_count_excel': len(duplicate_skus),
                    'high_price_count': high_price_count,
                    'missing_in_json': sorted(list(excel_set - json_set)),
                    'extra_in_json': sorted(list(json_set - excel_set)),
                    'common': sorted(list(excel_set & json_set)),
                    'duplicates': duplicate_skus,
                    'duplicates_json': duplicate_skus_json,
                    'duplicates_excel': duplicate_skus,
                    'missing_count': len(excel_set - json_set),
                    'high_price_extra_in_json': high_price_extra_in_json,
                    'high_price_existing': high_price_existing,
                    'added_high_price': added_high_price,
                    'added_high_price_count': len(added_high_price),
                    'added_products': sorted(added_products_all)[:100],
                    'added_products_count': len(added_products_all),
                    'removed_products': removed_products[:100],
                    'removed_products_count': len(removed_products),
                    'daily_profit_report': daily_profit_report,
                    'report_text': daily_profit_report
                }
                return jsonify(result)
            except Exception as e:
                return jsonify({'error': str(e)}, status_code=500)

        @app.get('/api/products')
        async def get_all_products():
            json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
            if not json_files:
                return jsonify({'error': '没有找到JSON文件'}, status_code=404)
            latest_file = max(json_files, key=os.path.getmtime)
            try:
                with open(latest_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data

                for p in products:
                    media_result = []
                    img_data = p.get('图片', '')
                    if img_data:
                        try:
                            if isinstance(img_data, list):
                                for b64_str in img_data:
                                    try:
                                        media_result.append(base64.b64decode(b64_str).decode('utf-8'))
                                    except Exception as e:
                                        logger.debug(f"Exception processing media: {e}")
                                        media_result.append(b64_str)
                            else:
                                try:
                                    media_result = base64.b64decode(img_data).decode('utf-8')
                                except Exception as e:
                                    logger.debug(f"Exception processing media: {e}")
                                    media_result = img_data
                        except Exception as e:
                            logger.debug(f"Exception processing media: {e}")
                            media_result = img_data
                    p['图片'] = media_result if media_result else img_data
                
                high_price_stats = data.get('高价商品统计', {})
                high_price_products = high_price_stats.get('商品列表', [])
                high_price_count = high_price_stats.get('数量', 0)
                
                total_price = 0
                total_fee = 0
                valid_price_count = 0
                
                for p in products:
                    try:
                        price_str = p.get('售价', '') or p.get('price', '')
                        if not price_str or not str(price_str).strip():
                            continue
                        price_clean = str(price_str).replace('¥', '').replace(',', '').strip()
                        price = float(price_clean)
                        if price > 0:
                            total_price += price
                            fee = price * 0.016
                            total_fee += fee
                            valid_price_count += 1
                    except Exception:
                        pass
                
                avg_price = total_price / valid_price_count if valid_price_count > 0 else 0
                
                storage_duration = None
                storage_times = []
                for p in products:
                    old_time = p.get('入库时间', '')
                    if old_time:
                        storage_times.append(old_time)
                
                if storage_times:
                    min_time_str = min(storage_times, key=lambda x: (
                        0 if '刚刚' in x or '分钟前' in x else
                        1 if '小时前' in x else
                        2 if '天前' in x else
                        3 if '周前' in x else
                        4 if '月前' in x else
                        5
                    ))
                    storage_duration = min_time_str
                
                created_time = None
                created_times = []
                for p in products:
                    created_time_str = p.get('入库时间戳', '')
                    if created_time_str:
                        created_times.append(created_time_str)
                
                if created_times:
                    created_time = min(created_times)
                
                return jsonify({
                    'filename': os.path.basename(latest_file), 
                    'total': len(products), 
                    'products': products[:100], 
                    'highPriceProducts': high_price_products[:500], 
                    'highPriceCount': high_price_count,
                    'totalPrice': f'¥{total_price:,.2f}',
                    'avgPrice': f'¥{avg_price:,.2f}',
                    'fee': f'¥{total_fee:,.2f}',
                    'system': Environment.SYSTEM,
                    'storage_duration': storage_duration,
                    'created_time': created_time
                })
            except Exception as e:
                return jsonify({'error': str(e)}, status_code=500)

        @app.get('/api/daily-profit')
        async def get_daily_profit(group_by: str = 'day', start_date: str = None, end_date: str = None):
            try:
                if pd is None or openpyxl is None:
                    return jsonify({'error': 'pandas或openpyxl未安装，每日利润报表功能不可用'}, status_code=500)
                
                excel_files_list, daily_profit_report = get_excel_files_with_report()
                
                table_data = []
                all_records = []
                
                for excel_file in excel_files_list:
                    if os.path.exists(excel_file):
                        try:
                            
                            wb = openpyxl.load_workbook(excel_file, data_only=True)
                            sheet_name = '每日利润'
                            if sheet_name in wb.sheetnames:
                                ws = wb[sheet_name]
                            else:
                                ws = wb.active
                            
                            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False)):
                                row_data = []
                                for cell in row:
                                    row_data.append(cell.value)
                                table_data.append(row_data)
                                
                                if row_idx > 0:
                                    try:
                                        amount = float(str(row_data[1] or 0).replace('¥', '').replace(',', '').strip()) if row_data[1] else 0
                                        cost = float(str(row_data[2] or 0).replace('¥', '').replace(',', '').strip()) if row_data[2] else 0
                                        profit = float(str(row_data[3] or 0).replace('¥', '').replace(',', '').strip()) if row_data[3] else 0
                                        date_val = row_data[4]
                                        remark = row_data[5] if len(row_data) > 5 else ''
                                        
                                        if isinstance(date_val, datetime):
                                            record_date_str = date_val.strftime('%Y-%m-%d')
                                        elif isinstance(date_val, str):
                                            date_str = date_val.strip()
                                            record_date_str = None
                                            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y年%m月%d日', '%Y.%m.%d']:
                                                try:
                                                    record_date = datetime.strptime(date_str.split()[0], fmt)
                                                    record_date_str = record_date.strftime('%Y-%m-%d')
                                                    break
                                                except Exception as e:
                                                    logger.debug(f"Exception in loop: {e}")
                                                    continue
                                            if record_date_str is None:
                                                parts = date_str.split()
                                                if len(parts) >= 2:
                                                    try:
                                                        day_part = parts[1].rstrip('日号号日.')
                                                        record_date = datetime.strptime(day_part, '%d')
                                                        record_date = record_date.replace(year=datetime.now().year)
                                                        if 'Dec' in date_str or 'De' in date_str:
                                                            record_date = record_date.replace(month=12)
                                                        elif 'Jan' in date_str:
                                                            record_date = record_date.replace(month=1)
                                                        elif 'Feb' in date_str:
                                                            record_date = record_date.replace(month=2)
                                                        elif 'Mar' in date_str:
                                                            record_date = record_date.replace(month=3)
                                                        elif 'Apr' in date_str:
                                                            record_date = record_date.replace(month=4)
                                                        elif 'May' in date_str:
                                                            record_date = record_date.replace(month=5)
                                                        elif 'Jun' in date_str:
                                                            record_date = record_date.replace(month=6)
                                                        elif 'Jul' in date_str:
                                                            record_date = record_date.replace(month=7)
                                                        elif 'Aug' in date_str:
                                                            record_date = record_date.replace(month=8)
                                                        elif 'Sep' in date_str:
                                                            record_date = record_date.replace(month=9)
                                                        elif 'Oct' in date_str:
                                                            record_date = record_date.replace(month=10)
                                                        elif 'Nov' in date_str:
                                                            record_date = record_date.replace(month=11)
                                                        record_date_str = record_date.strftime('%Y-%m-%d')
                                                    except Exception as e:
                                                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                                                        pass
                                                if record_date_str is None:
                                                    continue
                                        elif isinstance(date_val, (int, float)):
                                            try:
                                                record_date = datetime(1899, 12, 30) + timedelta(days=int(date_val))
                                                if record_date.year < 2000:
                                                    continue
                                                record_date_str = record_date.strftime('%Y-%m-%d')
                                            except Exception as e:
                                                logger.debug(f"Exception in loop: {e}")
                                                continue
                                        else:
                                            continue
                                        
                                        all_records.append({
                                            '项目': row_data[0],
                                            '金额': amount,
                                            '成本': cost,
                                            '纯利': profit,
                                            '日期': record_date_str,
                                            '备注': remark
                                        })
                                    except (ValueError, TypeError, IndexError):
                                        pass
                            
                            wb.close()
                            break
                        except Exception as e:
                            print(f'读取Excel文件失败: {excel_file} - {e}')
                            continue
                
                if not table_data:
                    return jsonify({'error': '未找到Excel数据'}, status_code=404)
                
                if start_date:
                    try:
                        all_records = [r for r in all_records if r['日期'] >= start_date]
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
                
                if end_date:
                    try:
                        all_records = [r for r in all_records if r['日期'] <= end_date]
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
                
                summary = {}
                for record in all_records:
                    date_key = record['日期']
                    project_name = record.get('项目', '未分类')
                    
                    if group_by == 'month':
                        key = date_key[:7]
                    elif group_by == 'year':
                        key = date_key[:4]
                    elif group_by == 'all':
                        key = '总计'
                    else:
                        key = date_key
                    
                    composite_key = key + '|' + project_name
                    
                    if composite_key not in summary:
                        summary[composite_key] = {'日期': key, '项目': project_name, '金额': 0, '成本': 0, '纯利': 0, '数量': 0}
                    
                    summary[composite_key]['金额'] += record['金额']
                    summary[composite_key]['成本'] += record['成本']
                    summary[composite_key]['纯利'] += record['纯利']
                    summary[composite_key]['数量'] += 1
                
                summary_list = sorted(summary.values(), key=lambda x: x['日期'])
                
                for row_data in table_data:
                    for col_idx, cell_val in enumerate(row_data):
                        if isinstance(cell_val, datetime):
                            row_data[col_idx] = cell_val.strftime('%Y-%m-%d')
                        elif isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool):
                            if cell_val > 40000 and cell_val < 100000:
                                try:
                                    converted = datetime(1899, 12, 30) + timedelta(days=int(cell_val))
                                    if converted.year >= 2000:
                                        row_data[col_idx] = converted.strftime('%Y-%m-%d')
                                except Exception as e:
                                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                                    pass
                
                result = {
                    'daily_profit_report': daily_profit_report,
                    'table_data': table_data,
                    'excel_file': os.path.basename(excel_files_list[0]) if excel_files_list else None,
                    'summary': summary_list,
                    'group_by': group_by,
                    'total_records': len(all_records),
                    'all_records': all_records
                }
                return jsonify(result)
            except Exception as e:
                error_detail = str(e) + '\n' + traceback.format_exc()
                print(f'get_daily_profit错误: {error_detail}')
                return jsonify({'error': str(e), 'detail': error_detail}, status_code=500)

        def get_all_products():
            json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
            if not json_files:
                return jsonify({'error': '没有找到JSON文件'}, status_code=404)
            latest_file = max(json_files, key=os.path.getmtime)
            try:
                with open(latest_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data

                for p in products:
                    media_result = []
                    img_data = p.get('图片', '')
                    if img_data:
                        try:
                            if isinstance(img_data, list):
                                for b64_str in img_data:
                                    try:
                                        media_result.append(base64.b64decode(b64_str).decode('utf-8'))
                                    except Exception as e:
                                        logger.debug(f"Exception processing media: {e}")
                                        media_result.append(b64_str)
                            else:
                                try:
                                    media_result = base64.b64decode(img_data).decode('utf-8')
                                except Exception as e:
                                    logger.debug(f"Exception processing media: {e}")
                                    media_result = img_data
                        except Exception as e:
                            logger.debug(f"Exception processing media: {e}")
                            media_result = img_data
                    p['图片'] = media_result if media_result else img_data
                
                high_price_products = []
                total_price = 0
                total_fee = 0
                valid_price_count = 0
                
                def safe_print(*args, **kwargs):
                    try:
                        print(*args, **kwargs)
                    except (IOError, OSError):
                        pass
                
                safe_print(f'开始处理 {len(products)} 个商品...')
                
                for p in products:
                    try:
                        price_str = p.get('售价', '') or p.get('price','')
                        if not price_str or not price_str.strip():
                            continue
                        
                        price_clean = price_str.replace('¥', '').replace(',', '').strip()
                        price = float(price_clean)
                        
                        if price >= 599:
                            high_price_products.append(p)
                        
                        # 计算总价和手续费
                        if price > 0:
                            total_price += price
                            # 计算闲鱼平台手续费（售价 * 1.6%）
                            fee = price * 0.016
                            total_fee += fee
                            valid_price_count += 1
                    except Exception as e:
                        safe_print(f'处理商品时出错: {e}, price_str: {p.get("售价", "")}')
                        pass
                
                safe_print(f'统计结果: valid_price_count={valid_price_count}, total_price={total_price}, high_price_count={len(high_price_products)}')
                
                # 计算平均价格
                avg_price = total_price / valid_price_count if valid_price_count > 0 else 0
                
                return jsonify({
                    'filename': os.path.basename(latest_file), 
                    'total': len(products), 
                    'products': products[:100], 
                    'highPriceProducts': high_price_products[:500], 
                    'highPriceCount': high_price_count,
                    'totalPrice': f'¥{total_price:,.2f}',
                    'avgPrice': f'¥{avg_price:,.2f}',
                    'fee': f'¥{total_fee:,.2f}',
                    'system': Environment.SYSTEM
                })
            except Exception as e:
                return jsonify({'error': str(e)}, status_code=500)

        @app.get('/api/product')
        async def get_product(sku: str = ''):
            sku = sku.strip()
            if not sku:
                return jsonify({'error': '请提供货号'}, status_code=400)
            json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
            if not json_files:
                return jsonify({'found': False, 'error': '没有找到JSON文件'})
            latest_file = max(json_files, key=os.path.getmtime)
            try:
                with open(latest_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                for p in products:
                    if str(p.get('货号')) == str(sku):
                        # 解码Base64图片URL
                        images = p.get('图片', [])
                        if images:
                            if isinstance(images, list):
                                decoded_images = []
                                for img in images:
                                    try:
                                        decoded = base64.b64decode(img).decode('utf-8')
                                        decoded_images.append(decoded)
                                    except Exception as e:
                                        logger.debug(f"Exception decoding image: {e}")
                                        decoded_images.append(img)
                                p['图片'] = decoded_images
                            elif isinstance(images, str):
                                try:
                                    decoded = base64.b64decode(images).decode('utf-8')
                                    p['图片'] = [decoded]
                                except Exception as e:
                                    logger.debug(f"Exception decoding image to list: {e}")
                                    p['图片'] = [images]
                            else:
                                p['图片'] = []
                        else:
                            p['图片'] = []
                        return jsonify({'found': True, 'product': p})
                return jsonify({'found': False, 'error': '未找到该商品'})
            except Exception as e:
                return jsonify({'found': False, 'error': str(e)})
        
        @app.get('/api/product/search')
        async def search_product(sku: str = ''):
            sku = sku.strip()
            if not sku:
                return jsonify({'error': '请提供货号'}, status_code=400)

            if len(sku) > 100:
                return jsonify({'error': '货号过长（最大100字符）'}, status_code=400)

            if not re.match(r'^[a-zA-Z0-9\u4e00-\u9fa5\-_]+$', sku):
                logger.warning(f'[search_product] 检测到可疑货号: {sku[:50]}')
                return jsonify({'error': '货号包含非法字符（仅允许字母、数字、中文、横线和下划线）'}, status_code=400)

            json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
            if not json_files:
                return jsonify({'error': '没有找到JSON文件'}, status_code=404)
            latest_file = max(json_files, key=os.path.getmtime)
            try:
                with open(latest_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                for p in products:
                    if p.get('货号') == sku:
                        media_result = []
                        new_image_url = p.get('图片', '')
                        if new_image_url:
                            try:
                                # 检查是否为有效的JSON字符串
                                if isinstance(new_image_url, str):
                                    # 防止HTML或非法数据导致解析失败
                                    if new_image_url.strip().startswith('<') or not new_image_url.strip().startswith('['):
                                        img_data = new_image_url
                                    else:
                                        try:
                                            img_data = json.loads(new_image_url)
                                        except (json.JSONDecodeError, TypeError):
                                            img_data = new_image_url
                                else:
                                    img_data = new_image_url
                            except Exception as e:
                                print(f'  ⚠️ 图片URL解析异常: {e}')
                                img_data = new_image_url
                            if isinstance(img_data, list):
                                for b64_str in img_data:
                                    try:
                                        decoded_url = base64.b64decode(b64_str).decode('utf-8')
                                        if decoded_url.startswith('http'):
                                            media_result.append(decoded_url)
                                        else:
                                            media_result.append(b64_str)
                                    except Exception as e:
                                        logger.debug(f"Exception processing media: {e}")
                                        media_result.append(b64_str)
                            else:
                                try:
                                    decoded_url = base64.b64decode(img_data).decode('utf-8')
                                    if decoded_url.startswith('http'):
                                        media_result = [decoded_url]
                                    else:
                                        media_result = [img_data]
                                except Exception as e:
                                    logger.debug(f"Exception processing media: {e}")
                                    media_result = [img_data]
                        p['图片'] = media_result
                        return jsonify({'found': True, 'product': p, 'filename': os.path.basename(latest_file), 'saved': True})
                return jsonify({'found': False, 'error': f'未找到货号为 {sku} 的商品'})
            except Exception as e:
                logger.error(f'[search_product] 搜索失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'error': '搜索商品失败'}, status_code=500)

        @app.get('/api/product/by-description')
        async def get_product_by_description(description: str = ''):
            description = description.strip()
            if not description:
                return JSONResponse(content={'error': '请提供商品描述'}, status_code=400)

            if len(description) > 500:
                return JSONResponse(content={'error': '商品描述过长（最大500字符）'}, status_code=400)

            if re.search(r'[<>"\'&]', description):
                logger.warning(f'[get_product_by_description] 检测到可疑字符 in description: {description[:50]}...')
                return JSONResponse(content={'error': '商品描述包含非法字符'}, status_code=400)

            json_files = glob.glob(os.path.join(PROJECT_DIR, 'file', '*微购相册*.json'))
            if not json_files:
                return jsonify({'found': False, 'error': '没有找到JSON文件'})
            latest_file = max(json_files, key=os.path.getmtime)
            try:
                with open(latest_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                products = data.get('商品列表', []) if isinstance(data, dict) else data
                for p in products:
                    stored_desc = p.get('商品描述', '')
                    if stored_desc.replace(' ', '') == description.replace(' ', ''):
                        images = p.get('图片', [])
                        if images:
                            if isinstance(images, list):
                                decoded_images = []
                                for img in images:
                                    try:
                                        decoded = base64.b64decode(img).decode('utf-8')
                                        decoded_images.append(decoded)
                                    except Exception as e:
                                        logger.debug(f"Exception decoding image: {e}")
                                        decoded_images.append(img)
                                p['图片'] = decoded_images
                            elif isinstance(images, str):
                                try:
                                    decoded = base64.b64decode(images).decode('utf-8')
                                    p['图片'] = [decoded]
                                except Exception as e:
                                    logger.debug(f"Exception decoding image to list: {e}")
                                    p['图片'] = [images]
                            else:
                                p['图片'] = []
                        else:
                            p['图片'] = []
                        return jsonify({'found': True, 'product': p})
                return jsonify({'found': False, 'error': '未找到该商品'})
            except Exception as e:
                logger.error(f'[get_product_by_description] 查询失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'found': False, 'error': '查询商品信息失败'})

        @app.post('/api/clean/list')
        async def api_clean_list(request: Request):
            try:
                data = await request.json()
                directory = data.get('directory', '')

                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                
                if not os.path.exists(directory):
                    return jsonify({'success': False, 'error': f'目录不存在: {directory}'})
                
                log_file = os.path.join(directory, 'clean_files.log')
                
                log_stream = io.StringIO()
                list_files(directory=directory, log_file=log_file, stream=log_stream)
                
                return jsonify({'success': True, 'output': log_stream.getvalue()})
            except Exception as e:
                logger.error(f'[api_clean_list] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '文件列表获取失败，请查看服务器日志'})

        @app.post('/api/clean/group')
        async def api_clean_group(request: Request):
            try:
                data = await request.json()
                directory = data.get('directory', '')
                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                dry_run = bool(data.get('dry_run', False))
                log_file = os.path.join(directory, 'clean_files.log')

                log_stream = io.StringIO()
                clean_old_files(directory=directory, dry_run=dry_run, log_file=log_file, stream=log_stream)

                return jsonify({'success': True, 'output': log_stream.getvalue()})
            except Exception as e:
                logger.error(f'[api_clean_group] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '文件分组清理失败，请查看服务器日志'})

        @app.post('/api/clean/time')
        async def api_clean_time():
            try:
                data = await request.json()
                directory = data.get('directory', '')
                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                minutes = data.get('minutes', 5)
                if not isinstance(minutes, (int, float)) or minutes < 0 or minutes > 525600:
                    return jsonify({'success': False, 'error': 'minutes参数无效（必须是0-525600之间的数字）'})
                dry_run = bool(data.get('dry_run', False))
                log_file = os.path.join(directory, 'clean_files.log')

                log_stream = io.StringIO()
                clean_old_files_by_time(directory=directory, minutes=minutes, dry_run=dry_run, log_file=log_file, stream=log_stream)
                
                return jsonify({'success': True, 'output': log_stream.getvalue()})
            except Exception as e:
                logger.error(f'[api_clean_time] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '按时间清理失败，请查看服务器日志'})

        @app.post('/api/clean/all')
        async def api_clean_all(request: Request):
            try:
                data = await request.json()
                directory = data.get('directory', '')
                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                dry_run = bool(data.get('dry_run', False))
                log_file = os.path.join(directory, 'clean_files.log')

                log_stream = io.StringIO()
                clean_all_files(directory=directory, dry_run=dry_run, log_file=log_file, stream=log_stream)

                output = log_stream.getvalue()
                response_data = json.dumps({'success': True, 'output': output}, ensure_ascii=False)
                return Response(response_data, media_type='application/json')
            except Exception as e:
                logger.error(f'[api_clean_all] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '全量清理失败，请查看服务器日志'})

        @app.post('/api/clean/png')
        async def api_clean_png(request: Request):
            try:
                data = await request.json()
                directory = data.get('directory', '')
                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                dry_run = bool(data.get('dry_run', False))
                log_file = os.path.join(directory, 'clean_files.log')

                log_stream = io.StringIO()
                clean_png_files(directory=directory, dry_run=dry_run, log_file=log_file, stream=log_stream)
                
                output = log_stream.getvalue()
                response_data = json.dumps({'success': True, 'output': output}, ensure_ascii=False)
                return Response(response_data, media_type='application/json')
            except Exception as e:
                logger.error(f'[api_clean_png] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': 'PNG清理失败，请查看服务器日志'})

        @app.post('/api/clean/media')
        async def api_clean_media(request: Request):
            try:
                data = await request.json()
                directory = data.get('directory', '')
                if not directory or directory.strip() == '':
                    directory = PROJECT_DIR
                elif not os.path.isabs(directory):
                    safe_dir, err = sec_sp(PROJECT_DIR, directory)
                    if not safe_dir:
                        return jsonify({'success': False, 'error': f'路径遍历攻击被阻止: {err}'})
                    directory = safe_dir
                dry_run = bool(data.get('dry_run', False))
                log_file = os.path.join(directory, 'clean_files.log')

                log_stream = io.StringIO()
                clean_media_files(directory=directory, dry_run=dry_run, log_file=log_file, stream=log_stream)
                
                output = log_stream.getvalue()
                response_data = json.dumps({'success': True, 'output': output}, ensure_ascii=False)
                return Response(response_data, media_type='application/json')
            except Exception as e:
                logger.error(f'[api_clean_media] 操作失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '媒体文件清理失败，请查看服务器日志'})

        @app.get('/api/version')
        async def get_version():
            return JSONResponse(content={'version': get_version_from_readme()})

        @app.get('/api/changelog')
        async def get_changelog():
            try:
                readme_path = os.path.join(PROJECT_DIR, 'README.md')
                with open(readme_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                lines = content.split('\n')
                changelog = []
                current_version = None
                current_date = None
                current_items = []
                current_item = None
                current_section = None
                in_changelog = False
                in_code_block = False
                for line in lines:
                    if '最新更新' in line.strip() and line.strip().startswith('##'):
                        in_changelog = True
                        continue
                    if not in_changelog:
                        continue
                    stripped = line.strip()
                    version_match = re.match(r'###\s+v([\d.]+)\s+\(([^)]+)\)', stripped)
                    if not version_match:
                        version_match = re.match(r'###\s+v([\d.]+)\s+\(([^)]+)\)', line.split(' - ')[0].strip())
                    if not version_match:
                        version_match = re.match(r'##\s+v([\d.]+)\s+\(([^)]+)\)', stripped)
                    if not version_match:
                        version_match = re.match(r'##\s+v([\d.]+)\s+\(([^)]+)\)', line.split(' - ')[0].strip())
                    if not version_match:
                        m_no_date = re.match(r'(?:###|##)\s+v([\d.]+)\s+', stripped)
                        if m_no_date:
                            version_match = type('VM', (), {'group': lambda self, n: m_no_date.group(n) if n <= m_no_date.lastindex else ''})()
                    if version_match:
                        if current_version:
                            if current_section:
                                current_items.append(current_section)
                                current_section = None
                            elif current_item:
                                current_items.append(current_item)
                                current_item = None
                            changelog.append({
                                'version': current_version,
                                'date': current_date,
                                'items': current_items
                            })
                        current_version = version_match.group(1)
                        current_date = version_match.group(2)
                        current_items = []
                        current_item = None
                        current_section = None
                        in_code_block = False
                        continue
                    if stripped.startswith('## ') and in_changelog and current_version:
                        break
                    if stripped.startswith('```'):
                        in_code_block = not in_code_block
                        if current_section:
                            current_section['content'] += line + '\n'
                        continue
                    section_match = re.match(r'^####\s+(.+)$', stripped)
                    if section_match and current_version:
                        if current_section:
                            current_items.append(current_section)
                        elif current_item:
                            current_items.append(current_item)
                            current_item = None
                        current_section = {
                            'type': 'section',
                            'title': section_match.group(1).strip(),
                            'content': '',
                            'sub_items': []
                        }
                        continue
                    item_match = re.match(r'^-\s+\*\*(.+?)\*\*\s*[-–]?\s*(.*)', stripped)
                    if item_match and current_version:
                        if current_section:
                            current_items.append(current_section)
                            current_section = None
                        elif current_item:
                            current_items.append(current_item)
                        current_item = {
                            'type': 'item',
                            'title': item_match.group(1),
                            'desc': item_match.group(2).strip(),
                            'sub_items': []
                        }
                        continue
                    sub_match = re.match(r'^-\s+(.*)', stripped)
                    if sub_match and (current_item or current_section):
                        is_indented = line.startswith('  ') or line.startswith('\t')
                        if current_item and is_indented:
                            current_item['sub_items'].append(sub_match.group(1).strip())
                        elif current_section:
                            current_section['sub_items'].append(sub_match.group(1).strip())
                        continue
                    if current_section:
                        if in_code_block:
                            current_section['content'] += line + '\n'
                        elif stripped:
                            current_section['content'] += line + '\n'
                if current_section:
                    current_items.append(current_section)
                elif current_item:
                    current_items.append(current_item)
                if current_version:
                    changelog.append({
                        'version': current_version,
                        'date': current_date,
                        'items': current_items
                    })
                result = {'success': True, 'changelog': changelog}
                _debug_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print(f'[{_debug_time}] [DEBUG] changelog API 返回: {len(changelog)} 个版本', file=sys.stderr)
                if changelog:
                    print(f'[{_debug_time}] [DEBUG] 最新版本: {changelog[0]["version"]}, 包含 {len(changelog[0]["items"])} 个项目', file=sys.stderr)
                    for idx, item in enumerate(changelog[0]['items']):
                        print(f'[{_debug_time}] [DEBUG]   项目{idx}: type={item.get("type")}, title={str(item.get("title", ""))[:50]}', file=sys.stderr)
                return jsonify(result)
            except Exception as e:
                _error_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print(f'[{_error_time}] [ERROR] changelog 解析失败: {e}', file=sys.stderr)
                traceback.print_exc(file=sys.stderr)
                logger.error(f'[api_changelog] 解析失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '更新日志解析失败'}, status_code=500)

        @app.get('/api/changelog-debug')
        def get_changelog_debug():
            try:
                readme_path = os.path.join(PROJECT_DIR, 'README.md')
                with open(readme_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                lines = content.split('\n')
                debug_lines = []
                in_changelog = False
                for i, line in enumerate(lines, 1):
                    if '最新更新' in line.strip() and line.strip().startswith('##'):
                        in_changelog = True
                    if in_changelog:
                        debug_lines.append(f'{i:4d}: {line}')
                        if line.strip().startswith('## ') and i > 10:
                            break
                        if len(debug_lines) > 200:
                            break
                return jsonify({
                    'success': True,
                    'total_lines': len(lines),
                    'changelog_preview': '\n'.join(debug_lines[:100])
                })
            except Exception as e:
                logger.error(f'[api_changelog_debug] 获取失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '调试信息获取失败'}, status_code=500)

        @app.get('/api/readme-sections')
        def get_readme_sections():
            try:
                readme_path = os.path.join(PROJECT_DIR, 'README.md')
                with open(readme_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                lines = content.split('\n')
                sections = {}
                current_h2 = None
                current_h3 = None
                current_lines = []
                for line in lines:
                    h2_match = re.match(r'^##\s+(.+)', line)
                    h3_match = re.match(r'^###\s+(.+)', line)
                    if h2_match:
                        if current_h2:
                            key = current_h2
                            sections[key] = {
                                'title': current_h2,
                                'content': '\n'.join(current_lines).strip(),
                                'subsections': {}
                            }
                        current_h2 = h2_match.group(1).strip()
                        current_h3 = None
                        current_lines = []
                        continue
                    if h3_match:
                        if current_h3 and current_h2:
                            sub_key = current_h3
                            if current_h2 not in sections:
                                sections[current_h2] = {
                                    'title': current_h2,
                                    'content': '',
                                    'subsections': {}
                                }
                            sections[current_h2]['subsections'][sub_key] = '\n'.join(current_lines).strip()
                        current_h3 = h3_match.group(1).strip()
                        current_lines = []
                        continue
                    current_lines.append(line)
                if current_h2:
                    if current_h2 not in sections:
                        sections[current_h2] = {
                            'title': current_h2,
                            'content': '\n'.join(current_lines).strip(),
                            'subsections': {}
                        }
                    elif current_h3:
                        sections[current_h2]['subsections'][current_h3] = '\n'.join(current_lines).strip()
                features = []
                feat_section = sections.get('📋 项目简介', sections.get('功能特性', {}))
                if feat_section:
                    for sub_title, sub_content in feat_section.get('subsections', {}).items():
                        items = []
                        for l in sub_content.split('\n'):
                            m = re.match(r'-\s+\*\*(.+?)\*\*[:：]?\s*(.*)', l.strip())
                            if m:
                                items.append({'title': m.group(1), 'desc': m.group(2).strip()})
                            elif l.strip().startswith('- '):
                                items.append({'title': l.strip()[2:], 'desc': ''})
                        clean_title = re.sub(r'^\d+\.\s*', '', sub_title)
                        features.append({'title': clean_title, 'items': items})
                tech_features = []
                tech_section = sections.get('🔧 核心模块说明', sections.get('技术特点', {}))
                if tech_section:
                    for sub_title, sub_content in tech_section.get('subsections', {}).items():
                        for l in sub_content.split('\n'):
                            m = re.match(r'-\s+\*\*(.+?)\*\*[:：]?\s*(.*)', l.strip())
                            if m:
                                tech_features.append({'title': m.group(1), 'desc': m.group(2).strip()})
                    for l in tech_section.get('content', '').split('\n'):
                        m = re.match(r'-\s+\*\*(.+?)\*\*[:：]?\s*(.*)', l.strip())
                        if m:
                            tech_features.append({'title': m.group(1), 'desc': m.group(2).strip()})
                usage_steps = []
                usage_section = sections.get('🚀 快速启动', sections.get('使用方法', {}))
                if usage_section:
                    for sub_title, sub_content in usage_section.get('subsections', {}).items():
                        clean_title = re.sub(r'^方法\d+[：:]\s*', '', sub_title)
                        usage_steps.append({'title': clean_title, 'content': sub_content})
                install_section = sections.get('⚙️ 配置说明', sections.get('安装和配置', {}))
                install_steps = []
                if install_section:
                    for sub_title, sub_content in install_section.get('subsections', {}).items():
                        clean_title = re.sub(r'^\d+\.\s*', '', sub_title)
                        install_steps.append({'title': clean_title, 'content': sub_content})
                return jsonify({
                    'success': True,
                    'features': features,
                    'tech_features': tech_features,
                    'usage_steps': usage_steps,
                    'install_steps': install_steps
                })
            except Exception as e:
                logger.error(f'[api_readme_sections] 获取失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': 'README章节解析失败'}, status_code=500)

        @app.get('/api/email/config')
        def get_email_config():
            notifier = EmailNotifier()
            config = notifier.get_email_config()
            if config['smtp_password']:
                config['smtp_password'] = '******'
            return jsonify({'success': True, 'config': config})

        @app.post('/api/email/config')
        async def save_email_config(request: Request):
            try:
                data = await request.json()

                smtp_host = data.get('smtp_host', 'smtp.qq.com')
                if not isinstance(smtp_host, str) or len(smtp_host) > 255 or not smtp_host.strip():
                    return jsonify({'success': False, 'error': 'SMTP主机地址无效'})

                try:
                    smtp_port = int(data.get('smtp_port', 587))
                except (ValueError, TypeError):
                    return jsonify({'success': False, 'error': 'SMTP端口号无效'})
                if smtp_port < 1 or smtp_port > 65535:
                    return jsonify({'success': False, 'error': 'SMTP端口超出有效范围(1-65535)'})

                smtp_user = data.get('smtp_user', '')
                if not isinstance(smtp_user, str) or len(smtp_user) > 128:
                    return jsonify({'success': False, 'error': 'SMTP用户名无效'})

                smtp_password = data.get('smtp_password', '')
                if not isinstance(smtp_password, str) or len(smtp_password) > 256:
                    return jsonify({'success': False, 'error': 'SMTP密码无效'})

                from_name = data.get('from_name', '公网IP监控')
                if not isinstance(from_name, str) or len(from_name) > 64:
                    return jsonify({'success': False, 'error': '发件人名称过长'})

                to_email = data.get('to_email', '980187223@qq.com')
                if not isinstance(to_email, str) or len(to_email) > 128:
                    return jsonify({'success': False, 'error': '收件人邮箱无效'})
                email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_regex, to_email.strip()):
                    return jsonify({'success': False, 'error': '收件人邮箱格式不正确'})

                notifier = EmailNotifier()
                notifier.save_email_config(
                    smtp_host=smtp_host.strip(),
                    smtp_port=smtp_port,
                    smtp_user=smtp_user,
                    smtp_password=smtp_password,
                    from_name=from_name[:64],
                    to_email=to_email.strip()
                )
                return jsonify({'success': True, 'message': '邮件配置已保存'})
            except Exception as e:
                logger.error(f'[save_email_config] 保存失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '邮件配置保存失败'})

        @app.post('/api/email/test')
        async def test_email(request: Request):
            try:
                data = await request.json()
                test_notifier = EmailNotifier()
                test_notifier.save_email_config(
                    smtp_host=data.get('smtp_host', 'smtp.qq.com'),
                    smtp_port=int(data.get('smtp_port', 587)),
                    smtp_user=data.get('smtp_user', ''),
                    smtp_password=data.get('smtp_password', ''),
                    from_name=data.get('from_name', '公网IP监控'),
                    to_email=data.get('to_email', '980187223@qq.com')
                )
                success = test_notifier.send_tunnel_notification('https://test.example.com', 'test')
                if success:
                    return jsonify({'success': True, 'message': '测试邮件发送成功'})
                else:
                    return jsonify({'success': False, 'error': '请先启用邮件通知'})
            except Exception as e:
                logger.error(f'[test_email] 发送失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({'success': False, 'error': '测试邮件发送失败'})

        @app.get('/api/server/info')
        def get_server_info():
            port = args.port

            # 获取局域网 IP（复用PathManager的方法）
            lan_ip = PathManager.get_lan_ip() or None
            
            return jsonify({
                'success': True,
                'local_url': f'http://localhost:{port}',
                'lan_url': f'http://{lan_ip}:{port}' if lan_ip else None,
                'lan_ip': lan_ip,
                'port': port,
                'version': get_version_from_readme()
            })

        tunnel_process = None
        tunnel_url = None
        tunnel_auto_restart = True
        tunnel_restart_thread = None
        tunnel_last_error = None
        tunnel_restart_count = 0
        tunnel_restart_delay = 0
        tunnel_heartbeat_thread = None
        tunnel_last_heartbeat = 0
        tunnel_heartbeat_failed = False
        tunnel_need_restart = False
        tunnel_daemon_started = False
        tunnel_type = 'hostc'
        user_selected_tunnel_type = 'hostc'
        tunnel_current_mode = None
        email_notifier = EmailNotifier()
        old_tunnel_url = None
        tunnel_consecutive_failures = 0
        tunnel_max_consecutive_failures = 5
        tunnel_backoff_delay = 5
        last_email_sent_time = 0
        email_cooldown = 60
        email_fail_count = 0
        email_max_fail_count = 3
        email_fail_cooldown = 300
        last_email_sent_url = None
        pending_email_url = None
        email_send_lock = threading.Lock()
        global_email_cooldown = 300
        global_last_email_sent_time = 0
        recent_sent_urls = {}
        url_dedup_window = 600
        
        stable_url = None
        stable_url_confirm_count = 0
        stable_url_min_confirms = 1
        url_first_seen_time = 0
        last_stable_notification_time = 0

        cf_process = None
        cf_url = None
        cf_mode = None
        cf_stable_url = None
        cf_stable_confirm_count = 0
        cf_stable_min_confirms = 1
        cf_url_first_seen_time = 0
        cf_last_stable_notification_time = 0
        cf_heartbeat_thread = None
        cf_last_email_sent_url = None
        cf_last_email_sent_time = 0
        
        def read_tunnel_urls_file():
            """读取 tunnel_url.txt 中已有的隧道 URL
            
            Returns:
                dict: {'hostc': url or None, 'cloudflare': url or None}
            """
            result = {'hostc': None, 'cloudflare': None}
            try:
                tunnel_file = PathManager.get_tunnel_url_file()
                if os.path.exists(tunnel_file):
                    with open(tunnel_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        hostc_match = re.search(r'hostc:\s*(https?://[^\s]+)', content)
                        cf_match = re.search(r'cloudflare:\s*(https?://[^\s]+)', content)
                        if hostc_match:
                            result['hostc'] = hostc_match.group(1).rstrip('/')
                        if cf_match:
                            result['cloudflare'] = cf_match.group(1).rstrip('/')
                        if not result['hostc']:
                            old_hostc_match = re.search(r'Public URL:\s*(https://[a-zA-Z0-9_-]+\.hostc\.dev)', content)
                            if old_hostc_match:
                                result['hostc'] = old_hostc_match.group(1).rstrip('/')
            except Exception as e:
                _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                pass
            return result
        
        def write_tunnel_urls_file(hostc_url=None, cf_url=None):
            """同时写入两个隧道的地址到 tunnel_url.txt
            
            Args:
                hostc_url: hostc 隧道的 URL（可选，None 表示保留已有值）
                cf_url: Cloudflare 隧道的 URL（可选，None 表示保留已有值）
            """
            try:
                existing = read_tunnel_urls_file()
                
                if hostc_url is None:
                    hostc_url = existing.get('hostc')
                if cf_url is None:
                    cf_url = existing.get('cloudflare')
                
                tunnel_file = PathManager.get_tunnel_url_file()
                with open(tunnel_file, 'w', encoding='utf-8') as f:
                    f.write(f"# Tunnel URLs - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"# Auto-generated by Szwego Crawler Tool\n\n")
                    
                    if hostc_url:
                        f.write(f"hostc: {hostc_url}\n")
                    
                    if cf_url:
                        f.write(f"cloudflare: {cf_url}\n")
                    
                    if not hostc_url and not cf_url:
                        f.write("# No active tunnels\n")
                
                print(f"[Tunnel] ✅ 已写入 tunnel_url.txt (hostc: {hostc_url or 'N/A'}, cf: {cf_url or 'N/A'})")
                return True
            except Exception as e:
                print(f"[Tunnel] ❌ 写入 tunnel_url.txt 失败: {e}")
                return False
        
        def send_tunnel_notification(new_url, event_type='new', force_send=False, tunnel_type='hostc'):
            """发送隧道通知邮件
            
            Args:
                new_url: 隧道 URL
                event_type: 事件类型 (new, stable_available, available, fallback_available, unavailable)
                force_send: 是否强制发送（跳过冷却期）
                tunnel_type: 隧道类型 (hostc, cloudflare)，用于独立去重
            """
            global last_email_sent_time, email_fail_count, last_email_sent_url, pending_email_url
            global cf_last_email_sent_time, cf_last_email_sent_url
            global global_last_email_sent_time, recent_sent_urls
            
            if tunnel_type == 'cloudflare':
                _last_sent_time = cf_last_email_sent_time
                _last_sent_url = cf_last_email_sent_url
            else:
                _last_sent_time = last_email_sent_time
                _last_sent_url = last_email_sent_url
            
            should_send = False
            
            with email_send_lock:
                current_time = time.time()

                if email_fail_count >= email_max_fail_count:
                    if current_time - last_email_sent_time < email_fail_cooldown:
                        print(f"[Email] 邮件发送失败次数过多 ({email_fail_count}次)，暂停发送 {email_fail_cooldown} 秒")
                        return
                    else:
                        print(f"[Email] 邮件发送失败冷却期已过，重置失败计数")
                        email_fail_count = 0

                expired_urls = [url for url, sent_time in recent_sent_urls.items() if current_time - sent_time > url_dedup_window]
                for url in expired_urls:
                    del recent_sent_urls[url]
                
                if new_url in recent_sent_urls:
                    sent_time = recent_sent_urls[new_url]
                    elapsed = int(current_time - sent_time)
                    print(f"[Email-{tunnel_type}] ⏭️ URL在去重窗口内已发送过，跳过: {new_url} (距上次{elapsed}秒)")
                    return
                
                if new_url == _last_sent_url and event_type != 'fallback_available':
                    print(f"[Email-{tunnel_type}] ⏭️ 相同URL已发送过，跳过: {new_url}")
                    return
                
                if not force_send:
                    if current_time - global_last_email_sent_time < global_email_cooldown:
                        elapsed = int(current_time - global_last_email_sent_time)
                        remaining = int(global_email_cooldown - elapsed)
                        print(f"[Email-{tunnel_type}] ⏳ 全局冷却中，跳过发送: {new_url} (剩余{remaining}秒)")
                        return
                    
                    if current_time - _last_sent_time < email_cooldown:
                        print(f"[Email-{tunnel_type}] 邮件冷却中，跳过发送: {new_url}")
                        return
                else:
                    if event_type not in ['unavailable', 'fallback_available']:
                        print(f"[Email-{tunnel_type}] ⚠️ 强制发送模式（事件: {event_type}）")
                
                should_send = True

            if not should_send:
                return

            def verify_and_send():
                global last_email_sent_time, email_fail_count, last_email_sent_url
                global cf_last_email_sent_time, cf_last_email_sent_url
                global global_last_email_sent_time, recent_sent_urls
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                thread_id = threading.current_thread().name
                try:
                    print(f"[{current_time_str}] [EmailNotifier-Thread:{thread_id}] 📧 开始发送邮件通知")
                    print(f"[{current_time_str}] [EmailNotifier-Thread:{thread_id}] 🎯 目标URL: {new_url}")
                    print(f"[{current_time_str}] [EmailNotifier-Thread:{thread_id}] 📋 事件类型: {event_type}")
                    print(f"[{current_time_str}] [EmailNotifier-Thread:{thread_id}] 🏷️ 隧道类型: {tunnel_type}")
                    
                    with email_send_lock:
                        success = email_notifier.send_tunnel_notification(new_url, event_type)
                        
                        if success:
                            send_time = time.time()
                            global_last_email_sent_time = send_time
                            recent_sent_urls[new_url] = send_time
                            if tunnel_type == 'cloudflare':
                                cf_last_email_sent_time = send_time
                                cf_last_email_sent_url = new_url
                            else:
                                last_email_sent_time = send_time
                                last_email_sent_url = new_url
                            email_fail_count = 0
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Email-{tunnel_type}] ✅✅✅ 邮件发送成功！")
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Email-{tunnel_type}] 🔗 隧道地址: {new_url}")
                        else:
                            email_fail_count += 1
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Email-{tunnel_type}] ❌❌❌ 邮件发送失败！")
                            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Email-{tunnel_type}] 📈 失败次数: {email_fail_count}/{email_max_fail_count}")

                except Exception as e:
                    email_fail_count += 1
                    error_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    print(f"[{error_time_str}] [Email-{tunnel_type}] 💥💥💥 邮件发送异常！")
                    print(f"[{error_time_str}] [Email-{tunnel_type}] 异常信息: {str(e)[:200]}")

            thread = threading.Thread(target=verify_and_send, daemon=True)
            thread.start()
        
        def check_and_send_pending_email():
            pass
        
        def verify_url(url, timeout=10, verbose=False, max_retries=3):
            for attempt in range(max_retries):
                try:
                    # 创建SSL上下文，保持默认的安全验证
                    ctx = ssl.create_default_context()
                    # 注意：不推荐禁用证书验证，除非有特殊需求
                    # 如果必须禁用，请添加配置选项和环境变量控制
                    
                    req = urllib.request.Request(url, method='HEAD')
                    req.add_header('User-Agent', 'hostc-verify/1.0')
                    
                    response = urllib.request.urlopen(req, timeout=timeout, context=ctx)
                    if response.status in [200, 301, 302, 307, 308]:
                        if verbose:
                            if attempt > 0:
                                print(f"[Email] ✅ URL验证成功 (第{attempt+1}次尝试): {url}")
                            else:
                                print(f"[Email] ✅ URL验证成功: {url}")
                        return True
                    return False
                except Exception as e:
                    if verbose:
                        print(f"[Email] URL验证失败 (第{attempt+1}/{max_retries}次): {url} - {str(e)[:100]}")
                    if attempt < max_retries - 1:
                        time.sleep(2)
                    continue
            return False
        
        def send_heartbeat():
            global tunnel_last_heartbeat, tunnel_heartbeat_failed
            web_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)
            if not web_url:
                tunnel_heartbeat_failed = True
                return False
            try:
                req = urllib.request.Request(web_url, method='HEAD')
                req.add_header('User-Agent', 'hostc-heartbeat/1.0')
                urllib.request.urlopen(req, timeout=3)
                tunnel_last_heartbeat = time.time()
                tunnel_heartbeat_failed = False
                return True
            except Exception as e:
                tunnel_heartbeat_failed = True
                return False
        
        def heartbeat_loop():
            global tunnel_process, tunnel_auto_restart, tunnel_need_restart, tunnel_url, tunnel_consecutive_failures
            global stable_url, stable_url_confirm_count, url_first_seen_time, last_stable_notification_time
            global tunnel_last_heartbeat, tunnel_heartbeat_failed, cf_url
            consecutive_failures = 0
            max_consecutive_failures = 5
            url_verify_failures = 0
            max_url_verify_failures = 2
            heartbeat_interval = 30
            last_log_time = 0
            grace_end_time = time.time() + 60
            last_url_sync_time = 0
            prev_web_url = None
            last_restart_state = False
            
            while tunnel_auto_restart:
                if tunnel_need_restart and not last_restart_state:
                    url_verify_failures = 0
                    stable_url_confirm_count = 0
                    stable_url = None
                    grace_end_time = time.time() + 60
                    print(f"[Tunnel] 🔄 检测到隧道重启，重置失败计数并进入60秒宽限期")
                last_restart_state = tunnel_need_restart
                
                # 优先使用内存中的 tunnel_url（解决文件锁定问题）
                if tunnel_url:
                    web_url = tunnel_url
                else:
                    web_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)
                is_tunnel_running = False
                url_verified = False

                if web_url:
                    if web_url != prev_web_url and prev_web_url is not None and stable_url_confirm_count >= stable_url_min_confirms:
                        grace_end_time = time.time() + 60
                        stable_url_confirm_count = 0
                        stable_url = None
                        print(f"[Tunnel] 🔄 URL变化({prev_web_url} → {web_url})，进入60秒宽限期")
                    prev_web_url = web_url

                    if time.time() < grace_end_time:
                        is_tunnel_running = True
                        remaining = int(grace_end_time - time.time())
                        if remaining >= 55:
                            print(f"[Tunnel] ⏳ 宽限期中（{remaining}秒），跳过URL验证")
                    else:
                        try:
                            url_verified = verify_url(web_url, timeout=10)
                            if url_verified:
                                is_tunnel_running = True
                                url_verify_failures = 0

                                if web_url != stable_url:
                                    stable_url = web_url
                                    stable_url_confirm_count = 1
                                    url_first_seen_time = time.time()
                                    print(f"[Tunnel] [搜索] 检测到新URL，开始稳定性验证 (1/{stable_url_min_confirms}): {web_url}")
                                elif stable_url_confirm_count < stable_url_min_confirms:
                                    stable_url_confirm_count += 1
                                    print(f"[Tunnel] ✅ URL稳定性验证 ({stable_url_confirm_count}/{stable_url_min_confirms}): {web_url}")
                                    if stable_url_confirm_count >= stable_url_min_confirms:
                                        elapsed = int(time.time() - url_first_seen_time)
                                        print(f"[Tunnel] 🎯 URL已确认为稳定！持续验证{stable_url_confirm_count}次，耗时{elapsed}秒")
                                        write_tunnel_urls_file(hostc_url=web_url, cf_url=cf_url)
                                        send_tunnel_notification(web_url, 'stable_available')
                                        last_stable_notification_time = time.time()
                            else:
                                url_verify_failures += 1
                                if stable_url_confirm_count > 0:
                                    print(f"[Tunnel] ⚠️ 公网地址不可用，重置稳定性计数 ({stable_url_confirm_count} -> 0)")
                                    stable_url_confirm_count = 0
                                    stable_url = None
                                    if cf_stable_url and cf_stable_confirm_count >= cf_stable_min_confirms:
                                        print(f"[Tunnel] 🔄 hostc 不可用，CF 仍有可用地址: {cf_stable_url}，发送备用地址通知")
                                        send_tunnel_notification(cf_stable_url, 'fallback_available', force_send=True)
                                if time.time() - last_log_time > 120:
                                    print(f"[Tunnel] ⚠️ 公网地址不可用: {web_url}，连续失败 {url_verify_failures}/{max_url_verify_failures} 次")
                                    last_log_time = time.time()
                                if url_verify_failures >= max_url_verify_failures:
                                    print(f"[Tunnel] 🚨 公网地址连续不可用{url_verify_failures}次，标记需要重启")
                                    send_tunnel_notification(web_url, 'unavailable', force_send=True)
                                    tunnel_need_restart = True
                        except Exception as e:
                            url_verify_failures += 1
                            if stable_url_confirm_count > 0:
                                stable_url_confirm_count = 0
                                stable_url = None
                            if time.time() - last_log_time > 120:
                                print(f"[Tunnel] URL验证异常: {e}")
                                last_log_time = time.time()
                            if url_verify_failures >= max_url_verify_failures:
                                tunnel_need_restart = True
                else:
                    if time.time() - last_log_time > 120:
                        print(f"[Tunnel] ⚠️ tunnel_url.txt 中未找到公网地址，隧道可能未启动")
                        last_log_time = time.time()
                    url_verify_failures += 1
                    if url_verify_failures >= max_url_verify_failures:
                        print(f"[Tunnel] 🚨 长时间未获取到公网地址，标记需要重启")
                        tunnel_need_restart = True
                
                if is_tunnel_running:
                    if url_verified:
                        tunnel_last_heartbeat = time.time()
                        tunnel_heartbeat_failed = False
                        if consecutive_failures > 0:
                            print(f"[Tunnel] 心跳恢复")
                            if url_verify_failures > 0 and web_url:
                                print(f"[Tunnel] 🎉 公网地址恢复，发送通知")
                                send_tunnel_notification(web_url, 'available')
                            last_log_time = time.time()
                        consecutive_failures = 0
                        check_and_send_pending_email()
                    else:
                        success = send_heartbeat()
                        if not success:
                            consecutive_failures += 1
                            if consecutive_failures >= max_consecutive_failures:
                                print(f"[Tunnel] 心跳连续失败 {consecutive_failures} 次，触发重启")
                                tunnel_need_restart = True
                                last_log_time = time.time()
                            elif consecutive_failures == 1 and time.time() - last_log_time > 10:
                                print(f"[Tunnel] 心跳异常: 网络不稳定 ({consecutive_failures}/{max_consecutive_failures})")
                                last_log_time = time.time()
                        else:
                            if consecutive_failures > 0:
                                print(f"[Tunnel] 心跳恢复")
                                last_log_time = time.time()
                            consecutive_failures = 0
                            check_and_send_pending_email()

                    if web_url and time.time() - last_url_sync_time > 300:
                        last_url_sync_time = time.time()
                        write_tunnel_urls_file(hostc_url=web_url, cf_url=cf_url)
                        web_output_file = PathManager.get_web_output_file()
                        try:
                            with open(web_output_file, 'a', encoding='utf-8') as wf:
                                wf.write(f"Public URL: {web_url}\n")
                        except Exception as e:
                            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                            pass
                time.sleep(heartbeat_interval)
        
        def start_tunnel_daemons():
            global tunnel_restart_thread, tunnel_heartbeat_thread, tunnel_daemon_started
            if tunnel_restart_thread is None or not tunnel_restart_thread.is_alive():
                if not tunnel_daemon_started:
                    tunnel_daemon_started = True
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] 启动自动重启守护进程")
                tunnel_restart_thread = threading.Thread(target=restart_tunnel, daemon=True)
                tunnel_restart_thread.start()
            if tunnel_heartbeat_thread is None or not tunnel_heartbeat_thread.is_alive():
                tunnel_heartbeat_thread = threading.Thread(target=heartbeat_loop, daemon=True)
                tunnel_heartbeat_thread.start()
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] 启动心跳守护进程（tunnel_url.txt 为唯一权威源）")

        def auto_start_tunnel(force_restart=False):
            global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_restart_thread, tunnel_restart_count, tunnel_last_error, tunnel_need_restart, tunnel_daemon_started, tunnel_type, old_tunnel_url, cf_url

            cf_binary = find_cloudflared_binary()
            if cf_binary and not force_restart:
                existing_urls = read_tunnel_urls_file()
                existing_cf = existing_urls.get('cloudflare')
                if existing_cf:
                    try:
                        is_cf_valid = verify_url(existing_cf, timeout=5, verbose=False)
                        if is_cf_valid:
                            print(f"[Tunnel] ✅ 发现可用CF地址，直接复用: {existing_cf}")
                            cf_url = existing_cf
                            start_cf_heartbeat()
                        else:
                            print(f"[Tunnel] ⚠️ 已有CF地址不可用: {existing_cf}，将启动新CF隧道")
                            existing_cf = None
                    except Exception as e:
                        print(f"[Tunnel] ⚠️ 验证已有CF地址失败: {e}，将启动新CF隧道")
                        existing_cf = None
                
                if existing_cf:
                    pass
                else:
                    port = args.port if "args" in globals() and hasattr(args, "port") else 8888
                    print(f"[Tunnel] 🚀 启动新的 Cloudflare Tunnel...")
                    cf_result = start_cloudflare_tunnel(port=port)
                    if cf_result and cf_result.get('success'):
                        print(f"[Tunnel] ✅ Cloudflare Tunnel 启动成功: {cf_result.get('url')}，等待心跳验证")
                        start_cf_heartbeat()
                    else:
                        cf_err = cf_result.get('error', '未知') if cf_result else '未知'
                        print(f"[Tunnel] ⚠️ Cloudflare Tunnel 启动失败: {cf_err}")
            elif cf_binary:
                port = args.port if "args" in globals() and hasattr(args, "port") else 8888
                print(f"[Tunnel] 🚀 强制重启 Cloudflare Tunnel...")
                cf_result = start_cloudflare_tunnel(port=port)
                if cf_result and cf_result.get('success'):
                    print(f"[Tunnel] ✅ Cloudflare Tunnel 启动成功: {cf_result.get('url')}，等待心跳验证")
                    start_cf_heartbeat()
                else:
                    cf_err = cf_result.get('error', '未知') if cf_result else '未知'
                    print(f"[Tunnel] ⚠️ Cloudflare Tunnel 启动失败: {cf_err}")
            else:
                print(f"[Tunnel] ⏭️ 未找到 cloudflared，跳过 Cloudflare Tunnel")

            if force_restart:
                print(f"[Tunnel] 🔄 强制重启模式，将清理旧进程并重新启动")
                sys.stdout.flush()
            else:
                has_hostc_process = Environment.check_process_running(Environment.HOSTC_PROCESS_NAME)
                web_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)

                if web_url and has_hostc_process:
                    tunnel_url = web_url
                    old_tunnel_url = web_url
                    print(f"[Tunnel] ✅ 发现公网地址: {web_url}，后台验证并发邮件")
                    sys.stdout.flush()

                    def _verify_and_notify_found_url(url):
                        global stable_url, stable_url_confirm_count, url_first_seen_time, last_stable_notification_time, last_email_sent_url
                        try:
                            url_verified = verify_url(url, timeout=10, verbose=True)
                        except Exception as e:
                            logger.debug(f"Exception verifying URL: {e}")
                            url_verified = False
                        if url_verified:
                            print(f"[Tunnel] 🎉 公网地址验证通过！立即发送邮件通知...")
                            send_tunnel_notification(url, 'stable_available')
                            stable_url = url
                            stable_url_confirm_count = stable_url_min_confirms
                            url_first_seen_time = time.time()
                            last_stable_notification_time = time.time()
                        else:
                            print(f"[Tunnel] ⏳ 公网地址暂不可用，将由心跳机制持续验证后发送邮件")

                    threading.Thread(target=_verify_and_notify_found_url, args=(web_url,), daemon=True).start()
                    return {'success': True, 'url': tunnel_url, 'message': f'发现已有URL: {tunnel_url}，后台验证中'}

                if web_url and not has_hostc_process:
                    print(f"[Tunnel] ⚠️ 发现旧URL {web_url} 但 hostc 进程已不在运行，hostc地址已失效")
                    try:
                        existing_urls = read_tunnel_urls_file()
                        cf_url = existing_urls.get('cloudflare')
                        if cf_url:
                            write_tunnel_urls_file(hostc_url=None, cf_url=cf_url)
                            print(f"[Tunnel] ✅ 已保留CF地址，仅清除hostc: {cf_url}")
                        else:
                            tunnel_file_to_clear = PathManager.get_tunnel_url_file()
                            with open(tunnel_file_to_clear, 'w', encoding='utf-8') as f:
                                f.write('')
                            print(f"[Tunnel] 已清除过期 tunnel_url.txt (无CF地址)")
                    except Exception as clear_err:
                        print(f"[Tunnel] 清除 hostc URL 失败: {clear_err}")
                    sys.stdout.flush()

                if has_hostc_process:
                    print(f"[Tunnel] [搜索] hostc在运行，后台等待URL出现后验证发邮件")
                    sys.stdout.flush()

                    def _wait_and_notify_hostc_url():
                        global tunnel_url, old_tunnel_url, stable_url, stable_url_confirm_count, url_first_seen_time, last_stable_notification_time, last_email_sent_url
                        global tunnel_need_restart
                        for _ in range(30):
                            time.sleep(2)
                            has_hostc = Environment.check_process_running(Environment.HOSTC_PROCESS_NAME)
                            if not has_hostc:
                                print(f"[Tunnel] ❌ hostc进程已退出，标记需要重启")
                                tunnel_need_restart = True
                                return
                            found_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)
                            if found_url:
                                tunnel_url = found_url
                                old_tunnel_url = found_url
                                print(f"[Tunnel] ✅ 后台获取到URL: {found_url}")
                                try:
                                    url_verified = verify_url(found_url, timeout=10, verbose=True)
                                except Exception as e:
                                    logger.debug(f"Exception verifying URL: {e}")
                                    url_verified = False
                                if url_verified:
                                    print(f"[Tunnel] 🎉 公网地址验证通过！立即发送邮件通知...")
                                    send_tunnel_notification(found_url, 'stable_available')
                                    stable_url = found_url
                                    stable_url_confirm_count = stable_url_min_confirms
                                    url_first_seen_time = time.time()
                                    last_stable_notification_time = time.time()
                                else:
                                    print(f"[Tunnel] ⏳ 公网地址暂不可用，将由心跳机制持续验证后发送邮件")
                                return
                        print(f"[Tunnel] ⏳ 后台等待URL超时，将由心跳机制继续获取")

                    threading.Thread(target=_wait_and_notify_hostc_url, daemon=True).start()
                    return {'success': True, 'url': None, 'message': 'hostc在运行，后台等待URL'}

                print(f"[Tunnel] 无hostc进程且无URL，需要启动新隧道")
                sys.stdout.flush()
            
            try:
                port = args.port
                tunnel_url = None
                url_ready = False
                tunnel_last_error = None
                tunnel_auto_restart = True

                print("[Tunnel] 启动 hostc 隧道")
                sys.stdout.flush()
                tunnel_type = 'hostc'

                # Fix: Clear old tunnel URL to prevent reading stale URL
                try:
                    tunnel_file_to_clear = PathManager.get_tunnel_url_file()
                    with open(tunnel_file_to_clear, 'w', encoding='utf-8') as f:
                        f.write('')
                    print("[Tunnel] Cleared old URL file, waiting for new address...")
                    sys.stdout.flush()
                except Exception as clear_err:
                    print(f"[Tunnel] Failed to clear old URL: {clear_err}")
                    sys.stdout.flush()

                # Clean up old hostc processes BEFORE starting new one
                Environment.kill_process_by_name(Environment.HOSTC_PROCESS_NAME)

                # 等待2秒确保旧进程完全退出（避免端口冲突）
                time.sleep(2)

                hostc_script = 'hostc' + ('.cmd' if Environment.IS_WINDOWS else '')
                hostc_bin = os.path.join(PROJECT_DIR, 'dist', 'node_modules', '.bin', hostc_script)
                if not os.path.isfile(hostc_bin):
                    hostc_bin = 'npx hostc'
                
                if not isinstance(port, int) or not (1 <= port <= 65535):
                    print(f"[Tunnel] ⚠️ 无效的端口号: {port}, 使用默认8888")
                    port = 8888
                
                env = os.environ.copy()
                env['HOSTC_DEBUG'] = '1'
                
                tunnel_process = subprocess.Popen(
                    [hostc_bin, str(port), '--local-host', 'localhost'] if hostc_bin != 'npx hostc' else ['npx', 'hostc', str(port), '--local-host', 'localhost'],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    stdin=subprocess.DEVNULL,
                    text=True,
                    bufsize=0,
                    env=env,
                    creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0) if Environment.IS_WINDOWS else 0
                )
                
                print(f"[Tunnel] 🆔 hostc进程已启动，PID: {tunnel_process.pid}")
                sys.stdout.flush()

                def read_output():
                    global tunnel_url, url_ready, old_tunnel_url, tunnel_consecutive_failures
                    global stable_url, stable_url_confirm_count, url_first_seen_time, last_stable_notification_time, last_email_sent_url
                    global tunnel_need_restart
                    if not tunnel_process or not tunnel_process.stdout:
                        return
                    
                    buffer = ''
                    while True:
                        if tunnel_process is None:
                            break
                        if tunnel_process.poll() is not None:
                            exit_code = tunnel_process.poll()
                            print(f"[Tunnel] ❌ hostc进程已退出 (exit code: {exit_code})，标记需要重启")
                            if buffer.strip():
                                print(f"[Tunnel] 📋 hostc输出内容:\n{buffer.strip()}")
                            tunnel_need_restart = True
                            sys.stdout.flush()
                            break
                        
                        try:
                            char = tunnel_process.stdout.read(1)
                            if char:
                                buffer += char
                                if '\n' in buffer:
                                    lines = buffer.split('\n')
                                    for line in lines[:-1]:
                                        if line.strip():
                                            print(f"[hostc] {line.strip()}")
                                    buffer = lines[-1]
                                match = re.search(r'(https://[a-zA-Z0-9_-]+\.hostc\.dev)', buffer)
                                if match:
                                    file_url = match.group(1).rstrip('/')
                                    if file_url and file_url != tunnel_url:
                                        print(f"[Tunnel] 从 hostc 输出获取到URL: {file_url}")
                                        
                                        write_tunnel_urls_file(hostc_url=file_url, cf_url=cf_url)
                                        
                                        try:
                                            web_output_file = PathManager.get_web_output_file()
                                            with open(web_output_file, 'a', encoding='utf-8') as wf:
                                                wf.write(f"Public URL: {file_url}\n")
                                            print(f"[Tunnel] 已写入 web_output.log")
                                        except Exception as e:
                                            pass
                                        
                                        tunnel_url = file_url
                                        url_ready = True
                                        tunnel_consecutive_failures = 0
                                        old_tunnel_url = file_url

                                        print(f"[Tunnel] ✅ URL已就绪: {tunnel_url}")
                                        
                                        url_verified = False
                                        try:
                                            url_verified = verify_url(file_url, timeout=10, verbose=True)
                                        except Exception as e:
                                            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                                            pass
                                        
                                        if url_verified:
                                            print(f"[Tunnel] 🎉 公网地址验证通过！立即发送邮件通知...")
                                            send_tunnel_notification(file_url, 'stable_available')
                                            stable_url = file_url
                                            stable_url_confirm_count = stable_url_min_confirms
                                            url_first_seen_time = time.time()
                                            last_stable_notification_time = time.time()
                                        else:
                                            print(f"[Tunnel] ⏳ 公网地址暂不可用，将由心跳机制持续验证后发送邮件")
                                        
                                        sys.stdout.flush()
                                        return
                                
                                if len(buffer) > 1000:
                                    buffer = buffer[-500:]
                            else:
                                time.sleep(0.1)
                        except Exception as e:
                            time.sleep(0.1)

                read_thread = threading.Thread(target=read_output, daemon=True)
                read_thread.start()
                
                if force_restart:
                    read_thread.join(timeout=10)
                    if tunnel_url:
                        print(f"[Tunnel] 隧道启动成功: {tunnel_url}")
                        return {'success': True, 'url': tunnel_url, 'message': f'隧道已启动，URL: {tunnel_url}'}
                    else:
                        print(f"[Tunnel] 启动超时，未获取到URL，将由心跳机制继续获取")
                        return {'success': True, 'url': None, 'message': '隧道已启动，URL由心跳机制获取'}
                else:
                    print(f"[Tunnel] 🚀 hostc已启动，URL将由心跳机制获取和验证")
                    return {'success': True, 'url': None, 'message': 'hostc已启动，URL由心跳机制获取'}
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        def restart_tunnel():
            global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_last_error, tunnel_restart_count, tunnel_need_restart, old_tunnel_url, tunnel_consecutive_failures, tunnel_backoff_delay, cf_url
            
            consecutive_restart_attempts = 0
            restart_wait_start = None
            grace_period_end = None
            verify_fail_count = 0
            
            def _do_restart(has_hostc_process, web_url, is_url_valid):
                nonlocal restart_wait_start, grace_period_end, consecutive_restart_attempts
                global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_last_error, tunnel_restart_count, tunnel_need_restart, old_tunnel_url, tunnel_consecutive_failures, tunnel_backoff_delay
                
                restart_wait_start = None
                tunnel_need_restart = False
                tunnel_restart_count += 1
                
                if consecutive_restart_attempts > 0:
                    backoff = min(60 * (2 ** (consecutive_restart_attempts - 1)), 300)
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] 🔄 连续重启失败{consecutive_restart_attempts}次，退避{backoff}秒后重试 (第{tunnel_restart_count}次)")
                    time.sleep(backoff)
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] 🔄 立即执行重启 (第{tunnel_restart_count}次)")
                
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] - hostc进程: {'运行中' if has_hostc_process else '未运行'}")
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] - 公网URL: {web_url if web_url else '无'}")
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] - URL有效: {'是' if is_url_valid else '否'}")
                sys.stdout.flush()
                
                Environment.kill_process_by_name(Environment.HOSTC_PROCESS_NAME)
                if tunnel_process:
                    try:
                        tunnel_process.terminate()
                        tunnel_process.wait(timeout=2)
                    except Exception as e:
                        logger.debug(f"Nested exception: {e}")
                        try:
                            tunnel_process.kill()
                        except Exception as e:
                            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                            pass
                
                saved_old_url = old_tunnel_url
                tunnel_process = None
                tunnel_url = None
                old_tunnel_url = None
                
                time.sleep(tunnel_restart_delay)
                
                if not tunnel_auto_restart:
                    return False
                
                try:
                    result = auto_start_tunnel()
                    if result['success']:
                        new_url = result.get('url')
                        if new_url and saved_old_url and saved_old_url != new_url:
                            print(f"[Tunnel] 隧道URL已变化: {saved_old_url} -> {new_url}")
                        consecutive_restart_attempts = 0
                    else:
                        consecutive_restart_attempts += 1
                except Exception as e:
                    consecutive_restart_attempts += 1
                    print(f"[Tunnel] 重启失败: {e}")
                
                grace_period_end = time.time() + 60
                return True
            
            while tunnel_auto_restart:
                now = time.time()
                
                if grace_period_end and now < grace_period_end:
                    time.sleep(3)
                    continue
                grace_period_end = None
                
                web_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)
                has_hostc_process = Environment.check_process_running(Environment.HOSTC_PROCESS_NAME)
                
                if has_hostc_process and web_url:
                    is_url_valid = False
                    if web_url:
                        try:
                            is_url_valid = verify_url(web_url)
                        except Exception as e:
                            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                            pass
                    
                    if is_url_valid:
                        verify_fail_count = 0
                        restart_wait_start = None
                        tunnel_need_restart = False
                        time.sleep(1)
                        continue
                    
                    verify_fail_count += 1
                    if verify_fail_count < 2:
                        time.sleep(5)
                        continue
                    
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] ⚠️ URL连续{verify_fail_count}次验证失败，触发重启")
                    sys.stdout.flush()
                    verify_fail_count = 0
                    if not _do_restart(has_hostc_process, web_url, False):
                        break
                    continue
                
                if tunnel_need_restart:
                    if not _do_restart(has_hostc_process, web_url, False):
                        break
                    continue
                
                if not has_hostc_process:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] ❌ hostc进程已退出，立即重启")
                    sys.stdout.flush()
                    if not _do_restart(False, web_url, False):
                        break
                    continue
                
                if not web_url:
                    if restart_wait_start is None:
                        restart_wait_start = time.time()
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] ⏳ hostc运行中但URL未就绪，等待...")
                        sys.stdout.flush()
                    
                    elapsed = now - restart_wait_start
                    if elapsed < 30:
                        time.sleep(3)
                        continue
                    
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [Tunnel] ⚠️ 等待超过30秒URL仍未就绪，触发重启")
                    sys.stdout.flush()
                    if not _do_restart(True, None, False):
                        break
        

        # ========================================
        # Cloudflare Tunnel 支持
        # ========================================

        def find_cloudflared_binary():
            """跨平台 cloudflared 二进制文件检测"""
            
            system = platform.system().lower()
            machine = platform.machine().lower()
            
            # 1. 优先检查项目目录（按操作系统分类）
            project_cf = None
            cf_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "cloudflared")
            if system == "windows":
                cf_dir = os.path.join(cf_base, "windows")
            elif system == "linux":
                cf_dir = os.path.join(cf_base, "linux")
            elif system == "darwin":
                cf_dir = os.path.join(cf_base, "macos")
            else:
                cf_dir = None
            if cf_dir and os.path.isdir(cf_dir):
                for fname in sorted(os.listdir(cf_dir), reverse=True):
                    fpath = os.path.join(cf_dir, fname)
                    if os.path.isfile(fpath) and (os.access(fpath, os.X_OK) or system == "windows"):
                        project_cf = fpath
                        break
            
            if project_cf and os.path.exists(project_cf):
                print(f"[Cloudflare] 在项目目录找到: {project_cf}")
                return project_cf
            
            # 2. 检查系统 PATH
            cf_in_path = shutil.which("cloudflared")
            if cf_in_path:
                print(f"[Cloudflare] 在系统 PATH 中找到: {cf_in_path}")
                return cf_in_path
            
            # 3. 检查系统常见路径
            if system == "windows":
                common_paths = []
                for env_var in ('PROGRAMFILES(X86)', 'PROGRAMFILES'):
                    base = os.environ.get(env_var, '')
                    if base:
                        common_paths.append(os.path.join(base, 'cloudflared', 'cloudflared' + Environment.EXE_SUFFIX))
                local_app = os.environ.get('LOCALAPPDATA', '')
                if local_app:
                    common_paths.append(os.path.join(local_app, 'cloudflared', 'cloudflared' + Environment.EXE_SUFFIX))
            elif system == "linux":
                common_paths = [
                    os.path.join('/usr', 'local', 'bin', 'cloudflared'),
                    os.path.join('/usr', 'bin', 'cloudflared'),
                ]
            elif system == "darwin":
                common_paths = [
                    os.path.join('/usr', 'local', 'bin', 'cloudflared'),
                ]
            else:
                common_paths = []
                for path in common_paths:
                    if os.path.exists(path):
                        print(f"[Cloudflare] 在常见路径找到: {path}")
                        return path
            
            print(f"[Cloudflare] 未找到 cloudflared (系统: {system})")
            return None

        def _detect_named_tunnel_config():
            """自动检测 .cloudflared 目录下的 Named Tunnel 配置
            返回: {'available': bool, 'tunnel_name': str, 'custom_domain': str, 'tunnel_id': str, 'config_yml_path': str}"""
            tunnel_dir = os.path.join(PROJECT_DIR, '.cloudflared')
            config_yml_path = os.path.join(tunnel_dir, 'config.yml')

            if not os.path.exists(config_yml_path):
                return {'available': False, 'tunnel_name': '', 'custom_domain': '', 'tunnel_id': '', 'config_yml_path': ''}

            try:
                with open(config_yml_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                tunnel_name = ''
                custom_domain = ''
                tunnel_id = ''

                for line in content.split('\n'):
                    line = line.strip()
                    if line.startswith('tunnel:'):
                        tunnel_name = line.split(':', 1)[1].strip()
                    elif line.startswith('hostname:'):
                        custom_domain = line.split(':', 1)[1].strip()

                credentials_path = os.path.join(tunnel_dir, f'{tunnel_name}.json')
                if os.path.exists(credentials_path):
                    with open(credentials_path, 'r') as f:
                        cred = json.load(f)
                    tunnel_id = cred.get('TunnelID') or cred.get('tunnel_id', '')

                if tunnel_name and custom_domain and tunnel_id:
                    return {
                        'available': True,
                        'tunnel_name': tunnel_name,
                        'custom_domain': custom_domain,
                        'tunnel_id': tunnel_id,
                        'config_yml_path': config_yml_path
                    }
            except Exception as e:
                _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                pass

            return {'available': False, 'tunnel_name': '', 'custom_domain': '', 'tunnel_id': '', 'config_yml_path': ''}

        def _ensure_named_tunnel_ready(cf_binary, tunnel_name, custom_domain, port):
            """确保 named tunnel 已创建并配置好 DNS 路由（首次运行时自动设置）
            返回: (tunnel_id, config_yml_path) 或 (None, None) 表示失败"""
            tunnel_dir = os.path.join(PROJECT_DIR, '.cloudflared')
            os.makedirs(tunnel_dir, exist_ok=True)
            config_yml_path = os.path.join(tunnel_dir, 'config.yml')
            credentials_path = os.path.join(tunnel_dir, f'{tunnel_name}.json')

            if os.path.exists(credentials_path) and os.path.exists(config_yml_path):
                try:
                    with open(credentials_path, 'r') as f:
                        cred = json.load(f)
                    tunnel_id = cred.get('TunnelID') or cred.get('tunnel_id', '')
                    if tunnel_id:
                        print(f"[Cloudflare] ✅ Named tunnel 已存在: {tunnel_name} (ID: {tunnel_id})")
                        return tunnel_id, config_yml_path
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass

            print(f"[Cloudflare] 🔧 首次使用 named tunnel，开始自动配置...")

            print(f"[Cloudflare] 步骤1/3: 创建 tunnel '{tunnel_name}'...")
            try:
                result = subprocess.run(
                    [cf_binary, "tunnel", "create", tunnel_name],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode != 0:
                    err_msg = result.stderr.strip() or result.stdout.strip()
                    print(f"[Cloudflare] ❌ 创建 tunnel 失败: {err_msg}")
                    if 'already exists' in err_msg.lower():
                        print(f"[Cloudflare] ⚠️ Tunnel 已存在，尝试列出获取 ID...")
                        list_result = subprocess.run(
                            [cf_binary, "tunnel", "list"],
                            capture_output=True, text=True, timeout=30
                        )
                        tunnel_id = None
                        for line in list_result.stdout.split('\n'):
                            if tunnel_name in line:
                                parts = line.split()
                                if parts:
                                    tunnel_id = parts[0]
                                    break
                        if tunnel_id:
                            print(f"[Cloudflare] ✅ 找到已有 tunnel ID: {tunnel_id}")
                        else:
                            return None, None
                    else:
                        return None, None
                else:
                    tunnel_id = None
                    for line in result.stdout.split('\n'):
                        id_match = re.search(r'([a-f0-9\-]{36})', line)
                        if id_match:
                            tunnel_id = id_match.group(1)
                            break
                    if not tunnel_id:
                        print(f"[Cloudflare] ❌ 无法从输出中解析 tunnel ID")
                        return None, None
                    print(f"[Cloudflare] ✅ Tunnel 创建成功，ID: {tunnel_id}")
            except Exception as e:
                print(f"[Cloudflare] ❌ 创建 tunnel 异常: {e}")
                return None, None

            print(f"[Cloudflare] 步骤2/3: 配置 DNS 路由 ({custom_domain} → {tunnel_name})...")
            try:
                result = subprocess.run(
                    [cf_binary, "tunnel", "route", "dns", tunnel_name, custom_domain],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode != 0:
                    err_msg = result.stderr.strip() or result.stdout.strip()
                    print(f"[Cloudflare] ⚠️ DNS 路由配置: {err_msg}")
                    if 'already' not in err_msg.lower() and 'exists' not in err_msg.lower():
                        print(f"[Cloudflare] ❌ DNS 路由配置失败，将降级到 quick tunnel")
                        return None, None
                else:
                    print(f"[Cloudflare] ✅ DNS 路由配置成功: {custom_domain} → {tunnel_name}")
            except Exception as e:
                print(f"[Cloudflare] ❌ DNS 路由配置异常: {e}")
                return None, None

            print(f"[Cloudflare] 步骤3/3: 生成 config.yml...")
            try:
                config_content = f"""tunnel: {tunnel_name}
credentials-file: {credentials_path}

ingress:
  - hostname: {custom_domain}
    service: http://localhost:{port}
  - service: http_status:404
"""
                with open(config_yml_path, 'w', encoding='utf-8') as f:
                    f.write(config_content)
                print(f"[Cloudflare] ✅ config.yml 已生成: {config_yml_path}")
            except Exception as e:
                print(f"[Cloudflare] ❌ 生成 config.yml 失败: {e}")
                return None, None

            return tunnel_id, config_yml_path

        def start_cloudflare_tunnel(port=8888, timeout=120):
            """启动 Cloudflare Tunnel（Plan A: Named Tunnel → Plan B: Quick Tunnel，保底至少成功一个）
            启动成功后由 cf_heartbeat_loop 独立验证并发邮件"""
            global cf_process, cf_url, cf_mode

            cf_binary = find_cloudflared_binary()
            if not cf_binary:
                return {"success": False, "error": "未找到 cloudflared"}

            named_config = _detect_named_tunnel_config()

            if named_config['available']:
                print(f"[Cloudflare] 🏠 Plan A: Named Tunnel (自定义域名: {named_config['custom_domain']})...")
                try:
                    print(f"[Cloudflare] 启动 Named Tunnel: {named_config['tunnel_name']}...")
                    cmd = [cf_binary, "tunnel", "run", named_config['tunnel_name'], "--config", named_config['config_yml_path'], "--no-autoupdate"]

                    cf_process = subprocess.Popen(
                        cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        bufsize=1
                    )

                    named_url = f"https://{named_config['custom_domain']}"
                    start_time = time.time()
                    while time.time() - start_time < 15:
                        if cf_process.poll() is not None:
                            break
                        time.sleep(1)

                    if cf_process.poll() is None:
                        cf_url = named_url
                        cf_mode = 'named'
                        print(f"[Cloudflare] ✅ Plan A 成功: Named Tunnel {cf_url}，等待心跳验证后发邮件")

                        existing = read_tunnel_urls_file()
                        write_tunnel_urls_file(hostc_url=existing.get('hostc'), cf_url=cf_url)

                        return {"success": True, "url": cf_url, "type": "cloudflare", "mode": "named"}
                    else:
                        print(f"[Cloudflare] ❌ Plan A 失败: Named Tunnel 进程退出 (code: {cf_process.returncode})，回退到 Plan B...")
                        cf_process = None
                except Exception as e:
                    print(f"[Cloudflare] ❌ Plan A 失败: Named Tunnel 启动异常: {e}，回退到 Plan B...")
                    cf_process = None
            else:
                print(f"[Cloudflare] ⏭️ Plan A 跳过: 未检测到 Named Tunnel 配置，直接 Plan B...")

            print(f"[Cloudflare] 🚀 Plan B: Quick Tunnel (临时域名, 端口 {port})...")
            try:
                cmd = [cf_binary, "tunnel", "--url", f"http://localhost:{port}", "--no-autoupdate"]

                cf_process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1
                )

                url_pattern = r"https://[a-z0-9\-]+\.trycloudflare\.com"
                rate_limit_pattern = r"(429|error code: 1015|Too Many Requests)"
                start_time = time.time()

                while time.time() - start_time < timeout:
                    if cf_process.poll() is not None:
                        return {"success": False, "error": f"Plan B 也失败了: Quick Tunnel 进程退出 (code: {cf_process.returncode})"}

                    line = cf_process.stdout.readline()
                    if line:
                        if re.search(rate_limit_pattern, line, re.IGNORECASE):
                            print(f"[Cloudflare] ⚠️ Quick Tunnel 请求被限流 (429 Too Many Requests)")
                            print(f"[Cloudflare] 💡 建议: 等待 5-10 分钟后重试，或配置 Named Tunnel")
                            return {"success": False, "error": "Quick Tunnel 限流，请稍后重试"}
                        
                        match = re.search(url_pattern, line)
                        if match:
                            cf_url = match.group(0)
                            cf_mode = 'quick'
                            print(f"[Cloudflare] ✅ Plan B 成功: Quick Tunnel {cf_url}，等待心跳验证后发邮件")

                            existing = read_tunnel_urls_file()
                            write_tunnel_urls_file(hostc_url=existing.get('hostc'), cf_url=cf_url)

                            return {"success": True, "url": cf_url, "type": "cloudflare", "mode": "quick"}

                    time.sleep(0.5)

                return {"success": False, "error": f"Plan B 也失败了: Quick Tunnel 等待 URL 超时 ({timeout}秒)"}

            except Exception as e:
                return {"success": False, "error": f"Plan B 也失败了: {str(e)}"}

        def cf_heartbeat_loop():
            """Cloudflare Tunnel 独立心跳验证 - 验证通过后发邮件通知"""
            global cf_process, cf_url, cf_mode
            global cf_stable_url, cf_stable_confirm_count, cf_url_first_seen_time, cf_last_stable_notification_time, cf_last_email_sent_url
            global stable_url

            interval = 30
            last_log_time = 0

            while True:
                time.sleep(interval)

                if cf_process is None or cf_process.poll() is not None:
                    if cf_url:
                        print(f"[CF-Heartbeat] ⚠️ CF 隧道进程已退出")
                        cf_url = None
                        cf_mode = None
                        cf_stable_url = None
                        cf_stable_confirm_count = 0
                        cf_last_email_sent_url = None
                    continue

                if not cf_url:
                    continue

                try:
                    url_verified = verify_url(cf_url, timeout=10, verbose=True)
                except Exception as e:
                    print(f"[CF-Heartbeat] ❌ CF URL 验证异常: {str(e)[:100]}")
                    url_verified = False

                if url_verified:
                    if cf_url != cf_stable_url:
                        cf_stable_url = cf_url
                        cf_stable_confirm_count = 1
                        cf_url_first_seen_time = time.time()
                        print(f"[CF-Heartbeat] [搜索] CF 新URL，开始稳定性验证 (1/{cf_stable_min_confirms}): {cf_url}")
                        print(f"[CF-Heartbeat] 🎯 CF URL 已确认稳定！持续验证{cf_stable_confirm_count}次，耗时0秒")
                        write_tunnel_urls_file(hostc_url=stable_url, cf_url=cf_url)
                        if cf_url != cf_last_email_sent_url:
                            print(f"[CF-Heartbeat] 🎉 公网地址验证通过！立即发送邮件通知...")
                            send_tunnel_notification(cf_url, 'stable_available', tunnel_type='cloudflare')
                            cf_last_stable_notification_time = time.time()
                            cf_last_email_sent_url = cf_url
                        else:
                            print(f"[CF-Heartbeat] ⏭️ CF URL 已发送过邮件，跳过重复发送")
                    elif cf_stable_confirm_count < cf_stable_min_confirms:
                        cf_stable_confirm_count += 1
                        print(f"[CF-Heartbeat] ✅ CF 稳定性验证 ({cf_stable_confirm_count}/{cf_stable_min_confirms}): {cf_url}")
                        elapsed = int(time.time() - cf_url_first_seen_time)
                        print(f"[CF-Heartbeat] 🎯 CF URL 已确认稳定！持续验证{cf_stable_confirm_count}次，耗时{elapsed}秒")
                        write_tunnel_urls_file(hostc_url=stable_url, cf_url=cf_url)
                        if cf_url != cf_last_email_sent_url:
                            print(f"[CF-Heartbeat] 🎉 公网地址验证通过！立即发送邮件通知...")
                            send_tunnel_notification(cf_url, 'stable_available', tunnel_type='cloudflare')
                            cf_last_stable_notification_time = time.time()
                            cf_last_email_sent_url = cf_url
                        else:
                            print(f"[CF-Heartbeat] ⏭️ CF URL 已发送过邮件，跳过重复发送")
                else:
                    if cf_stable_confirm_count > 0:
                        print(f"[CF-Heartbeat] ⚠️ CF URL 不可用，重置稳定性计数 ({cf_stable_confirm_count} -> 0)")
                        cf_stable_confirm_count = 0
                        cf_stable_url = None
                        if stable_url and stable_url_confirm_count >= stable_url_min_confirms:
                            print(f"[CF-Heartbeat] 🔄 CF 不可用，hostc 仍有可用地址: {stable_url}，发送备用地址通知")
                            send_tunnel_notification(stable_url, 'fallback_available', force_send=True)
                    if time.time() - last_log_time > 120:
                        print(f"[CF-Heartbeat] ⚠️ CF URL 不可用: {cf_url}")
                        last_log_time = time.time()

        def start_cf_heartbeat():
            """启动 CF 心跳守护线程"""
            global cf_heartbeat_thread
            if cf_heartbeat_thread is None or not cf_heartbeat_thread.is_alive():
                cf_heartbeat_thread = threading.Thread(target=cf_heartbeat_loop, daemon=True)
                cf_heartbeat_thread.start()
                print(f"[CF-Heartbeat] 启动 Cloudflare Tunnel 心跳守护进程")

        @app.api_route('/api/tunnel/type', methods=['GET', 'POST'])
        def tunnel_type_api():
            """获取隧道类型状态（CF 和 hostc 同时运行）"""
            cf_available = find_cloudflared_binary() is not None
            hostc_running = Environment.check_process_running(Environment.HOSTC_PROCESS_NAME)
            cf_running = cf_process is not None and cf_process.poll() is None
            
            current = 'hostc'
            if cf_running and hostc_running:
                current = 'cloudflare'
            elif cf_running:
                current = 'cloudflare'
            elif hostc_running:
                current = 'hostc'
            elif cf_available:
                current = 'cloudflare'
            else:
                current = 'hostc'
            
            return jsonify({
                'mode': 'dual',
                'current': current,
                'hostc': {'available': True, 'running': hostc_running},
                'cloudflare': {'available': cf_available, 'running': cf_running}
            })

        @app.post('/api/tunnel/start')
        def start_tunnel():
            global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_restart_thread, tunnel_restart_count, tunnel_last_error, tunnel_need_restart, tunnel_daemon_started, tunnel_type, tunnel_consecutive_failures
            global stable_url, stable_url_confirm_count, url_first_seen_time

            tunnel_need_restart = False
            tunnel_consecutive_failures = 0
            
            # 重置稳定性检测状态，确保从头开始验证
            old_stable_url = stable_url
            stable_url = None
            stable_url_confirm_count = 0
            url_first_seen_time = time.time()
            
            print(f"[Tunnel/API] 🚀 收到手动启动请求（优先复用，强制重启作为备用）")
            print(f"[Tunnel/API] 📊 重置稳定性检测 (旧URL: {old_stable_url})")
            
            if tunnel_process and tunnel_process.poll() is None:
                return jsonify({
                    'success': True,
                    'url': tunnel_url,
                    'message': '隧道已在运行',
                    'status': 'running',
                    'stable_confirmed': stable_url == tunnel_url and stable_url_confirm_count >= stable_url_min_confirms
                })

            result = auto_start_tunnel(force_restart=False)
            if result['success'] and result.get('url'):
                new_url = result.get('url')
                _min_confirms_api = globals().get('stable_url_min_confirms', 3)
                response_data = {
                    'success': True,
                    'url': new_url,
                    'message': f'复用已有可用隧道地址 ({_min_confirms_api}次连续验证)',
                    'status': 'verifying',
                    'verify_progress': {
                        'current': 0,
                        'required': stable_url_min_confirms,
                        'estimated_time_seconds': stable_url_min_confirms * 30
                    },
                    'next_notification': '等待稳定性确认后自动发送邮件通知'
                }
                print(f"[Tunnel/API] ✅ 复用可用隧道: {new_url}")
                sys.stdout.flush()
                return jsonify(response_data)

            if not result.get('url'):
                print(f"[Tunnel/API] ⚠️ 正常模式未获取到URL，启用备用：强制重启...")
                sys.stdout.flush()
                result = auto_start_tunnel(force_restart=True)

            if result['success']:
                new_url = result.get('url')
                _min_confirms_api = globals().get('stable_url_min_confirms', 3)
                response_data = {
                    'success': True,
                    'url': new_url,
                    'message': f'备用方案：强制重启隧道，正在验证稳定性 ({_min_confirms_api}次连续验证)',
                    'status': 'verifying',
                    'verify_progress': {
                        'current': 0,
                        'required': stable_url_min_confirms,
                        'estimated_time_seconds': stable_url_min_confirms * 30
                    },
                    'next_notification': '等待稳定性确认后自动发送邮件通知'
                }
                
                if new_url:
                    print(f"[Tunnel/API] ✅ 备用方案启动成功: {new_url}")
                    print(f"[Tunnel/API] ⏳ 进入稳定性验证模式 (需要{_min_confirms_api}次通过)")
                    print(f"[Tunnel/API] 📧 邮件将在验证通过后自动发送")
                else:
                    print(f"[Tunnel/API] ⚠️ 隧道进程已启动但URL未就绪")
                    response_data['message'] = '隧道进程已启动，等待获取公网地址...'
                    response_data['status'] = 'waiting_for_url'
                
                sys.stdout.flush()
                return jsonify(response_data)
            else:
                error_msg = result.get('error', '启动失败')
                print(f"[Tunnel/API] ❌ 启动失败: {error_msg}")
                sys.stdout.flush()
                return jsonify({
                    'success': False,
                    'error': error_msg,
                    'status': 'failed'
                })

        last_url_invalid_log_time = 0  # 上次打印URL不可用日志的时间

        @app.get('/api/url-source/status')
        def url_source_status():
            """获取URL源状态和配置"""
            try:
                status = PathManager.get_url_source_status()
                
                # 执行健康检查（强制执行，忽略间隔）
                health = PathManager.check_url_files_health()
                
                return jsonify({
                    'success': True,
                    'config': status['config'],
                    'last_log': status['last_log'],
                    'health_check': health,
                    'timestamp': datetime.now().isoformat()
                })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': str(e)[:200]
                }), 500
        
        @app.post('/api/url-source/configure')
        async def url_source_configure():
            """配置URL获取策略"""
            try:
                data = await request.json()
                
                if not data:
                    return jsonify({
                        'success': False,
                        'error': '请提供配置数据'
                    }), 400
                
                # 允许的配置项
                allowed_keys = [
                    'primary_source',
                    'fallback_source', 
                    'enable_logging',
                    'enable_health_check',
                    'auto_sync_interval',
                    'validate_url',
                    'url_validation_timeout'
                ]
                
                # 过滤只允许的配置项
                config_data = {k: v for k, v in data.items() if k in allowed_keys}
                
                if not config_data:
                    return jsonify({
                        'success': False,
                        'error': '没有有效的配置项'
                    }), 400
                
                # 应用配置
                new_config = PathManager.configure_url_source(**config_data)
                
                return jsonify({
                    'success': True,
                    'message': '配置已更新',
                    'config': new_config
                })
            except Exception as e:
                logger.error(f'[tunnel_update_config] 更新失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({
                    'success': False,
                    'error': '隧道配置更新失败'
                }), 500

        @app.post('/api/url-source/health-check')
        def url_source_force_health_check():
            """强制执行健康检查"""
            try:
                # 重置健康检查时间戳，强制执行检查
                PathManager._url_health_check_time = 0
                
                health = PathManager.check_url_files_health()
                
                return jsonify({
                    'success': True,
                    'message': '健康检查已完成',
                    'health_result': health
                })
            except Exception as e:
                logger.error(f'[url_source_health_check] 健康检查失败: {type(e).__name__}: {e}', exc_info=True)
                return jsonify({
                    'success': False,
                    'error': 'URL健康检查失败'
                }), 500

        @app.get('/api/tunnel/status')
        def tunnel_status():
            global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_restart_count, tunnel_last_error, tunnel_last_heartbeat, tunnel_daemon_started, tunnel_restart_thread, tunnel_heartbeat_thread, tunnel_need_restart, last_url_invalid_log_time
            
            heartbeat_str = datetime.fromtimestamp(tunnel_last_heartbeat).strftime('%Y-%m-%d %H:%M:%S') if tunnel_last_heartbeat > 0 else None
            
            if tunnel_url:
                web_url = tunnel_url
            else:
                web_url = PathManager.get_public_url_from_web_log(skip_validation=True, quiet=True)
            
            hostc_process_running = Environment.check_process_running(Environment.HOSTC_PROCESS_NAME)
            cf_process_running = cf_process is not None and cf_process.poll() is None
            
            is_running = hostc_process_running and web_url is not None
            
            url_valid = (stable_url == web_url and 
                        stable_url_confirm_count >= stable_url_min_confirms and 
                        web_url is not None)
            
            stable_confirmed = (stable_url == web_url and 
                              stable_url_confirm_count >= stable_url_min_confirms and 
                              url_valid)
            
            verify_status = {
                'is_verifying': stable_url_confirm_count > 0 and stable_url_confirm_count < stable_url_min_confirms,
                'current_count': stable_url_confirm_count,
                'required_count': stable_url_min_confirms,
                'progress_percent': int((stable_url_confirm_count / stable_url_min_confirms) * 100) if stable_url_min_confirms > 0 else 0,
                'stable_url': stable_url,
                'time_elapsed_seconds': int(time.time() - url_first_seen_time) if url_first_seen_time > 0 and stable_url_confirm_count > 0 else 0,
                'estimated_remaining_seconds': max(0, (stable_url_min_confirms - stable_url_confirm_count) * 30) if stable_url_confirm_count < stable_url_min_confirms else 0
            }
            
            if stable_confirmed:
                detailed_status = 'stable'
                status_message = f'✅ 公网地址已稳定可用 (已连续验证{stable_url_confirm_count}次)'
            elif verify_status['is_verifying']:
                detailed_status = 'verifying'
                status_message = f'⏳ 正在验证稳定性 ({verify_status["current_count"]}/{verify_status["required_count"]})，预计还需{verify_status["estimated_remaining_seconds"]}秒'
            elif web_url and not url_valid:
                detailed_status = 'unstable'
                status_message = '⚠️ URL不可用，等待重新获取...'
            elif hostc_process_running and not web_url:
                detailed_status = 'starting'
                status_message = '🔄 隧道启动中，等待获取公网地址...'
            else:
                detailed_status = 'stopped'
                status_message = '⏹️ 隧道未运行'

            cf_stable_confirmed = (cf_stable_url == cf_url and 
                                   cf_stable_confirm_count >= cf_stable_min_confirms and 
                                   cf_url is not None)

            cf_status = {
                'running': cf_process_running,
                'url': cf_url,
                'mode': cf_mode,
                'stable': cf_stable_confirmed,
                'verify_count': cf_stable_confirm_count,
                'verify_required': cf_stable_min_confirms
            }

            start_tunnel_daemons()

            return {
                'running': is_running,
                'url': web_url,
                'url_valid': url_valid,
                'auto_restart': tunnel_auto_restart,
                'restart_count': tunnel_restart_count,
                'last_error': tunnel_last_error or ('URL无效，正在重启...' if tunnel_need_restart else None),
                'last_heartbeat': heartbeat_str,
                'tunnel_type': 'hostc',
                'detailed_status': detailed_status,
                'status_message': status_message,
                'stable_confirmed': stable_confirmed,
                'verify_progress': verify_status,
                'cloudflare': cf_status,
                'email_notification_status': {
                    'will_notify': not stable_confirmed and web_url is not None,
                    'notification_type': 'stable_available',
                    'enabled': email_notifier.get_email_config().get('enabled', False) if hasattr(email_notifier, 'get_email_config') else False,
                    'recipient': email_notifier.get_email_config().get('to_email', '') if hasattr(email_notifier, 'get_email_config') else '',
                    'sender': email_notifier.get_email_config().get('smtp_user', '') if hasattr(email_notifier, 'get_email_config') else '',
                    'sender_name': email_notifier.get_email_config().get('from_name', '公网IP监控') if hasattr(email_notifier, 'get_email_config') else '公网IP监控',
                    'condition': f'需要连续{stable_url_min_confirms}次验证通过',
                    'current_progress': f'{stable_url_confirm_count}/{stable_url_min_confirms}',
                    'last_stable_notification': datetime.fromtimestamp(last_stable_notification_time).strftime('%Y-%m-%d %H:%M:%S') if last_stable_notification_time > 0 else None,
                    'preview_subject': f'【✅ 公网地址已稳定可用】{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}' if not stable_confirmed and web_url else None,
                    'preview_body': f'''✅ 公网地址已稳定可用
时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
公网地址: {web_url or "待确定"}
✅ 稳定性验证：已连续通过 {stable_url_min_confirms} 次验证
📊 验证耗时：{int(time.time() - url_first_seen_time) if url_first_seen_time > 0 else 0} 秒
🎯 状态：确认稳定可用，可放心使用''' if not stable_confirmed and web_url else None
                }
            }

        @app.post('/api/tunnel/stop')
        def stop_tunnel():
            global tunnel_process, tunnel_url, tunnel_auto_restart, tunnel_need_restart, tunnel_restart_count, tunnel_last_error, tunnel_consecutive_failures
            global cf_process, cf_url, cf_mode
            tunnel_auto_restart = False
            tunnel_need_restart = False
            Environment.kill_process_by_name(Environment.HOSTC_PROCESS_NAME)
            if tunnel_process:
                try:
                    tunnel_process.terminate()
                    tunnel_process.wait(timeout=2)
                except Exception as e:
                    logger.debug(f"Nested exception: {e}")
                    try:
                        tunnel_process.kill()
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
            tunnel_process = None
            tunnel_url = None
            if cf_process:
                try:
                    cf_process.terminate()
                    cf_process.wait(timeout=2)
                except Exception as e:
                    logger.debug(f"Nested exception: {e}")
                    try:
                        cf_process.kill()
                    except Exception as e:
                        _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                        pass
            cf_process = None
            cf_url = None
            cf_mode = None
            return jsonify({'success': True, 'message': '所有隧道已停止'})

        # 初始化日志输出到 web_output.log
        setup_web_logging()
        
        # 启动时清理 temp 目录（超过3MB则清理）
        auto_clean_temp_dir()
        
        # 后台定期清理 temp 目录（每1分钟检查一次，超过3MB立即清理）
        def temp_cleanup_loop():
            while True:
                time.sleep(60)
                try:
                    auto_clean_temp_dir()
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
                    pass
        
        threading.Thread(target=temp_cleanup_loop, daemon=True).start()
        
        # 启动前获取一次局域网 IP 用于显示
        lan_ip_startup = None
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(2)
            s.connect((os.environ.get('LAN_IP_DETECT_HOST', '8.8.8.8'), int(os.environ.get('LAN_IP_DETECT_PORT', '80'))))
            lan_ip_startup = s.getsockname()[0]
        except Exception as e:
            _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
            pass
        finally:
            if s:
                try:
                    s.close()
                except Exception as e:
                    _module_logger.debug(f'静默异常: {type(e).__name__}: {e}', exc_info=True)
        
        print("=" * 50)
        print("Szwego商品爬虫 - Web服务")
        print("=" * 50)
        print(f"访问地址: http://localhost:{args.port}")
        print(f"局域网地址: http://{lan_ip_startup}:{args.port}" if lan_ip_startup else "")
        print("[Tunnel] 正在自动启动隧道...")
        
        tunnel_result = auto_start_tunnel()
        if tunnel_result['success']:
            url = tunnel_result.get('url')
            if url:
                print(f"[Tunnel] 隧道启动成功: {url}")
            else:
                print(f"[Tunnel] 隧道已就绪，公网地址将由心跳机制获取和验证")
        else:
            print(f"[Tunnel] 隧道启动失败: {tunnel_result.get('error', '未知错误')}")
        
        start_tunnel_daemons()
        
        print("按 Ctrl+C 停止服务")
        print("=" * 50)
        
        @app.get('/{invalid_path:path}')
        async def handle_invalid_path(invalid_path: str, request: Request):
            path = request.url.path
            if path.startswith('/dist/'):
                raise HTTPException(status_code=404, detail='File not found')
            if path.startswith('/api/') or path.startswith('/run') or path.startswith('/input') or path.startswith('/kill') or path.startswith('/output/'):
                raise HTTPException(status_code=404, detail=f'接口不存在: {path}')
            return await index()
        
        @app.get('/favicon.ico')
        async def favicon():
            return FileResponse(path=os.path.join(PROJECT_DIR, 'dist', 'favicon', 'favicon.ico'))


        print(f"\n🚀 FastAPI 服务启动中...")
        print(f"   地址: http://0.0.0.0:{args.port}")
        print(f"   文档: http://0.0.0.0:{args.port}/docs")
        uvicorn.run(app, host='0.0.0.0', port=args.port, log_level="info")
    else:
        main()

# SSRF Defense System Added




# ============================================================
# SSRF安全防御体系 (v3.8.89.20) - 基于SSRF攻击视频的完整防护
# 攻击场景: 内网探测、云元数据窃取、文件读取、协议走私、DNS Rebinding
# ============================================================





class SSRFProtection:
    BLOCKED_SCHEMES={'file','gopher','dict','ftp','sftp','ldap','tftp','netdoc','jar'}
    ALLOWED_SCHEMES={'http','https'}
    PRIVATE_IP_RANGES=[('10.0.0.0',8),('172.16.0.0',12),('192.168.0.0',16),('127.0.0.0',8),('169.254.0.0',16),('100.64.0.0',10),('198.18.0.0',15),('0.0.0.0',8)]
    CLOUD_METADATA_ENDPOINTS={'metadata.google.internal','169.254.169.254','metadata.azure.com'}
    SENSITIVE_PORTS={22,23,25,53,135,139,445,1433,1521,3306,3389,5432,5900,6379,8080,9200,27017}
    BLOCKED_HOSTNAMES={'localhost','localhost.localdomain','ip6-localhost','ip6-loopback','metadata'}

    def __init__(self,enabled=True,max_redirects=3,max_response_size=5*1024*1024,connect_timeout=5,read_timeout=10,dns_cache_ttl=300,block_private_ips=True,allow_localhost=False):
        self.enabled=enabled;self.max_redirects=max_redirects;self.max_response_size=max_response_size
        self.connect_timeout=connect_timeout;self.read_timeout=read_timeout;self.dns_cache_ttl=dns_cache_ttl
        self.block_private_ips=block_private_ips;self.allow_localhost=allow_localhost
        self._dns_cache={};self._dns_cache_time={};self._request_count=0;self._block_count=0
        self._log=logging.getLogger('SSRF')
        if not self._log.handlers:
            h=logging.StreamHandler();h.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
            self._log.addHandler(h);self._log.setLevel(logging.WARNING)

    def _log_event(self,t,u,d='',l='warning'):
        m=f'[{t}] {u}';getattr(self._log,l)(m+(f' | {d}' if d else ''))
        if t=='BLOCKED':self._block_count+=1

    def normalize_ip(self,ip):
        if not ip:return None
        try:
            ip=str(ip).strip().lower()
            if ':' in ip:return self._norm_ipv6(ip)
            return self._norm_ipv4(ip)
        except:return None

    def _norm_ipv4(self,ip):
        ip=ip.strip().strip('.')
        if not ip:return None
        if re.match(r'^\d+$',ip):
            v=int(ip)
            if 0<=v<=0xFFFFFFFF:return socket.inet_ntoa(struct.pack('!I',v&0xFFFFFFFF))
            return None
        parts=ip.split('.');np=[]
        for p in parts[:4]:
            p=p.strip()
            if not p:np.append('0');continue
            if p.startswith(('0x','0X')):v=int(p,16)
            elif len(p)>1 and p.startswith('0') and all(c in '01234567' for c in p):v=int(p,8)
            elif p.isdigit():v=int(p,10)
            else:return None
            if not(0<=v<=255):return None
            np.append(str(v))
        while len(np)<4:np.append('0')
        return '.'.join(np[:4])

    def is_private_ip(self,ip):
        if not ip:return True
        n=self.normalize_ip(ip)
        if not n:return True
        for net,pfx in self.PRIVATE_IP_RANGES:
            try:
                if ':' in n or ':' in net:continue
                i=struct.unpack('!I',socket.inet_aton(n))[0]
                s=struct.unpack('!I',socket.inet_aton(net))[0]
                m=(0xFFFFFFFF<<(32-pfx))&0xFFFFFFFF
                if(i&m)==(s&m):return True
            except:continue
        return False

    def is_blocked_hostname(self,h):
        if not h:return True
        h=h.lower().strip().rstrip('.')
        if h in self.BLOCKED_HOSTNAMES:return True
        for e in self.CLOUD_METADATA_ENDPOINTS:
            if e in h or h.endswith(e.replace('169.254.169.254','')):return True
        if h.startswith('metadata.') or h.endswith('.internal'):return True
        return False

    def is_safe_url(self,url):
        if not self.enabled:return True,None
        if not url or not isinstance(url,str):return False,'Invalid URL'
        url=url.strip()
        if len(url)>8192:return False,'URL too long'
        try:

            p=urllib.parse.urlparse(url)
            if not p.scheme or not p.netloc:return False,'Invalid format'
            s=p.scheme.lower()
            if s in self.BLOCKED_SCHEMES:return False,f'Blocked: {s}'
            if s and s not in self.ALLOWED_SCHEMES:return False,f'Unsupported: {s}'
            h=p.hostname
            if not h:return False,'Missing host'
            if self.is_blocked_hostname(h):return False,f'Blocked host: {h}'
            pt=p.port
            if pt and pt in self.SENSITIVE_PORTS:return False,f'Sensitive port: {pt}'
            return True,None
        except Exception as e:return False,f'Error: {e}'

    def safe_request(self,url,method='GET',headers=None,data=None,timeout=None,**kw):
        if timeout is None:timeout=self.connect_timeout
        safe,err=self.is_safe_url(url)
        if not safe:self._log_event('BLOCKED',url,err,'error');return None,f'SSRF Blocked: {err}'
        try:

            req=urllib.request.Request(url,method=method.upper(),data=data,**kw)
            hd={'User-Agent':'SecureClient/2.0','Accept':'*/*'}
            if headers and isinstance(headers,dict):hd.update(headers)
            for d in ['Host','Origin','Referer','Cookie','Authorization']:hd.pop(d,None)
            for k,v in hd.items():
                if k!='Content-Length':req.add_header(k,v)
            ctx=ssl.create_default_context();ctx.check_hostname=True;ctx.verify_mode=ssl.CERT_REQUIRED
            resp=urllib.request.urlopen(req,timeout=min(timeout,self.read_timeout),context=ctx)
            cl=resp.headers.get('Content-Length')
            if cl and int(cl)>self.max_response_size:resp.close();return None,f'Too large: {int(cl)}'
            self._request_count+=1
            return resp,None
        except urllib.error.HTTPError as e:return None,f'HTTP [{e.code}]'
        except urllib.error.URLError as e:return None,f'Failed: {e.reason}'
        except socket.timeout:return None,'Timeout'
        except ssl.SSLError as e:return None,f'SSL: {e}'
        except Exception as e:return None,f'Err: {type(e).__name__}'

    def get_stats(self):
        return {'enabled':self.enabled,'requests':self._request_count,'blocked':self._block_count}

def ssrf_safe_request(func):
    p=SSRFProtection()
    @wraps(func)
    def wrapper(*args,**kwargs):
        url=args[0] if args else kwargs.get('url','')
        if url:
            s,e=p.is_safe_url(str(url))
            if not s:raise ValueError(f'Blocked: {e}')
        return func(*args,**kwargs)
    return wrapper

ssrf_protector=SSRFProtection(enabled=True,block_private_ips=True,allow_localhost=False)

print('='*60)
print('OK - SSRF Defense System v3.8.89.20 Loaded Successfully!')
print('='*60)

# ============================================================
# 安全检查与审计系统 (v3.8.89.31) - 整合quick_security_check + security_audit + config_secure
# 包含: Playwright移动端安全检查、依赖审计、配置加密管理
# ============================================================

_SECURITY_CHECKS=None

def _get_security_checks():
    global _SECURITY_CHECKS
    if _SECURITY_CHECKS is not None:
        return _SECURITY_CHECKS
    _main_file=PROJECT_DIR/'main.py'
    _app_js=PROJECT_DIR/'dist'/'app.js'
    _gitignore=PROJECT_DIR/'.gitignore'
    _SECURITY_CHECKS={
        'injection':[
            {'name':'SQL注入防护','description':'检查是否存在不安全的SQL操作','files':[_main_file],'patterns':[
                (r'cursor\.execute\s*\([^)]*%\s*[sd]',False,'字符串格式化SQL查询'),
                (r'cursor\.execute\s*\([^)]*\+\s*["\']',False,'字符串拼接SQL'),
                (rf'f[\"\'].*SELECT.*FROM.*\{{\{{',False,'f-string SQL注入'),
                (r'pymysql\.connect|sqlite3\.connect',None,'数据库连接（需确认参数化）'),
            ]},
            {'name':'命令注入防护','description':'检查subprocess/os.system调用安全性','files':[_main_file],'patterns':[
                (r'os\.system\s*\(',False,'os.system()危险调用'),
                (r'subprocess\.(call|run|Popen)\s*\([^)]*shell\s*=\s*True',False,'shell=True危险模式'),
                (r'subprocess\.(call|run|Popen)\s*\([^)]*\+\s*["\']',False,'命令拼接'),
                (r'subprocess\.(call|run|Popen)\s*\(\s*\[',True,'安全的列表参数传递'),
                (r'shell\s*=\s*False',True,'禁用shell解释'),
                (r'dangerous_patterns|BLACKLISTED_COMMANDS',True,'存在命令黑名单检测'),
            ]},
            {'name':'XSS防护','description':'检查前端XSS防护机制','files':[_app_js],'patterns':[
                (r'function\s+escapeHtml',True,'HTML转义函数定义'),
                (r'innerHTML\s*=.*escapeHtml',True,'innerHTML使用转义'),
                (r'innerHTML\s*=\s*[^;]*[^e]scapeHtml',False,'未转义的innerHTML赋值'),
                (r'\.html\s*\([^)]*[^e]scapeHtml',False,'jQuery.html()未转义'),
                (r'document\.write\s*\(',False,'document.write()危险调用'),
                (r'eval\s*\(',False,'eval()危险函数'),
            ]},
        ],
        'deserialization':[
            {'name':'反序列化安全','description':'检查反序列化操作的安全性','files':[_main_file],'patterns':[
                (r'pickle\.loads?\s*\(',False,'pickle反序列化（代码执行风险）'),
                (r'yaml\.load\s*\((?!.*SafeLoader)',False,'yaml.load不安全加载器'),
                (r'marshal\.loads?\s*\(',False,'marshal反序列化（代码执行风险）'),
                (r'eval\s*\(',False,'eval()代码执行'),
                (r'exec\s*\(',False,'exec()代码执行'),
                (r'json\.loads?\s*\(',True,'JSON安全反序列化'),
                (r'yaml\.safe_load',True,'YAML安全加载器'),
            ]},
        ],
        'sensitive_data':[
            {'name':'硬编码凭证检测','description':'检查是否有硬编码的密码/密钥/Token','files':[_main_file],'patterns':[
                (r'password\s*=\s*[\'"][^\'"]{8,}[\'"]',False,'硬编码密码'),
                (r'secret_key\s*=\s*[\'"][^\'"]+?[\'"]',False,'硬编码密钥'),
                (r'api_key\s*=\s*[\'"][^\'"]+?[\'"]',False,'硬编码API Key'),
                (r'token\s*=\s*[\'"][^\'"]+?[\'"]',False,'硬编码Token'),
                (r'os\.environ\.get\s*\([\'"](?:PASSWORD|SECRET|KEY|TOKEN)',True,'环境变量读取敏感信息'),
                (r'config\[.*(?:password|secret|token)',None,'配置文件读取（需确认加密）'),
            ]},
            {'name':'日志信息泄露','description':'检查日志中是否泄露敏感信息','files':[_main_file],'patterns':[
                (r'(?:logger|print)\(.*(?:password|token|secret).*f[\"\']',False,'日志打印密码/token'),
                (r'print\(.*Token.*\{.*\}',False,'print输出Token信息'),
                (r'logger\.(?:debug|info)\(.*(?:password|token)',False,'日志记录敏感字段'),
                (r'except.*:\s*(?:return|raise).*str\(e\)',False,'异常信息直接返回给客户端'),
                (r'logger\.error\(.*str\(e\)',None,'异常日志（需确认无敏感数据）'),
            ]},
        ],
        'access_control':[
            {'name':'路径遍历防护','description':'检查文件路径操作的安全性','files':[_main_file],'patterns':[
                (r'def\s+sec_sp',True,'路径安全函数定义'),
                (r'os\.path\.realpath',True,'路径规范化处理'),
                (r'startswith.*base_dir',True,'前缀匹配检查'),
                (r'open\s*\([^)]*\+\s*(?:user|request)\.',False,'用户输入直接拼接文件路径'),
                (r'os\.path\.join\s*\([^)]*\.\.[\/\\]',False,'路径包含..（可能遍历）'),
            ]},
            {'name':'认证与授权','description':'检查API访问控制机制','files':[_main_file],'patterns':[
                (r'secrets\.compare_digest',True,'时间安全比较'),
                (r'@app\.(post|put|delete).*async def',True,'写操作端点定义'),
                (r'x-api-key',True,'API Key头检查'),
                (r'CSRF_EXEMPT_PATHS|csrf_token',True,'CSRF防护机制'),
                (r'request\.client\.host',True,'IP地址获取（用于白名单）'),
                (r'LOCAL_TRUSTED_ORIGINS',True,'可信来源白名单'),
            ]},
            {'name':'SSRF防护','description':'检查服务端请求伪造防护','files':[_main_file],'patterns':[
                (r'class\s+SecureClient',True,'安全HTTP客户端类'),
                (r'PRIVATE_IP_RANGES',True,'私有IP黑名单'),
                (r'is_safe_url',True,'URL安全检查函数'),
                (r'SENSITIVE_PORTS',True,'敏感端口过滤'),
                (r'BLOCKED_HOSTNAMES',True,'阻止的主机名列表'),
                (r'requests\.get\s*\([^)]*url\s*=\s*(?:user|request)',False,'用户输入URL直接请求'),
            ]},
        ],
        'configuration':[
            {'name':'CORS配置安全','description':'检查跨域资源共享配置','files':[_main_file],'patterns':[
                (r'allow_origins\s*=\s*\[\s*\*',False,'CORS允许所有来源（危险）'),
                (r'allow_origins\s*=\s*\[',True,'CORS源白名单配置'),
                (r'allow_credentials\s*=\s*True',None,'允许凭据（需配合严格白名单）'),
                (r'CORSMiddleware',True,'CORS中间件启用'),
            ]},
            {'name':'DEBUG模式控制','description':'检查调试模式设置','files':[_main_file],'patterns':[
                (r'DEBUG\s*=\s*True',False,'生产环境开启DEBUG'),
                (r'app\.debug\s*=\s*True',False,'应用调试模式开启'),
                (r'logging\.setLevel\s*\(.*DEBUG',None,'日志级别DEBUG（开发用）'),
                (r'logger\s*=\s*logging\.getLogger',True,'结构化日志使用'),
            ]},
            {'name':'.gitignore完整性','description':'检查敏感文件是否被忽略','files':[_gitignore],'patterns':[
                (r'config/.*\.json',True,'配置JSON文件已忽略'),
                (r'\.env',True,'环境变量文件已忽略'),
                (r'\.encryption_key',True,'加密密钥文件已忽略'),
                (r'__pycache__/',True,'Python缓存已忽略'),
                (r'\.pyc$',True,'编译文件已忽略'),
            ]},
        ],
        'cryptography':[
            {'name':'密码学实践','description':'检查随机数生成和密钥管理','files':[_main_file],'patterns':[
                (r'secrets\.token_urlsafe',True,'安全令牌生成'),
                (r'secrets\.choice',True,'加密安全随机选择'),
                (r'secrets\.compare_digest',True,'时间安全比较'),
                (r'random\.randint|random\.choice|random\.random',False,'不安全随机数'),
                (r'hashlib\.md5(?!\()',None,'MD5哈希（仅用于非安全场景）'),
                (r'PBKDF2HMAC|bcrypt|argon2',True,'安全密钥派生'),
            ]},
        ],
        'playwright_mobile':[
            {'name':'浏览器上下文隔离','description':'检查Playwright浏览器会话隔离，防止Cookie/LocalStorage泄露','files':[_main_file],'patterns':[
                (r'browser\.new_context\(\)',True,'独立浏览器上下文创建'),
                (r'context\.close\(\)',True,'上下文关闭清理'),
                (r'await\s+browser\.close\(\)',True,'浏览器实例关闭'),
                (r'browser_context.*storage_state',None,'存储状态管理（需确认隔离）'),
                (r'context\.add_cookies',None,'Cookie操作（需确认作用域）'),
            ]},
            {'name':'动态内容操作安全','description':'检查自动化点击操作是否有防误触机制','files':[_main_file],'patterns':[
                (r'page\.click\(',None,'页面点击操作（需确认目标验证）'),
                (r'page\.wait_for_selector',True,'元素等待机制'),
                (r'page\.url|page\.title\(\)',True,'页面状态验证'),
                (r'\.is_visible\(\)|\.is_enabled\(\)',True,'元素可见性检查'),
                (r'confirm_dialog|accept_dialog',None,'弹窗处理（需二次确认）'),
            ]},
            {'name':'文件下载上传安全','description':'检查文件操作的类型校验和路径限制','files':[_main_file],'patterns':[
                (r'download\.(save_as|path)',None,'文件下载（需限制目录）'),
                (r'set_input_files\(',None,'文件上传（需MIME检测）'),
                (r'filetype\.|mime_type|MIMEType',True,'文件类型检测'),
                (r'\.download_path|download_dir',True,'下载目录限制'),
                (r'allow_files|accept',None,'文件白名单过滤'),
            ]},
            {'name':'浏览器指纹与反检测','description':'检查User-Agent和浏览器特征伪装安全性','files':[_main_file],'patterns':[
                (r'user_agent.*Mozilla',True,'User-Agent设置'),
                (r'viewport.*width.*height',True,'视口配置'),
                (r'--disable-blink-features=AutomationControlled',True,'反自动化检测'),
                (r'--no-sandbox',None,'沙箱禁用（开发环境可接受）'),
                (r'locale|timezone_id',None,'地区设置（需一致性）'),
            ]},
            {'name':'移动端环境安全','description':'检查移动端模拟器和设备参数安全性','files':[_main_file],'patterns':[
                (r'device_scale_factor|is_mobile|has_touch',True,'移动设备模拟参数'),
                (r'user_agent.*Mobile|iPhone|Android',True,'移动端UA标识'),
                (r'viewport.*\{.*width:\s*375',None,'iPhone视口配置'),
                (r'viewport.*\{.*width:\s*360',None,'Android视口配置'),
                (r'geolocation|latitude|longitude',None,'GPS位置模拟（需服务端校验）'),
            ]},
            {'name':'网络流量安全','description':'检查HTTPS强制和代理防护','files':[_main_file],'patterns':[
                (r'ignore_https_errors\s*=\s*True',False,'忽略SSL证书错误（危险）'),
                (r'--proxy-server|--proxy-bypass-list',None,'代理配置（测试环境谨慎）'),
                (r'route\.continue\(|route\.fetch\(\)',None,'路由拦截（可能篡改流量）'),
                (r'context\.set_extra_http_headers',True,'自定义请求头'),
                (r'Accept-Encoding|Accept-Language',None,'请求头完整性'),
            ]},
            {'name':'截图与快照安全','description':'检查页面截图是否泄露敏感信息','files':[_main_file],'patterns':[
                (r'page\.screenshot\(',None,'页面截图（避免包含敏感数据）'),
                (r'page\.pdf\(\)',None,'PDF生成（可能包含隐私信息）'),
                (r'screenshot\.png|screenshot\.jpg',None,'截图文件存储（需加密或限时删除）'),
                (r'full_page\s*=\s*True',None,'全页截图（风险更高）'),
                (r'clip.*\{.*x:.*y:',True,'区域截图（更安全）'),
            ]},
            {'name':'Playwright进程安全','description':'检查浏览器子进程管理和资源清理','files':[_main_file],'patterns':[
                (r'async with async_playwright',True,'上下文管理器使用'),
                (r'try.*finally.*browser\.close',True,'异常时资源清理'),
                (r'taskkill.*chrome|taskkill.*chromium|pkill.*chrome',True,'残留进程清理'),
                (r'executable_path.*chrome',None,'自定义浏览器路径（需验证完整性）'),
                (r'headless\s*=\s*(?:False|0)',None,'有头模式（生产环境建议无头）'),
            ]},
        ],
    }
    return _SECURITY_CHECKS

_CATEGORY_DISPLAY_NAMES={
    'injection':'一、注入攻击防护',
    'deserialization':'二、反序列化漏洞防护',
    'sensitive_data':'三、敏感信息泄露防护',
    'access_control':'四、权限控制失效防护',
    'configuration':'五、依赖与配置安全',
    'cryptography':'六、密码学安全实践',
    'playwright_mobile':'七、Playwright + 移动端安全',
}

class SecurityChecker:
    """安全检查器 - 整合原quick_security_check.py"""

    def __init__(self,verbose=False):
        self.verbose=verbose
        self.results={}
        self.total_checks=0
        self.passed_checks=0
        self.failed_checks=0

    def _check_file(self,file_path):
        if not os.path.exists(file_path):return ""
        try:
            with open(file_path,'r',encoding='utf-8',errors='ignore') as f:return f.read()
        except Exception:return ""

    def _run_pattern_check(self,content,pattern,expected,description):
        matches=re.findall(pattern,content,re.IGNORECASE|re.MULTILINE)
        found=len(matches)>0
        if expected is None:
            message=f"{description} - {'发现' if found else '未发现'}"
            return True,message
        elif expected and found:
            return True,f"✅ {description}"
        elif not expected and not found:
            return True,f"✅ {description} (未检测到)"
        elif expected and not found:
            return False,f"❌ 缺少: {description}"
        else:
            return False,f"❌ 发现安全问题: {description}"

    def run_all_checks(self):
        checks=_get_security_checks()
        categories_passed={}
        for category_name,category_checks in checks.items():
            display_name=_CATEGORY_DISPLAY_NAMES.get(category_name,category_name)
            category_ok=True
            for check in category_checks:
                for file_path in check['files']:
                    content=self._check_file(file_path)
                    if not content:continue
                    for pattern,expected,desc in check['patterns']:
                        self.total_checks+=1
                        passed,_=self._run_pattern_check(content,pattern,expected,desc)
                        if passed:
                            self.passed_checks+=1
                        else:
                            self.failed_checks+=1
                            category_ok=False
            categories_passed[category_name]=category_ok
        score=(self.passed_checks/self.total_checks*100) if self.total_checks>0 else 0
        return {
            'timestamp':datetime.now().isoformat(),
            'total':self.total_checks,
            'passed':self.passed_checks,
            'failed':self.failed_checks,
            'score':round(score,1),
            'categories':categories_passed,
        }

_KNOWN_VULNERABLE_VERSIONS={
    'requests':{'<2.31.0':'CVE-2023-32681 (Header注入)'},
    'urllib3':{'<2.0.7':'CVE-2023-45803 (Cookie注入)'},
    'pillow':{'<10.1.0':'CVE-2023-44471 (缓冲区溢出)'},
    'pyyaml':{'<6.0.1':'CVE-2024-34351 (YAML反序列化)'},
}

_SECURE_MINIMUM_VERSIONS={
    'fastapi':'0.100.0','uvicorn':'0.23.0','pydantic':'2.0.0',
    'playwright':'1.45.0','pymysql':'1.1.0','psutil':'5.9.0',
    'cryptography':'42.0.0','pillow':'10.0.0','requests':'2.32.0',
}

class DependencyAuditor:
    """依赖安全审计器 - 整合原security_audit.py"""

    def __init__(self):
        self.installed_packages={}
        self.vulnerabilities=[]

    def get_installed_packages(self):
        try:
            result=subprocess.run([sys.executable,'-m','pip','list','--format=json'],capture_output=True,text=True,timeout=30)
            packages=json.loads(result.stdout)
            return {pkg['name'].lower():pkg['version'] for pkg in packages}
        except Exception:return {}

    def parse_requirements(self):
        req_file=PROJECT_DIR/'requirements.txt'
        if not os.path.exists(req_file):return []
        requirements=[]
        with open(req_file,'r',encoding='utf-8') as f:
            for line in f:
                line=line.strip()
                if line and not line.startswith('#'):
                    parts=line.split('>=') if '>=' in line else line.split('==')
                    if len(parts)==2:
                        requirements.append((parts[0].strip().lower(),parts[1].strip()))
        return requirements

    def audit(self):
        self.installed_packages=self.get_installed_packages()
        required_packages=self.parse_requirements()
        total_issues=0
        for package,req_ver in required_packages:
            inst_ver=self.installed_packages.get(package.lower())
            if inst_ver and package.lower() in _KNOWN_VULNERABLE_VERSIONS:
                for vuln_ver,desc in _KNOWN_VULNERABLE_VERSIONS[package.lower()].items():
                    min_ver=vuln_ver.lstrip('<').strip()
                    if self._ver_lt(inst_ver,min_ver):
                        self.vulnerabilities.append({'package':package,'installed':inst_ver,'issue':desc})
                        total_issues+=1
        return {
            'total_packages':len(required_packages),
            'installed_count':len(self.installed_packages),
            'vulnerabilities_found':total_issues,
            'vulnerabilities':self.vulnerabilities,
        }

    def _ver_lt(self,v1,v2):
        if packaging_version is not None:
            try:
                return packaging_version.parse(v1)<packaging_version.parse(v2)
            except Exception:pass
        p1=[int(x) for x in v1.split('.') if x.isdigit()]
        p2=[int(x) for x in v2.split('.') if x.isdigit()]
        for i in range(max(len(p1),len(p2))):
            a=p1[i] if i<len(p1) else 0
            b=p2[i] if i<len(p2) else 0
            if a<b:return True
            elif a>b:return False
        return False

_SENSITIVE_CONFIG_FIELDS=['login.password','headers.cookie','email.smtp_password']

class SecureConfigManager:
    """安全配置管理器 - 整合原config_secure_template.py"""

    _instance=None
    _fernet=None

    def __new__(cls):
        if cls._instance is None:
            cls._instance=super().__new__(cls)
            cls._instance._initialize()
        return cls._instance

    def _initialize(self):
        key_file=PROJECT_DIR/'config'/'.encryption_key'
        salt_file=PROJECT_DIR/'config'/'.salt'
        if os.path.exists(key_file) and os.path.exists(salt_file):
            key=self._load_key(key_file,salt_file)
            if key:
                if Fernet is not None:
                    try:
                        self._fernet=Fernet(key)
                    except Exception:pass

    def _load_key(self,key_file,salt_file):
        env_key=os.environ.get('CONFIG_ENCRYPTION_KEY')
        if env_key:
            if hashes is not None and PBKDF2HMAC is not None:
                try:
                    with open(salt_file,'rb') as f:salt=f.read()
                    kdf=PBKDF2HMAC(algorithm=hashes.SHA256(),length=32,salt=salt,iterations=480000)
                    return base64.urlsafe_b64encode(kdf.derive(env_key.encode()))
                except Exception:return None
        if os.path.exists(key_file):
            with open(key_file,'rb') as f:return f.read()
        return None

    def encrypt_value(self,value):
        if not self._fernet:raise RuntimeError("加密器未初始化")
        encrypted=self._fernet.encrypt(value.encode())
        return f"ENC({encrypted.decode()})"

    def decrypt_value(self,encrypted_value):
        if not encrypted_value.startswith("ENC("):return encrypted_value
        if not self._fernet:raise RuntimeError("加密器未初始化")
        try:
            encrypted=encrypted_value[4:-1]
            return self._fernet.decrypt(encrypted.encode()).decode()
        except Exception:return ""

    def load_config(self):
        config_file=PROJECT_DIR/'config'/'config.json'
        if not os.path.exists(config_file):return {}
        with open(config_file,'r',encoding='utf-8') as f:config=json.load(f)
        return self._decrypt_sensitive(config)

    def save_config(self,config):
        config=self._encrypt_sensitive(config)
        config_file=PROJECT_DIR/'config'/'config.json'
        os.makedirs(os.path.dirname(config_file),exist_ok=True)
        with open(config_file,'w',encoding='utf-8') as f:
            json.dump(config,f,ensure_ascii=False,indent=2)

    def _get_nested(self,obj,path):
        keys=path.split('.')
        current=obj
        for key in keys:
            if isinstance(current,dict) and key in current:current=current[key]
            else:return None
        return current

    def _set_nested(self,obj,path,value):
        keys=path.split('.')
        current=obj
        for key in keys[:-1]:
            if key not in current or not isinstance(current[key],dict):current[key]={}
            current=current[key]
        current[keys[-1]]=value

    def _encrypt_sensitive(self,config):
        for field_path in _SENSITIVE_CONFIG_FIELDS:
            value=self._get_nested(config,field_path)
            if value and not str(value).startswith("ENC("):
                self._set_nested(config,field_path,self.encrypt_value(str(value)))
        return config

    def _decrypt_sensitive(self,config):
        for field_path in _SENSITIVE_CONFIG_FIELDS:
            value=self._get_nested(config,field_path)
            if value and str(value).startswith("ENC("):
                self._set_nested(config,field_path,self.decrypt_value(str(value)))
        return config

    @classmethod
    def initialize_encryption(cls,password=None):
        """初始化加密系统"""
        key_file=PROJECT_DIR/'config'/'.encryption_key'
        salt_file=PROJECT_DIR/'config'/'.salt'
        if os.path.exists(key_file):
            return True,'加密系统已初始化'
        if not password or len(password)<8:
            return False,'密码长度至少需要8个字符'
        if Fernet is None or hashes is None or PBKDF2HMAC is None:
            return False,'cryptography库未安装'
        try:
            salt=os.urandom(16)
            kdf=PBKDF2HMAC(algorithm=hashes.SHA256(),length=32,salt=salt,iterations=480000)
            key=base64.urlsafe_b64encode(kdf.derive(password.encode()))
            os.makedirs(os.path.dirname(key_file),exist_ok=True)
            with open(key_file,'wb') as f:f.write(key)
            with open(salt_file,'wb') as f:f.write(salt)
            config_file=PROJECT_DIR/'config'/'config.json'
            if os.path.exists(config_file):
                mgr=cls()
                if mgr._fernet is None:
                    mgr._fernet=Fernet(key)
                with open(config_file,'r',encoding='utf-8') as f:config=json.load(f)
                config=mgr._encrypt_sensitive(config)
                with open(config_file,'w',encoding='utf-8') as f:
                    json.dump(config,f,ensure_ascii=False,indent=2)
            return True,'加密系统初始化成功'
        except Exception as e:
            return False,f'初始化失败: {e}'

print('='*60)
print('OK - Security Check & Audit System v3.8.89.31 Loaded!')
print('='*60)


def _auto_encrypt_config():
    """自动加密config.json中的明文敏感字段（首次运行或加密未激活时）"""
    try:
        config_file = Path(PROJECT_DIR) / 'config' / 'config.json'
        if not config_file.exists():
            return
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        has_plaintext = False
        for field_path in _SENSITIVE_CONFIG_FIELDS:
            keys = field_path.split('.')
            current = config
            for key in keys:
                if isinstance(current, dict) and key in current:
                    current = current[key]
                else:
                    current = None
                    break
            if current and isinstance(current, str) and not current.startswith('ENC('):
                has_plaintext = True
                break
        if not has_plaintext:
            return
        mgr = SecureConfigManager()
        if mgr._fernet is None:
            password = os.environ.get('CONFIG_ENCRYPTION_KEY', '')
            if not password:
                password = secrets.token_urlsafe(32)
                os.environ['CONFIG_ENCRYPTION_KEY'] = password
            success, _ = SecureConfigManager.initialize_encryption(password)
            if not success:
                return
            mgr = SecureConfigManager()
        if mgr._fernet is not None:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            config = mgr._encrypt_sensitive(config)
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            print('[SecureConfig] ✅ 已自动加密config.json中的敏感字段')
    except Exception as e:
        print(f'[SecureConfig] ⚠️ 自动加密失败: {e}')

_auto_encrypt_config()